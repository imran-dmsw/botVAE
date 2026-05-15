from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

AUDIT_COLUMNS = (
    ("BASE_REFERENCE_MODEL", "BasePrice_RefYear"),
    ("BASE_REFERENCE_MODEL", "Units_RefYear"),
    ("BASE_REFERENCE_MODEL", "Range"),
    ("BASE_REFERENCE_MODEL", "Segment"),
    ("FIRMS", "Share_RefYear"),
    ("FIRMS", "Units_RefYear"),
)


@dataclass(frozen=True)
class ExcelFieldProbe:
    sheet: str
    cell: str
    field: str
    raw_value: Any
    validated_value: Any
    used_value: Any


@dataclass(frozen=True)
class ExcelWorkbookAudit:
    workbook_path: str
    uses_validated_excel: bool
    source_label: str
    probes: tuple[ExcelFieldProbe, ...]


def _header_map(ws) -> dict[str, int]:
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def _cell_ref(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(col)}{row}"


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _coerce_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _pick_used_value(validated: Any, raw: Any) -> Any:
    if validated is not None:
        return validated
    if _coerce_number(raw) is not None:
        return raw
    return validated


def _append_probe(
    probes: list[ExcelFieldProbe],
    *,
    sheet: str,
    row: int,
    col: int,
    field: str,
    ws_formula,
    ws_values,
) -> None:
    raw = ws_formula.cell(row, col).value
    validated = ws_values.cell(row, col).value
    probes.append(
        ExcelFieldProbe(
            sheet=sheet,
            cell=_cell_ref(row, col),
            field=field,
            raw_value=raw,
            validated_value=validated,
            used_value=_pick_used_value(validated, raw),
        )
    )


def _iter_brm_rows(ws, firm: str) -> Iterable[tuple[int, dict[str, int]]]:
    headers = _header_map(ws)
    firm_col = headers.get("Firm")
    if not firm_col:
        return
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, firm_col).value != firm:
            continue
        yield row, headers


def audit_workbook_fields(workbook_path: Path | str, firm: str) -> ExcelWorkbookAudit:
    path = Path(workbook_path).expanduser()
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formula = openpyxl.load_workbook(path, data_only=False)

    probes: list[ExcelFieldProbe] = []
    brm = wb_values["BASE_REFERENCE_MODEL"]
    brm_f = wb_formula["BASE_REFERENCE_MODEL"]
    for row, headers in _iter_brm_rows(brm, firm):
        for field in ("BasePrice_RefYear", "Units_RefYear", "Range", "Segment"):
            col = headers.get(field)
            if not col:
                continue
            _append_probe(
                probes,
                sheet="BASE_REFERENCE_MODEL",
                row=row,
                col=col,
                field=field,
                ws_formula=brm_f,
                ws_values=brm,
            )

    firms_ws = wb_values["FIRMS"]
    firms_f = wb_formula["FIRMS"]
    for row in range(2, firms_ws.max_row + 1):
        if firms_ws.cell(row, 1).value != firm:
            continue
        for col, field in ((3, "Units_RefYear"), (4, "Share_RefYear")):
            _append_probe(
                probes,
                sheet="FIRMS",
                row=row,
                col=col,
                field=field,
                ws_formula=firms_f,
                ws_values=firms_ws,
            )

    uses_validated = True
    for probe in probes:
        if _is_formula(probe.raw_value) and probe.validated_value is None:
            uses_validated = False
            break
        if probe.used_value is None:
            uses_validated = False
            break

    label = (
        "Source des données : valeurs validées Excel"
        if uses_validated
        else "Source des données : valeurs brutes non validées"
    )
    return ExcelWorkbookAudit(
        workbook_path=str(path),
        uses_validated_excel=uses_validated,
        source_label=label,
        probes=tuple(probes),
    )
