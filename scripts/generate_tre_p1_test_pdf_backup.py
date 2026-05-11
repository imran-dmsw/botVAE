"""
Genere un rapport PDF pour le test bout en bout TRE P1 (simulation VAE).
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import shutil

from fpdf import FPDF

import openpyxl

# Memes chemins que run_tre_p1_e2e_test.py
from run_tre_p1_e2e_test import (
    MARKETING_TRE,
    PRODUCTIONS,
    OUT as TEST_XLSX,
    WORKBOOK_DEFAULT,
    budget_base_adjusted_tre_p1,
    evaluate_controles_tre_p1_python,
    inject_valid_decisions,
    price_list_p040_est,
    seuils_marketing,
    tre_revenue_ref_year,
)

# Fallback si PARAM absent (documentation / alignement script utilisateur)
MARKETING_REF = {
    "AVE": 80_000,
    "CAN": 80_000,
    "EBI": 80_000,
    "GIA": 80_000,
    "PED": 80_000,
    "RID": 80_000,
    "SUR": 80_000,
    "TRE": 120_000,
    "VEL": 80_000,
}
MARKETING_MIN_ABS = 50_000
MARKETING_MIN_PCT = 0.03


def calcul_seuils_marketing(firm: str, period: int, total_marketing_precedent: float | None = None):
    """Meme logique que PARAM / CONTROLE_DECISIONS (sans classeur)."""
    ref = MARKETING_REF.get(firm, 80_000)
    if period == 1:
        base = ref
    else:
        base = total_marketing_precedent if total_marketing_precedent else ref
    min_marketing = max(MARKETING_MIN_ABS, MARKETING_MIN_PCT * base)
    max_marketing = base * 1.10
    return min_marketing, max_marketing

PDF_OUT = Path("/Users/imran/BotMarketing/reports/Rapport_test_TRE_P1_VAE.pdf")
PDF_DOWNLOADS = Path("/Users/imran/Downloads/Rapport_test_TRE_P1_VAE.pdf")

# Police Unicode (macOS)
FONT_PATHS = [
    Path("/Library/Fonts/Arial Unicode.ttf"),
    Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
]


class RapportPDF(FPDF):
    font_main = "Helvetica"

    def header(self):
        self.set_font(getattr(self, "font_main", "Helvetica"), "", 9)
        self.set_text_color(80, 80, 80)
        self.cell(0, 8, "Simulation marche VAE — Test TRE periode 1", align="R")
        self.ln(12)

    def footer(self):
        self.set_y(-15)
        self.set_font(getattr(self, "font_main", "Helvetica"), "", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


def setup_font(pdf: FPDF) -> str:
    for fp in FONT_PATHS:
        if fp.exists():
            pdf.add_font("UFont", "", str(fp))
            return "UFont"
    return "Helvetica"


def set_section_title(pdf: FPDF, ff: str, size: int = 12):
    if ff == "UFont":
        pdf.set_font(ff, "", size + 1)
    else:
        pdf.set_font(ff, "B", size)


def safe_txt(s: str) -> str:
    s = s.replace("\u26a0", "!").replace("⚠", "!")
    if _u_font_ok:
        return s
    return (
        s.replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
        .replace("î", "i")
        .replace("ç", "c")
        .replace("É", "E")
        .replace("€", "EUR")
        .replace("✓", "OK")
        .replace("✗", "X")
    )


_u_font_ok = False


def main():
    global _u_font_ok
    wb_path = TEST_XLSX if TEST_XLSX.exists() else WORKBOOK_DEFAULT
    wb = openpyxl.load_workbook(wb_path)
    inject_valid_decisions(wb)
    rev = tre_revenue_ref_year(wb)
    bba = budget_base_adjusted_tre_p1(wb)
    rd = int(round(0.02 * bba))
    plist = price_list_p040_est(wb)
    mkt_min_ref, mkt_max_ref = seuils_marketing(wb, "TRE", 1)
    m_min_naive, m_max_naive = calcul_seuils_marketing("TRE", 1)
    report, n_blocked = evaluate_controles_tre_p1_python(wb)

    pdf = RapportPDF()
    pdf.alias_nb_pages()
    ff = setup_font(pdf)
    pdf.font_main = ff
    _u_font_ok = ff == "UFont"

    pdf.add_page()
    set_section_title(pdf, ff, 15)
    pdf.set_text_color(28, 43, 74)
    pdf.multi_cell(0, 10, safe_txt("Rapport — Test bout en bout TRE (periode 1)"))
    pdf.ln(4)
    pdf.set_font(ff, "", 10)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(
        0,
        6,
        safe_txt(
            f"Fichier classeur : {wb_path.name}\n"
            f"Script associe : run_tre_p1_e2e_test.py / generate_tre_p1_test_pdf.py\n"
            f"Revenu de reference TRE (somme PxU) : {rev:,.0f} $\n"
            f"BudgetBase ajuste P1 (x1,12 x1,07) : {bba:,.0f} $\n"
            f"Seuils marketing TRE P1 (PARAM) : min {mkt_min_ref:,.0f} $ / max {mkt_max_ref:,.0f} $\n"
            f"Meme regle (fallback sans classeur) : min {m_min_naive:,.0f} $ / max {m_max_naive:,.0f} $\n"
            f"Budget R&D injecte (2 %) : {rd:,} $\n"
            f"Prix liste estime P040 (base x inflation P1) : {plist:,.0f} $\n"
        ),
    )
    pdf.ln(6)

    set_section_title(pdf, ff, 12)
    pdf.cell(0, 8, safe_txt("1. Decisions injectees — INPUT_FIRM (TRE P1)"))
    pdf.ln(10)
    pdf.set_font(ff, "", 9)
    for k, v in MARKETING_TRE.items():
        pdf.cell(80, 6, safe_txt(k), border=0)
        pdf.cell(0, 6, safe_txt(f"{v:,.0f} $"), ln=1)
    pdf.cell(80, 6, safe_txt("Total marketing"))
    pdf.cell(0, 6, safe_txt(f"{sum(MARKETING_TRE.values()):,.0f} $"), ln=1)
    pdf.cell(80, 6, safe_txt("BudgetRD"))
    pdf.cell(0, 6, safe_txt(f"{rd:,} $"), ln=1)
    pdf.cell(80, 6, safe_txt("SustainabilityInvestPct"))
    pdf.cell(0, 6, "0,005", ln=1)
    pdf.cell(80, 6, safe_txt("LancementPrevu"))
    pdf.cell(0, 6, "0", ln=1)
    pdf.ln(4)

    set_section_title(pdf, ff, 12)
    pdf.cell(0, 8, safe_txt("2. INPUT_MODEL — Productions TRE P1"))
    pdf.ln(10)
    pdf.set_font(ff, "", 9)
    for pk in sorted(PRODUCTIONS.keys()):
        pdf.cell(40, 5, pk)
        pdf.cell(0, 5, safe_txt(f"{PRODUCTIONS[pk]:,} unites"), ln=1)
    pdf.ln(4)

    set_section_title(pdf, ff, 12)
    pdf.cell(0, 8, safe_txt("3. CONTROLE_DECISIONS — Synthese Python (sans recalcul Excel)"))
    pdf.ln(10)
    pdf.set_font(ff, "", 8)
    pdf.set_fill_color(230, 236, 245)
    pdf.cell(55, 7, safe_txt("Type"), 1, 0, "L", True)
    pdf.cell(25, 7, safe_txt("Statut"), 1, 0, "C", True)
    pdf.cell(100, 7, safe_txt("Message"), 1, 1, "L", True)
    pdf.set_font(ff, "", 8)
    for row in report:
        st = str(row["statut"])
        if "✗" in st or "X" in st:
            fill = (248, 234, 234)
        elif "Attention" in st or "!" in st:
            fill = (253, 243, 227)
        else:
            fill = (232, 245, 240)
        pdf.set_fill_color(*fill)
        msg = (row.get("msg") or "")[:85]
        pdf.cell(55, 6, safe_txt(str(row["type"])[:40]), 1, 0, "L", True)
        pdf.cell(25, 6, safe_txt(str(row["statut"])[:14]), 1, 0, "C", True)
        pdf.cell(100, 6, safe_txt(msg), 1, 1, "L", True)
    pdf.ln(4)

    set_section_title(pdf, ff, 12)
    pdf.cell(0, 8, safe_txt("4. Bandeau et interpretation"))
    pdf.ln(8)
    pdf.set_font(ff, "", 9)
    pdf.multi_cell(
        0,
        5,
        safe_txt(
            "Cellule A1 de CONTROLE_DECISIONS : formule COUNTIF sur la colonne Statut (bloques ✗).\n"
            "Les alertes ⚠ (marketing sous plancher) ne declenchent pas le bandeau rouge ; "
            "depassement du plafond marketing → ✗ Bloque.\n"
            f"Nombre de lignes en ✗ Bloque (evaluation Python) : {n_blocked}.\n"
        ),
    )
    pdf.ln(4)

    set_section_title(pdf, ff, 12)
    pdf.cell(0, 8, safe_txt("5. Test invalide (documentation)"))
    pdf.ln(8)
    pdf.set_font(ff, "", 9)
    pdf.multi_cell(
        0,
        5,
        safe_txt(
            "Scenario avec BudgetInfluencer eleve (depassement plafond ~132 k$ TRE P1), promotion P040 = -15 %, "
            "LancementPrevu = 1. Le portefeuille plein (NbActifs >= NbInit+2) depend du moteur Excel.\n"
            "Feuille TEST_TRE_P1_INVALIDE dans le classeur v4_TEST.xlsx.\n"
        ),
    )

    PDF_OUT.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(PDF_OUT))
    shutil.copy2(PDF_OUT, PDF_DOWNLOADS)
    print(f"PDF genere : {PDF_OUT}")
    print(f"Copie : {PDF_DOWNLOADS}")


if __name__ == "__main__":
    main()
