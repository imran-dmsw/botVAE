"""
Genere un PDF de rapport firme (ex. TRE) a partir de la feuille RAPPORT_<FIRME>_P<n>
du classeur VAE (openpyxl + ReportLab + vae_report_style).
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from reportlab.lib import colors
from reportlab.lib.units import mm

from vae_report_style import (
    CONTENT_WIDTH,
    GREEN_LT,
    AMBER_LT,
    RED_LT,
    H1,
    Paragraph,
    P_INS,
    P_JUST,
    P_SMALL,
    Spacer,
    box_info,
    build_doc,
    cover_banner,
    pdf_escape,
    table_standard,
)


YELLOW_LT = colors.HexColor("#FFFDE7")

DEFAULT_WB = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v6.xlsx")
OUT_DEFAULT = Path("reports/Rapport_firme_{firm}_P{periode}_VAE.pdf")
DOWNLOADS_DIR = Path("/Users/imran/Downloads")


def row_fill_alerte(alerte: str):
    a = pdf_escape(str(alerte))
    if any(x in a for x in ("Rupture", "Surproduction", "invalide", "X Bloque")):
        return RED_LT
    if "Prudente" in a or "Liquidation" in a:
        return AMBER_LT
    if "Sécuritaire" in a or "Securitaire" in a:
        return YELLOW_LT
    if "Équilibrée" in a or "Equilibree" in a or "OK" == a.strip():
        return GREEN_LT
    return None


def fmt_pct(x) -> str:
    if x is None:
        return "—"
    try:
        return f"{float(x) * 100:.1f} %"
    except (TypeError, ValueError):
        return str(x)


def fmt_num(x) -> str:
    if x is None:
        return "—"
    try:
        return f"{int(round(float(x))):,}".replace(",", " ")
    except (TypeError, ValueError):
        return str(x)


def load_sheet_rows(ws: openpyxl.worksheet.worksheet.Worksheet):
    products = []
    r = 3
    while ws.cell(r, 1).value:
        row = [ws.cell(r, c).value for c in range(1, 16)]
        products.append(row)
        r += 1
    kpis = []
    while r <= ws.max_row:
        label = ws.cell(r, 1).value
        if label is None:
            r += 1
            continue
        if str(label).strip() == "KPIs firme":
            r += 1
            continue
        kpis.append((str(label), ws.cell(r, 2).value))
        r += 1
    return products, kpis


def main():
    ap = argparse.ArgumentParser(description="PDF rapport firme VAE depuis feuille RAPPORT_*")
    ap.add_argument("--firm", default="TRE", help="Code firme (ex. TRE)")
    ap.add_argument("--periode", type=int, default=1, help="Numero de periode")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WB, help="Chemin classeur .xlsx")
    ap.add_argument("--out", type=Path, default=None, help="Fichier PDF sortie")
    args = ap.parse_args()

    wb_path = args.workbook.expanduser().resolve()
    if not wb_path.exists():
        raise SystemExit(f"Classeur introuvable : {wb_path}")

    sheet_name = f"RAPPORT_{args.firm}_P{args.periode}"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Feuille absente : {sheet_name} (disponibles : RAPPORT_* )")

    ws = wb[sheet_name]
    products, kpis = load_sheet_rows(ws)

    story = []
    story.append(
        cover_banner(
            f"RAPPORT FIRME {args.firm} — PÉRIODE {args.periode}",
            f"Source : {wb_path.name} — feuille {sheet_name}",
        )
    )
    story.append(Spacer(1, 12))

    intro = (
        "Vue synthétique par produit : niveaux de production, demande active, écoulement, "
        "taux de couverture (stock départ + production) / demande, zone d’alerte et production indicielle."
    )
    story.append(Paragraph(pdf_escape(intro), P_JUST))
    story.append(Spacer(1, 14))

    H1(story, "Tableau par modèle")
    pdf_headers = [
        "Produit",
        "Modèle",
        "Segment",
        "Gamme",
        "Prix liste",
        "Production",
        "Demande",
        "Ventes",
        "Stock fin",
        "Marge",
        "Alerte",
        "Taux couv.",
        "Zone",
        "Prod. reco.",
    ]
    rows_pdf = []
    fills = []
    for row in products:
        pk, mn, seg, rng, pl, prod, dem, ts, se, mrg, alerte, tc, zone, prec, reco = row[:15]
        rows_pdf.append(
            [
                pk or "—",
                (mn or pk or "—") if isinstance(mn, str) else str(mn or "—"),
                seg or "—",
                rng or "—",
                fmt_num(pl),
                fmt_num(prod),
                fmt_num(dem),
                fmt_num(ts),
                fmt_num(se),
                fmt_pct(mrg),
                str(alerte or "—"),
                fmt_pct(tc),
                str(zone or "—"),
                fmt_num(prec),
            ]
        )
        fills.append(row_fill_alerte(str(alerte or "")))

    # Colonnes étroites : ~ proportions sur CONTENT_WIDTH
    w = CONTENT_WIDTH
    col_widths = [
        22 * mm,
        26 * mm,
        22 * mm,
        18 * mm,
        22 * mm,
        18 * mm,
        18 * mm,
        18 * mm,
        18 * mm,
        16 * mm,
        22 * mm,
        18 * mm,
        22 * mm,
        18 * mm,
    ]
    scale = w / sum(col_widths)
    col_widths = [x * scale for x in col_widths]

    story.append(
        table_standard(
            pdf_headers,
            rows_pdf,
            col_widths,
            row_fills=fills,
            repeat_rows=1,
        )
    )
    story.append(Spacer(1, 10))

    H1(story, "Recommandations (extrait)")
    rec_lines = []
    for row in products:
        reco = row[14] if len(row) > 14 else None
        pk = row[0]
        if reco and str(reco).strip() and str(reco) != "Maintenir la strategie":
            rec_lines.append(f"<b>{pdf_escape(str(pk))}</b> — {pdf_escape(str(reco))}")
    if rec_lines:
        story.append(
            box_info(
                Paragraph("<br/>".join(rec_lines[:25]), P_INS),
            )
        )
    else:
        story.append(Paragraph(pdf_escape("Aucune recommandation spécifique — stratégie à maintenir."), P_SMALL))

    story.append(Spacer(1, 14))
    H1(story, "Indicateurs agrégés (feuille Excel)")
    kpi_rows = [[pdf_escape(str(l)), pdf_escape(str(v if v is not None else "—"))] for l, v in kpis]
    if kpi_rows:
        kw = CONTENT_WIDTH
        story.append(
            table_standard(
                ["Indicateur", "Valeur"],
                kpi_rows,
                [kw * 0.45, kw * 0.55],
                row_fills=None,
            )
        )
    else:
        story.append(Paragraph(pdf_escape("KPIs non disponibles (recalculer le classeur dans Excel)."), P_SMALL))

    out_rel = args.out or Path(str(OUT_DEFAULT).format(firm=args.firm, periode=args.periode))
    if not out_rel.is_absolute():
        out_rel = Path.cwd() / out_rel
    out_rel.parent.mkdir(parents=True, exist_ok=True)

    build_doc(
        out_rel,
        story,
        firm_code=args.firm,
        left_text=f"Simulation marché VAE — Rapport {args.firm} P{args.periode}",
    )

    dl = DOWNLOADS_DIR / out_rel.name
    try:
        shutil.copy2(out_rel, dl)
        print(f"PDF : {out_rel}")
        print(f"Copie : {dl}")
    except OSError:
        print(f"PDF : {out_rel}")


if __name__ == "__main__":
    main()
