"""Rapport PDF VAE moderne (vue firme + fiches modèles)."""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from reportlab.lib.colors import Color
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle

from config.market_config import MARKET_CONFIG
from engine.budget_allocation import snap_rd_firm_pct
from engine.market_matrix import build_cross_matrix_pct_total_market
from engine.models import MarketingChannels, ScenarioInput, SimulationResult
from engine.rapport_alertes import build_prioritized_alerts, flatten_prioritized_alerts
from engine.simulation import period_inflation_factor, period_to_year, segment_size, simulate, total_market_size
from vae_rapport_firme_logic import default_excel_path, resolve_firm_code
from vae_report_style import (
    CONTENT_WIDTH,
    GREEN_LT,
    NAVY,
    P_INS,
    P_JUST,
    P_SMALL,
    P_TOC,
    H1,
    PageBreak,
    Paragraph,
    Spacer,
    build_doc,
    cover_banner,
    kpi_block,
    pdf_escape,
    pdf_xml_fragment,
    table_standard,
    toc_block,
)
from vae_report_tables import (
    canonical_gamme_label,
    competitor_reaction_blurb,
    portfolio_anomaly_note,
    production_band_note,
    scenario_display_title,
    top3_channels_cell,
)
from vae_rapport_firme_logic import build_firm_recommendations, build_model_recommendations

DEFAULT_XLSX = default_excel_path()
OUT_PDF = Path("reports/Rapport_{firm}_VAE_{date}.pdf")
DOWNLOADS = Path("/Users/imran/Downloads")

FIRM_LEGEND = {
    "TRE": "TrailRidge Mobility (TRE)",
    "AVE": "AVE Mobility (AVE)",
    "CAN": "CAN Performance (CAN)",
    "EBI": "EBI Confort (EBI)",
    "GIA": "GIA Famille (GIA)",
    "PED": "PED Prudent (PED)",
    "RID": "RID Terre & Trail (RID)",
    "SUR": "SUR Loisirs (SUR)",
    "VEL": "VEL Endurance (VEL)",
}

SEG_IDX_TO_KEY = {
    1: "urbains_presses",
    2: "prudentes_confort",
    3: "endurants_performants",
    4: "nomades_multimodaux",
    5: "familles_cargo",
    6: "aventuriers_tt",
}
SEG_KEY_TO_IDX = {v: k for k, v in SEG_IDX_TO_KEY.items()}
RANGE_MAP = {
    "Basic": "entry",
    "Medium": "mid",
    "Standard": "mid",
    "Premium": "premium",
    "Bas": "entry",
    "Moyen": "mid",
    "Haut": "premium",
}

AUTO_SCENARIO_TAGS = (
    "_FocalPremium_P2",
    "_FocalPremium_P1",
    "_Marketing_Max_P1",
    "_RD_5pct_P2",
    "_SousProduction_P1",
    "_Durabilite_2_tranches",
    "_Figee_P4",
    "_Figee_P8",
)


def fmt_money(x: float) -> str:
    return f"{int(round(float(x))):,}".replace(",", " ")


def fmt_pct(x: float, dec: int = 1) -> str:
    return f"{100.0 * float(x):.{dec}f} %"


def fmt_pct_and_money(pct: float, amount: float) -> str:
    return f"{pct * 100:.1f} % ({fmt_money(amount)} $)".replace(".", ",")


def product_type_for_segment(seg_idx: int) -> str:
    if seg_idx == 6:
        return "vtt_enduro"
    if seg_idx == 5:
        return "cargo_familial"
    if seg_idx == 3:
        return "route_connecte"
    if seg_idx in (2, 4):
        return "vtc_polyvalent"
    return "ville_quotidien"


@dataclass
class ProdRow:
    product_key: str
    model_name: str
    segment_idx: int
    range_raw: str
    base_price: float
    units: float

    @property
    def revenue(self) -> float:
        return float(self.base_price) * float(self.units)


def load_products(wb: openpyxl.Workbook, firm: str) -> list[ProdRow]:
    ws = wb["BASE_REFERENCE_MODEL"]
    h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    out: list[ProdRow] = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, h["Firm"]).value or "").upper() != firm:
            continue
        seg = ws.cell(r, h["Segment"]).value
        try:
            seg_i = int(seg)
        except (TypeError, ValueError):
            continue
        out.append(
            ProdRow(
                product_key=str(ws.cell(r, h["ProductKey"]).value or "").strip(),
                model_name=str(ws.cell(r, h["ModelName"]).value or "").strip(),
                segment_idx=seg_i,
                range_raw=str(ws.cell(r, h["Range"]).value or "Medium").strip(),
                base_price=float(ws.cell(r, h["BasePrice_RefYear"]).value or 0.0),
                units=float(ws.cell(r, h["Units_RefYear"]).value or 0.0),
            )
        )
    return out


def pick_focal_product(products: list[ProdRow]) -> ProdRow:
    if not products:
        raise ValueError("Aucun produit pour choisir un focal product.")
    for p in products:
        if p.segment_idx == 3 and canonical_gamme_label(p.range_raw) == "Premium":
            return p
    return max(products, key=lambda p: p.revenue)


def firm_ref_revenue(products: list[ProdRow]) -> float:
    return max(sum(p.revenue for p in products), 150_000.0)


def segment_label_from_idx(wb: openpyxl.Workbook, idx: int) -> str:
    ws = wb["SEGMENTS"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == idx:
            return str(ws.cell(r, 2).value or f"Segment {idx}")
    return f"Segment {idx}"


def load_mm_rows(wb: openpyxl.Workbook) -> list[list[str]]:
    ws = wb["MARKETING_MATRIX"]
    out: list[list[str]] = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        out.append([str(ws.cell(r, c).value or "") for c in range(1, 7)])
    return out


def load_firms_market_rows(wb: openpyxl.Workbook) -> list[tuple[str, float, float]]:
    ws = wb["FIRMS"]
    rows: list[tuple[str, float, float]] = []
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        if not code:
            continue
        rows.append(
            (
                code,
                float(ws.cell(r, 3).value or 0.0),
                float(ws.cell(r, 4).value or 0.0),
            )
        )
    rows.sort(key=lambda row: row[2], reverse=True)
    return rows


def growth_rows(firm: str) -> tuple[list[list[str]], str]:
    _ = firm
    rows: list[list[str]] = []
    p1_tot = total_market_size(1)
    seg_key = "endurants_performants"
    seg_label = "Seg. 3 (Endurants)"
    for p in range(1, 9):
        tot = total_market_size(p)
        seg = segment_size(p, seg_key)
        growth = "" if p == 1 else f"+{(tot / p1_tot - 1) * 100:.0f} %"
        rows.append(
            [
                str(p),
                str(period_to_year(p)),
                f"{tot:,.0f}".replace(",", " "),
                f"{seg:,.0f}".replace(",", " "),
                growth,
            ]
        )
    return rows, seg_label


def _scenario_prefix(firm: str, prod: ProdRow) -> str:
    clean = "".join(ch if (ch.isalnum() or ch in "_-") else "_" for ch in str(prod.product_key))
    return f"{firm}_{clean}"[:40]


def _build_scenario(
    prod: ProdRow,
    firm: str,
    period: int,
    sid: str,
    adj_firm_budget: float,
    *,
    marketing_mult: float = 1.0,
    rd_pct: float = 0.02,
    promo: float = 0.0,
    production_mult: float = 1.0,
    sustainability_tranches: int = 0,
    opening_stock: int = 0,
    allocation_weight: float = 1.0,
) -> ScenarioInput:
    seg_key = SEG_IDX_TO_KEY.get(prod.segment_idx, "urbains_presses")
    rng = RANGE_MAP.get(canonical_gamme_label(prod.range_raw), "mid")
    infl = period_inflation_factor(period)
    price = max(500.0, prod.base_price * infl)
    production = max(100, int(round(prod.units * production_mult)))

    mkt_cap = adj_firm_budget * MARKET_CONFIG["constraints"]["marketing_max_pct"]
    firm_mkt = min(max(adj_firm_budget * 0.08 * marketing_mult, 0.0), mkt_cap)
    rd_ratio = snap_rd_firm_pct(rd_pct)
    firm_rd = adj_firm_budget * rd_ratio
    product_mkt = max(firm_mkt * allocation_weight, 0.0)
    product_rd = max(firm_rd * allocation_weight, 0.0)
    product_adj = max(adj_firm_budget * allocation_weight, 1.0)
    rep = float(MARKET_CONFIG["firms"][firm]["base_rep"])
    return ScenarioInput(
        firm_name=firm,
        period=period,
        scenario_name=sid,
        model_name=prod.model_name,
        product_type=product_type_for_segment(prod.segment_idx),
        segment=seg_key,
        model_range=rng,
        product_status="active",
        marketing_budget=round(product_mkt, 2),
        marketing_channels=MarketingChannels(),
        rd_budget=round(product_rd, 2),
        price=round(price, 2),
        promotion_rate=promo,
        production=production,
        opening_stock=max(0, int(opening_stock)),
        adjusted_budget=round(product_adj, 2),
        allocation_weight=allocation_weight,
        firm_marketing_budget_total=round(firm_mkt, 2),
        firm_rd_budget_total=round(firm_rd, 2),
        previous_innovation_score=rep,
        previous_sustainability_score=5.0,
        sustainability_tranches=sustainability_tranches,
        competitor_attractiveness=14.0,
    )


def scenario_from_product(
    prod: ProdRow,
    firm_code: str,
    period: int,
    name: str,
    adj_budget: float,
    *,
    products: list[ProdRow] | None = None,
    marketing_mult: float = 1.0,
    rd_pct: float = 0.02,
    promo: float = 0.0,
    production_mult: float = 1.0,
    sustainability_tranches: int = 0,
    opening_stock: int = 0,
    **_: Any,
) -> ScenarioInput:
    """API stable pour excel_2250_reference et scripts externes."""
    portfolio = products or [prod]
    total_ref_rev = max(sum(p.revenue for p in portfolio), 1.0)
    weight = max(prod.revenue / total_ref_rev, 0.01)
    return _build_scenario(
        prod,
        firm_code,
        period,
        name,
        adj_budget,
        marketing_mult=marketing_mult,
        rd_pct=rd_pct,
        promo=promo,
        production_mult=production_mult,
        sustainability_tranches=sustainability_tranches,
        opening_stock=opening_stock,
        allocation_weight=weight,
    )


def build_scenario_grid(wb: openpyxl.Workbook, primary_firm: str) -> list[tuple[str, str, ScenarioInput]]:
    products = load_products(wb, primary_firm)
    if not products:
        return []
    focal = pick_focal_product(products)
    adj = firm_ref_revenue(products)
    grid: list[tuple[str, str, ScenarioInput]] = []
    sid_p2 = f"{primary_firm}_FocalPremium_P2"
    s_p2 = _build_scenario(
        focal,
        primary_firm,
        2,
        sid_p2,
        adj,
        marketing_mult=1.0,
        rd_pct=0.05,
        production_mult=1.08,
        sustainability_tranches=2,
    )
    grid.append((sid_p2, primary_firm, s_p2))
    sid_p1 = f"{primary_firm}_FocalPremium_P1"
    s_p1 = _build_scenario(focal, primary_firm, 1, sid_p1, adj, production_mult=1.05)
    grid.append((sid_p1, primary_firm, s_p1))
    sid_mkt = f"{primary_firm}_Marketing_Max_P1"
    s_mkt = _build_scenario(focal, primary_firm, 1, sid_mkt, adj, marketing_mult=1.35, production_mult=1.05)
    grid.append((sid_mkt, primary_firm, s_mkt))
    sid_rd = f"{primary_firm}_RD_5pct_P2"
    s_rd = _build_scenario(focal, primary_firm, 2, sid_rd, adj, rd_pct=0.05, production_mult=1.06)
    grid.append((sid_rd, primary_firm, s_rd))
    sid_under = f"{primary_firm}_SousProduction_P1"
    s_under = _build_scenario(focal, primary_firm, 1, sid_under, adj, production_mult=0.72)
    grid.append((sid_under, primary_firm, s_under))
    sid_dur = f"{primary_firm}_Durabilite_2_tranches"
    s_dur = _build_scenario(
        focal,
        primary_firm,
        1,
        sid_dur,
        adj,
        production_mult=1.02,
        sustainability_tranches=2,
    )
    grid.append((sid_dur, primary_firm, s_dur))
    s_p4 = s_p1.model_copy(update={"period": 4, "scenario_name": f"{primary_firm}_Figee_P4"})
    s_p8 = s_p1.model_copy(update={"period": 8, "scenario_name": f"{primary_firm}_Figee_P8"})
    grid.append((s_p4.scenario_name, primary_firm, s_p4))
    grid.append((s_p8.scenario_name, primary_firm, s_p8))
    return grid


def run_grid(grid: list[tuple[str, str, ScenarioInput]]) -> list[tuple[str, str, ScenarioInput, SimulationResult]]:
    out: list[tuple[str, str, ScenarioInput, SimulationResult]] = []
    for sid, owner, sc in grid:
        out.append((sid, owner, sc, simulate(sc)))
    out.sort(key=lambda row: row[3].profit, reverse=True)
    return out


def firm_scenario_results(
    results: list[tuple[str, str, ScenarioInput, SimulationResult]],
    firm: str,
) -> list[tuple[str, str, ScenarioInput, SimulationResult]]:
    return [row for row in results if row[1] == firm]


def firm_adjusted_budget_from_scenario(scen: ScenarioInput) -> float:
    w = max(float(scen.allocation_weight), 1e-9)
    return float(scen.adjusted_budget) / w


def _projection_1500_2250_3000(ref_scen: ScenarioInput) -> list[tuple[int, SimulationResult]]:
    out: list[tuple[int, SimulationResult]] = []
    for units in (1500, 2250, 3000):
        sc = ref_scen.model_copy(update={"production": units, "scenario_name": f"{ref_scen.scenario_name}_proj_{units}"})
        out.append((units, simulate(sc)))
    return out


def _build_product_bundle(prod: ProdRow, firm: str, adj_firm: float, allocation_weight: float) -> dict[str, Any]:
    pref = _scenario_prefix(firm, prod)
    ref_scen = _build_scenario(
        prod,
        firm,
        2,
        f"{pref}_P2ref",
        adj_firm,
        rd_pct=0.05,
        production_mult=1.08,
        sustainability_tranches=2,
        allocation_weight=allocation_weight,
    )
    ref_res = simulate(ref_scen)
    lever_scenarios = [
        ("Référence P1", _build_scenario(prod, firm, 1, f"{pref}_P1", adj_firm, production_mult=1.05, allocation_weight=allocation_weight)),
        (
            "Marketing maximum",
            _build_scenario(prod, firm, 1, f"{pref}_MktMax", adj_firm, marketing_mult=1.35, production_mult=1.05, allocation_weight=allocation_weight),
        ),
        (
            "Promotion -5 %",
            _build_scenario(prod, firm, 1, f"{pref}_Promo5", adj_firm, promo=-0.05, production_mult=1.05, allocation_weight=allocation_weight),
        ),
        (
            "Sous-production (-28 %)",
            _build_scenario(prod, firm, 1, f"{pref}_SousProd", adj_firm, production_mult=0.72, allocation_weight=allocation_weight),
        ),
    ]
    levers = [(lab, simulate(sc), sc) for lab, sc in lever_scenarios]
    r_p1 = levers[0][1]
    s_p4 = lever_scenarios[0][1].model_copy(update={"period": 4, "scenario_name": f"{pref}_FigP4"})
    s_p8 = lever_scenarios[0][1].model_copy(update={"period": 8, "scenario_name": f"{pref}_FigP8"})
    grouped = build_prioritized_alerts(scenario=ref_scen, result=ref_res, product=prod)
    return {
        "prod": prod,
        "ref_scen": ref_scen,
        "ref_res": ref_res,
        "levers": [(lab, res) for lab, res, _ in levers],
        "lever_scenarios": [(lab, sc) for lab, _, sc in levers],
        "r_p1": r_p1,
        "r_p4": simulate(s_p4),
        "r_p8": simulate(s_p8),
        "alerts_grouped": grouped,
        "alerts": flatten_prioritized_alerts(grouped),
        "projection_1500_2250_3000": _projection_1500_2250_3000(ref_scen),
    }


def _safe_cross_matrix(firm: str, scen: ScenarioInput) -> dict[str, Any]:
    return build_cross_matrix_pct_total_market(period=1, user_firm=firm, user_scenario=scen)


def build_report_payload(xlsx: Path, firm: str) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    produits = load_products(wb, firm)
    if not produits:
        wb.close()
        raise ValueError(f"Aucun produit trouvé pour la firme {firm}.")
    focal_prod = pick_focal_product(produits)
    mm_rows = load_mm_rows(wb)
    grid = build_scenario_grid(wb, firm)
    results = run_grid(grid)
    firm_rows = firm_scenario_results(results, firm)
    if not firm_rows:
        wb.close()
        raise ValueError(f"Aucun scénario pour la firme {firm}.")
    sid_ref, _owner_ref, ref_scen, ref_res = next((r for r in firm_rows if "_FocalPremium_P2" in r[0]), firm_rows[0])
    adj_firm = firm_adjusted_budget_from_scenario(ref_scen)
    total_ref_rev = max(sum(p.revenue for p in produits), 1.0)
    product_bundles: dict[str, dict[str, Any]] = {}
    for prod in produits:
        w = max(prod.revenue / total_ref_rev, 0.01)
        product_bundles[prod.product_key] = _build_product_bundle(prod, firm, adj_firm, w)
    r_p1 = next((r[3] for r in firm_rows if "_FocalPremium_P1" in r[0]), ref_res)
    r_p4 = next((r[3] for r in firm_rows if "_Figee_P4" in r[0]), ref_res)
    r_p8 = next((r[3] for r in firm_rows if "_Figee_P8" in r[0]), ref_res)
    recommandations_firme = build_firm_recommendations(
        firm=firm,
        legend=FIRM_LEGEND.get(firm, firm),
        prods=produits,
        focal_prod=focal_prod,
        ref_scen=ref_scen,
        ref_res=ref_res,
        r_p8=r_p8,
        results=firm_rows,
    )
    fiches_modeles = {
        prod.product_key: {
            "prod": prod,
            "bundle": product_bundles[prod.product_key],
            "recommandations": build_model_recommendations(prod, product_bundles[prod.product_key], mm_rows),
        }
        for prod in produits
    }
    vue_firme = {
        "legend": FIRM_LEGEND.get(firm, firm),
        "firm_results": firm_rows,
        "ref_sid": sid_ref,
        "ref_scen": ref_scen,
        "ref_res": ref_res,
        "r_p1": r_p1,
        "r_p4": r_p4,
        "r_p8": r_p8,
        "projection_1500_2250_3000": _projection_1500_2250_3000(ref_scen),
        "cross_matrix_pct": _safe_cross_matrix(firm, ref_scen),
    }
    return {
        "xlsx": xlsx,
        "wb": wb,
        "firm": firm,
        "legend": FIRM_LEGEND.get(firm, firm),
        "produits": produits,
        "focal_prod": focal_prod,
        "mm_rows": mm_rows,
        "grid": grid,
        "results": results,
        "product_bundles": product_bundles,
        "vue_firme": vue_firme,
        "fiches_modeles": fiches_modeles,
        "recommandations_firme": recommandations_firme,
        "cross_matrix_pct": vue_firme["cross_matrix_pct"],
    }


def append_reading_note(story: list, text: str) -> None:
    story.append(Spacer(1, 3))
    story.append(Paragraph(f"Note de lecture : {pdf_escape(text)}", P_SMALL))


def _append_cover_exec_toc(story: list, code_firme: str, vue_firme: dict[str, Any], fiches_modeles: dict[str, Any]) -> None:
    legend = vue_firme["legend"]
    best = max(vue_firme["firm_results"], key=lambda row: row[3].profit)
    summary = (
        f"<b>Résumé exécutif.</b> {pdf_escape(legend)} — {len(vue_firme['firm_results'])} scénarios firmes simulés. "
        f"Meilleur profit: <b>{pdf_escape(scenario_display_title(best[0]))}</b> ({fmt_money(best[3].profit)} $). "
        f"Le rapport regroupe une vue firme (sections 1.1 à 1.7, 1.10) puis {len(fiches_modeles)} fiches modèle."
    )
    story.append(cover_banner("RAPPORT VAE - ANALYSE STRATEGIQUE", f"Firme {code_firme} - {legend}"))
    story.append(Spacer(1, 10))
    story.append(Paragraph(summary, P_JUST))
    story.append(PageBreak())
    story.append(Paragraph("<b>Table des matieres</b>", P_INS))
    toc_items = [
        "PARTIE VUE FIRME",
        "1.1 Tableau de bord scenarios (firme)",
        "1.2 Parts de marche firmes",
        "1.3 Segments de marche",
        "1.4 Croissance du marche",
        "Tableau croise firme x segment",
        "1.5 Portefeuille et performance",
        "1.6 Leviers strategiques",
        "1.7 Reference + projections 1500/2250/3000",
        "1.10 Recommandations",
        "PARTIE VUE MODELE",
    ]
    for k in fiches_modeles.keys():
        toc_items.append(f"Modele {k}: contexte, portefeuille, budgets, leviers, performance, alertes, recommandations")
    story.extend(toc_block(toc_items))
    story.append(PageBreak())


def _append_section_11_dashboard(story: list, firm_results: list[tuple[str, str, ScenarioInput, SimulationResult]], firm: str) -> None:
    H1(story, "1.1 Tableau de bord scenarios - vue firme")
    kpi_items = [
        ("Scenarios", str(len(firm_results)), NAVY),
        ("Meilleur profit", f"{fmt_money(max(r[3].profit for r in firm_results))} $", NAVY),
        ("Marge max", fmt_pct(max(r[3].margin for r in firm_results)), NAVY),
        ("Service max", fmt_pct(max(r[3].service_rate for r in firm_results), 0), NAVY),
    ]
    story.append(kpi_block(kpi_items))
    rows: list[list[str]] = []
    fills: list[Color | None] = []
    for sid, owner, scen, res in sorted(firm_results, key=lambda row: row[3].profit, reverse=True):
        rows.append(
            [
                scenario_display_title(sid),
                owner,
                str(scen.period),
                str(res.sales),
                fmt_money(res.revenue),
                fmt_money(res.profit),
                fmt_pct(res.margin),
                fmt_pct(res.market_share),
                fmt_pct(res.service_rate, 0),
                competitor_reaction_blurb(sid, firm),
            ]
        )
        fills.append(GREEN_LT)
    w = CONTENT_WIDTH
    story.append(
        table_standard(
            ["Scenario", "Firme", "P", "Ventes", "CA", "Profit", "Marge", "PDM", "Service", "Reaction concurrents"],
            rows,
            [w * 0.19, w * 0.06, w * 0.05, w * 0.07, w * 0.09, w * 0.09, w * 0.08, w * 0.08, w * 0.08, w * 0.21],
            row_fills=fills,
        )
    )
    append_reading_note(story, "Colonne 'Reaction concurrents' ajoutee pour lecture rapide des contre-mouvements.")


def _append_sections_12_14_context(story: list, wb: openpyxl.Workbook, firm: str) -> None:
    H1(story, "1.2 Parts de marche firmes")
    rows_f = [[code, f"{units:,.0f}".replace(",", " "), fmt_pct(share)] for code, units, share in load_firms_market_rows(wb)]
    story.append(table_standard(["Firme", "Unites ref.", "PDM"], rows_f, [CONTENT_WIDTH * 0.28, CONTENT_WIDTH * 0.35, CONTENT_WIDTH * 0.37]))
    append_reading_note(story, "PDM lue depuis la feuille FIRMS (annee de reference).")
    story.append(Spacer(1, 10))
    H1(story, "1.3 Structure des segments")
    rows_s: list[list[str]] = []
    ws = wb["SEGMENTS"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        rows_s.append(
            [
                str(ws.cell(r, 1).value),
                str(ws.cell(r, 2).value or ""),
                fmt_pct(float(ws.cell(r, 3).value or 0.0)),
                f"{float(ws.cell(r, 5).value or 0):,.0f}".replace(",", " "),
                fmt_pct(float(ws.cell(r, 6).value or 0.0)),
            ]
        )
    story.append(
        table_standard(
            ["Segment", "Description", "Part", "Unites ref.", "Online"],
            rows_s,
            [CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.43, CONTENT_WIDTH * 0.13, CONTENT_WIDTH * 0.18, CONTENT_WIDTH * 0.16],
        )
    )
    append_reading_note(story, "Le prix prefere n'est volontairement pas reproduit dans cette version.")
    story.append(Spacer(1, 10))
    H1(story, "1.4 Croissance du marche")
    gr, seg_col = growth_rows(firm)
    story.append(
        table_standard(
            ["Periode", "Annee", "Marche total", seg_col, "Croiss. vs P1"],
            gr,
            [CONTENT_WIDTH * 0.13, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.23],
        )
    )
    append_reading_note(story, "Projection 8 periodes, croissance globale +12 %/an.")


def _append_cross_matrix(story: list, cross: dict[str, Any]) -> None:
    H1(story, "Tableau croise firme x segment")
    firms = cross["firms"]
    segs = cross["segments"]
    grid = cross["pct_grid"]
    headers = ["Firme"] + [f"S{i}" for i in range(1, len(segs) + 1)] + ["TOTAL"]
    rows: list[list[str]] = []
    for i, firm in enumerate(firms):
        rows.append([firm] + [fmt_pct(grid[i][j] / 100.0, 2) for j in range(len(segs))] + [fmt_pct(cross["row_totals_pct"][i] / 100.0, 2)])
    rows.append(["TOTAL"] + [fmt_pct(cross["col_totals_pct"][j] / 100.0, 2) for j in range(len(segs))] + [fmt_pct(cross["grand_total_pct"] / 100.0, 2)])
    story.append(table_standard(headers, rows, [CONTENT_WIDTH / len(headers)] * len(headers)))
    if cross.get("footnote"):
        append_reading_note(story, str(cross["footnote"]))


def _append_section_15_portfolio(story: list, wb: openpyxl.Workbook, produits: list[ProdRow], product_bundles: dict[str, Any]) -> None:
    H1(story, "1.5 Portefeuille produits et grille performance")
    rows_p: list[list[str]] = []
    for p in produits:
        note = portfolio_anomaly_note(p.product_key) or production_band_note(p.units)
        rows_p.append(
            [
                p.product_key,
                p.model_name[:34],
                f"{p.segment_idx} - {segment_label_from_idx(wb, p.segment_idx)[:20]}",
                canonical_gamme_label(p.range_raw),
                fmt_money(p.base_price),
                fmt_money(p.units),
                note or "",
            ]
        )
    story.append(
        table_standard(
            ["Code", "Modele", "Segment", "Gamme", "Prix ref.", "Unites ref.", "Alerte / note"],
            rows_p,
            [CONTENT_WIDTH * 0.09, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.21, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.15],
        )
    )
    append_reading_note(story, "Les anomalies mettent en evidence cannibalisation, prix et volume hors bandes cibles.")
    story.append(Spacer(1, 10))
    rows_perf: list[list[str]] = []
    for p in produits:
        b = product_bundles[p.product_key]
        rs = b["ref_res"]
        rows_perf.append(
            [
                p.product_key,
                fmt_pct(rs.market_share),
                fmt_pct(rs.market_share_segment),
                fmt_pct(rs.margin),
                fmt_pct(rs.service_rate, 0),
                fmt_money(rs.forecast_ending_stock_units),
                fmt_money(rs.profit),
            ]
        )
    story.append(
        table_standard(
            ["Modele", "PDM glob.", "PDM seg.", "Marge", "Service", "Stock fin", "Profit"],
            rows_perf,
            [CONTENT_WIDTH * 0.15, CONTENT_WIDTH * 0.13, CONTENT_WIDTH * 0.13, CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.16, CONTENT_WIDTH * 0.19],
        )
    )
    append_reading_note(story, "Grille performance calculee sur les scenarios P2ref de chaque modele.")


def _append_section_16_levers(story: list, mm_rows: list[list[str]], firm_rows: list[tuple[str, str, ScenarioInput, SimulationResult]]) -> None:
    H1(story, "1.6 Leviers strategiques")
    rows_l: list[list[str]] = []
    for sid, _owner, _sc, res in sorted(firm_rows, key=lambda row: row[3].profit, reverse=True):
        rows_l.append(
            [
                scenario_display_title(sid),
                fmt_money(res.revenue),
                fmt_money(res.profit),
                fmt_pct(res.margin),
                fmt_pct(res.service_rate, 0),
                fmt_money(max(0, res.demand - res.sales)),
            ]
        )
    story.append(
        table_standard(
            ["Levier / scenario", "CA", "Profit", "Marge", "Service", "Demande non servie"],
            rows_l,
            [CONTENT_WIDTH * 0.36, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.14],
        )
    )
    append_reading_note(story, "Aucun sweep promo AVE n'est inclus dans la section leviers de cette version.")
    story.append(Spacer(1, 8))
    rows_m = [row + [top3_channels_cell(row)] for row in mm_rows]
    story.append(
        table_standard(
            ["Segment", "Digital", "Social", "Influenceur", "Affichage", "Evenements", "Top 3 canaux"],
            rows_m,
            [CONTENT_WIDTH * 0.09, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.38],
        )
    )
    append_reading_note(story, "Top 3 canaux derive de MARKETING_MATRIX.")


def _append_section_17_reference(story: list, vue_firme: dict[str, Any]) -> None:
    H1(story, "1.7 Scenario de reference et projection 1500 / 2250 / 3000")
    ref_scen: ScenarioInput = vue_firme["ref_scen"]
    ref_res: SimulationResult = vue_firme["ref_res"]
    params_rows = [
        ["Scenario", scenario_display_title(vue_firme["ref_sid"])],
        ["Periode", f"{ref_scen.period} ({period_to_year(ref_scen.period)})"],
        ["Prix catalogue", f"{fmt_money(ref_scen.price)} $"],
        ["Promotion", fmt_pct(ref_scen.promotion_rate)],
        ["Production", str(ref_scen.production)],
        ["Budget marketing (firme)", fmt_pct_and_money(ref_scen.firm_marketing_budget_total / max(firm_adjusted_budget_from_scenario(ref_scen), 1), ref_scen.firm_marketing_budget_total)],
        ["Budget R&D (firme)", fmt_pct_and_money(ref_scen.firm_rd_budget_total / max(firm_adjusted_budget_from_scenario(ref_scen), 1), ref_scen.firm_rd_budget_total)],
    ]
    story.append(table_standard(["Parametre", "Valeur"], params_rows, [CONTENT_WIDTH * 0.40, CONTENT_WIDTH * 0.60]))
    append_reading_note(story, "Budgets affiches au niveau firme, puis alloues aux produits par poids strategique.")
    story.append(Spacer(1, 8))
    proj_rows = []
    for units, res in vue_firme["projection_1500_2250_3000"]:
        proj_rows.append(
            [
                str(units),
                str(res.sales),
                fmt_pct(res.service_rate, 0),
                fmt_money(res.forecast_ending_stock_units),
                fmt_money(res.profit),
                "OK" if res.service_rate >= 0.85 else "Sous seuil 85 %",
            ]
        )
    story.append(
        table_standard(
            ["Production", "Ventes", "Service", "Stock fin", "Profit", "Lecture"],
            proj_rows,
            [CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.18, CONTENT_WIDTH * 0.20, CONTENT_WIDTH * 0.22],
        )
    )
    append_reading_note(story, "La grille 1500/2250/3000 aide a calibrer la capacite cible.")
    story.append(Spacer(1, 8))
    fig_rows = [
        ["P1", fmt_money(vue_firme["r_p1"].profit), fmt_pct(vue_firme["r_p1"].margin), fmt_pct(vue_firme["r_p1"].service_rate, 0)],
        ["P4", fmt_money(vue_firme["r_p4"].profit), fmt_pct(vue_firme["r_p4"].margin), fmt_pct(vue_firme["r_p4"].service_rate, 0)],
        ["P8", fmt_money(vue_firme["r_p8"].profit), fmt_pct(vue_firme["r_p8"].margin), fmt_pct(vue_firme["r_p8"].service_rate, 0)],
    ]
    story.append(
        table_standard(
            ["Strategie figee", "Profit", "Marge", "Service"],
            fig_rows,
            [CONTENT_WIDTH * 0.22, CONTENT_WIDTH * 0.28, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.25],
        )
    )
    append_reading_note(story, "Indicateur de derive temporelle sur strategie non ajustee.")


def _append_section_110_reco(story: list, recommandations: list[Any]) -> None:
    H1(story, "1.10 Recommandations")
    rows = [[r.action, r.impact, r.horizon, r.priority] for r in recommandations]
    story.append(
        table_standard(
            ["Action", "Impact attendu", "Horizon", "Priorite"],
            rows,
            [CONTENT_WIDTH * 0.35, CONTENT_WIDTH * 0.33, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.18],
        )
    )
    append_reading_note(story, "Priorisation issue des regles stratégiques firme.")


def _append_model_part(
    story: list,
    wb: openpyxl.Workbook,
    code_firme: str,
    mm_rows: list[list[str]],
    fiche: dict[str, Any],
) -> None:
    prod: ProdRow = fiche["prod"]
    bundle = fiche["bundle"]
    ref_scen: ScenarioInput = bundle["ref_scen"]
    ref_res: SimulationResult = bundle["ref_res"]
    story.append(PageBreak())
    part_style = ParagraphStyle(
        "PART_MODEL",
        parent=P_INS,
        fontSize=11,
        leading=13,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        textColor=NAVY,
    )
    story.append(Paragraph(f"PARTIE VUE MODELE - {pdf_escape(prod.product_key)}", part_style))
    story.append(Spacer(1, 6))
    H1(story, "2. Contexte concurrentiel")
    story.append(
        Paragraph(
            f"Modele <b>{pdf_xml_fragment(prod.product_key)}</b> - segment {prod.segment_idx} ({pdf_xml_fragment(segment_label_from_idx(wb, prod.segment_idx))}).",
            P_JUST,
        )
    )
    gr, seg_col = growth_rows(code_firme)
    story.append(
        table_standard(
            ["Periode", "Annee", "Marche total", seg_col, "Croiss. vs P1"],
            gr,
            [CONTENT_WIDTH * 0.13, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.25, CONTENT_WIDTH * 0.23],
        )
    )
    append_reading_note(story, "Contexte resume pour eviter de dupliquer toute la vue firme.")
    H1(story, "3. Portefeuille (focus modele)")
    story.append(
        table_standard(
            ["Code", "Modele", "Segment", "Gamme", "Prix ref.", "Unites ref."],
            [[
                prod.product_key,
                prod.model_name,
                f"{prod.segment_idx} - {segment_label_from_idx(wb, prod.segment_idx)}",
                canonical_gamme_label(prod.range_raw),
                fmt_money(prod.base_price),
                fmt_money(prod.units),
            ]],
            [CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.33, CONTENT_WIDTH * 0.20, CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.12],
        )
    )
    H1(story, "3.1 Budgets du modele")
    adj_firm = firm_adjusted_budget_from_scenario(ref_scen)
    rows_budget = [
        ["Budget ajuste firme", f"{fmt_money(adj_firm)} $"],
        ["Budget marketing firme", fmt_pct_and_money(ref_scen.firm_marketing_budget_total / max(adj_firm, 1), ref_scen.firm_marketing_budget_total)],
        ["Budget R&D firme", fmt_pct_and_money(ref_scen.firm_rd_budget_total / max(adj_firm, 1), ref_scen.firm_rd_budget_total)],
        ["Poids strategique modele", f"{ref_scen.allocation_weight * 100:.1f} %".replace(".", ",")],
        ["Budget marketing modele", f"{fmt_money(ref_scen.marketing_budget)} $"],
        ["Budget R&D modele", f"{fmt_money(ref_scen.rd_budget)} $"],
    ]
    story.append(table_standard(["Indicateur", "Valeur"], rows_budget, [CONTENT_WIDTH * 0.45, CONTENT_WIDTH * 0.55]))
    append_reading_note(story, "Les lignes 'Part marketing/R&D imputee au modele' ont ete retirees.")
    H1(story, "4. Leviers")
    rows_lev = []
    for lab, res in bundle["levers"]:
        rows_lev.append([lab, fmt_money(res.revenue), fmt_money(res.profit), fmt_pct(res.margin), fmt_pct(res.service_rate, 0)])
    story.append(
        table_standard(
            ["Levier", "CA", "Profit", "Marge", "Service"],
            rows_lev,
            [CONTENT_WIDTH * 0.36, CONTENT_WIDTH * 0.18, CONTENT_WIDTH * 0.18, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.14],
        )
    )
    seg_mm = next((row for row in mm_rows if row[0] == str(prod.segment_idx)), None)
    if seg_mm:
        story.append(
            table_standard(
                ["Segment", "Digital", "Social", "Influenceur", "Affichage", "Evenements", "Top 3 canaux"],
                [seg_mm + [top3_channels_cell(seg_mm)]],
                [CONTENT_WIDTH * 0.09, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.11, CONTENT_WIDTH * 0.38],
            )
        )
    append_reading_note(story, "Leviers limites au modele courant.")
    H1(story, "Fiche performance modele")
    rows_perf = [
        ["Ventes", str(ref_res.sales)],
        ["CA", f"{fmt_money(ref_res.revenue)} $"],
        ["Profit", f"{fmt_money(ref_res.profit)} $"],
        ["Marge", fmt_pct(ref_res.margin)],
        ["PDM globale", fmt_pct(ref_res.market_share)],
        ["PDM segment", fmt_pct(ref_res.market_share_segment)],
        ["Service", fmt_pct(ref_res.service_rate, 0)],
        ["Stock final", str(ref_res.forecast_ending_stock_units)],
    ]
    story.append(table_standard(["KPI", "Valeur"], rows_perf, [CONTENT_WIDTH * 0.40, CONTENT_WIDTH * 0.60]))
    H1(story, "Alertes modele")
    alerts = bundle.get("alerts", [])
    rows_alerts = [[a.level.upper(), a.code, a.message] for a in alerts] if alerts else [["INFO", "-", "Aucune alerte priorisee."]]
    story.append(
        table_standard(
            ["Niveau", "Code", "Message"],
            rows_alerts,
            [CONTENT_WIDTH * 0.12, CONTENT_WIDTH * 0.20, CONTENT_WIDTH * 0.68],
        )
    )
    append_reading_note(story, "Alertes issues de engine.rapport_alertes.")
    H1(story, "Recommandations modele")
    rec_rows = [[r.action, r.impact, r.horizon, r.priority] for r in fiche["recommandations"]]
    story.append(
        table_standard(
            ["Action", "Impact", "Horizon", "Priorite"],
            rec_rows,
            [CONTENT_WIDTH * 0.35, CONTENT_WIDTH * 0.33, CONTENT_WIDTH * 0.14, CONTENT_WIDTH * 0.18],
        )
    )


def generer_pdf_rapport(
    out: Path,
    code_firme: str,
    vue_firme: dict[str, Any],
    fiches_modeles: dict[str, Any],
    recommandations: list[Any],
    donnees: dict[str, Any],
) -> None:
    story: list = []
    _append_cover_exec_toc(story, code_firme, vue_firme, fiches_modeles)
    part_style = ParagraphStyle(
        "PART_FIRM",
        parent=P_INS,
        fontSize=12,
        leading=14,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        textColor=NAVY,
    )
    story.append(Paragraph("PARTIE VUE FIRME", part_style))
    story.append(Spacer(1, 6))
    _append_section_11_dashboard(story, vue_firme["firm_results"], code_firme)
    _append_sections_12_14_context(story, donnees["wb"], code_firme)
    _append_cross_matrix(story, vue_firme["cross_matrix_pct"])
    _append_section_15_portfolio(story, donnees["wb"], donnees["produits"], donnees["product_bundles"])
    _append_section_16_levers(story, donnees["mm_rows"], vue_firme["firm_results"])
    _append_section_17_reference(story, vue_firme)
    _append_section_110_reco(story, recommandations)
    for _, fiche in fiches_modeles.items():
        _append_model_part(story, donnees["wb"], code_firme, donnees["mm_rows"], fiche)
    out = out.expanduser()
    if not out.is_absolute():
        out = Path.cwd() / out
    out.parent.mkdir(parents=True, exist_ok=True)
    footer = f"Rapport VAE - {vue_firme['legend']} - usage interne"
    build_doc(
        out,
        story,
        firm_code=code_firme,
        left_text=footer,
        report_date=date.today().isoformat(),
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="Generateur rapport PDF VAE moderne")
    ap.add_argument("--all-firms", action="store_true", help="Generer un rapport pour chaque firme.")
    ap.add_argument("--firme", "--firm", default="TRE")
    ap.add_argument("--workbook", type=Path, default=None)
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()

    xlsx = args.workbook.expanduser() if args.workbook else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Classeur introuvable : {xlsx}")

    if args.all_firms:
        for firm in sorted(MARKET_CONFIG["firms"].keys()):
            donnees = build_report_payload(xlsx, firm)
            out = Path(str(OUT_PDF).format(firm=firm, date=date.today().isoformat()))
            generer_pdf_rapport(
                out,
                firm,
                donnees["vue_firme"],
                donnees["fiches_modeles"],
                donnees["recommandations_firme"],
                donnees,
            )
            try:
                shutil.copy2(out, DOWNLOADS / out.name)
            except OSError:
                pass
            finally:
                donnees["wb"].close()
            print(f"[OK] PDF : {out}")
        return

    firm = resolve_firm_code(args.firme)
    if firm not in MARKET_CONFIG["firms"]:
        raise SystemExit(f"Firme inconnue : {firm!r}")
    donnees = build_report_payload(xlsx, firm)
    out = args.out or Path(str(OUT_PDF).format(firm=firm, date=date.today().isoformat()))
    generer_pdf_rapport(
        out,
        firm,
        donnees["vue_firme"],
        donnees["fiches_modeles"],
        donnees["recommandations_firme"],
        donnees,
    )
    try:
        shutil.copy2(out, DOWNLOADS / out.name)
    except OSError:
        pass
    finally:
        donnees["wb"].close()
    print(f"[OK] PDF : {out}")


if __name__ == "__main__":
    main()
