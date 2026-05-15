#!/usr/bin/env python3
"""
Export Excel du rapport VAE — une feuille par tableau (openpyxl).
Usage : python3 scripts/generate_excel.py --firme TRE --out reports/Rapport_TRE_tables.xlsx
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from generate_rapport_vae_style_pdf import (  # noqa: E402
    RANGE_MAP,
    build_report_payload,
    firm_adjusted_budget_from_scenario,
    firm_scenario_results,
    fmt_pct_and_money,
    growth_rows,
    load_firms_market_rows,
    load_mm_rows,
    scenario_display_title,
    segment_label_from_idx,
    simulate,
)
from engine.budget_allocation import snap_rd_firm_pct  # noqa: E402
from vae_report_tables import (  # noqa: E402
    FIRM_ONLINE_PCT,
    FIRM_ONLINE_TREND,
    canonical_gamme_label,
    competitor_reaction_blurb,
    portfolio_anomaly_note,
    production_band_note,
    top3_channels_cell,
)
from excel_2250_reference import enrich_workbook_product_scenarios  # noqa: E402
from vae_rapport_firme_logic import default_excel_path, resolve_firm_code  # noqa: E402
from config.market_config import MARKET_CONFIG  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1A82D4")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
NEG_FONT = Font(name="Arial", size=10, color="CC0000")
ALT_FILL = PatternFill("solid", fgColor="F0F0F0")


def _write_sheet(ws, headers: list[str], rows: list[list], *, pct_cols: set[int] | None = None) -> None:
    pct_cols = pct_cols or set()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(ri, ci, val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(val, (int, float)) and val < 0:
                cell.font = NEG_FONT
            if ri % 2 == 0:
                cell.fill = ALT_FILL
            if ci - 1 in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            elif isinstance(val, (int, float)) and ci - 1 not in pct_cols:
                cell.number_format = "#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(48, 10 + 1.2 * len(str(headers[c - 1])))


def _budget_rows_for_scenario(ref_scen, ref_res) -> list[list[str]]:
    adj_firm = firm_adjusted_budget_from_scenario(ref_scen)
    firm_mkt = float(ref_scen.firm_marketing_budget_total)
    firm_rd = float(ref_scen.firm_rd_budget_total)
    mkt_ratio = firm_mkt / adj_firm
    rd_ratio = snap_rd_firm_pct(firm_rd / adj_firm if adj_firm > 0 else 0.0)
    wgt = float(ref_scen.allocation_weight)
    return [
        ["Budget marketing (firme)", fmt_pct_and_money(mkt_ratio, firm_mkt)],
        ["Budget R&D (firme)", fmt_pct_and_money(rd_ratio, firm_rd)],
        ["Poids stratégique du modèle", f"{wgt * 100:.1f} %".replace(".", ",")],
        ["Contribution profit (modèle)", f"{ref_res.profit:,.2f}".replace(",", " ").replace(".", ",") + " $"],
        ["Marge (modèle)", f"{ref_res.margin * 100:.2f} %".replace(".", ",")],
    ]


def main() -> None:
    ap = argparse.ArgumentParser(description="Export Excel des tableaux du rapport VAE")
    ap.add_argument("--firme", "--firm", default="TRE")
    ap.add_argument("--workbook", type=Path, default=None)
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Chemin .xlsx de sortie (défaut : reports/Export_{firme}_{date}.xlsx)",
    )
    args = ap.parse_args()

    firm = resolve_firm_code(getattr(args, "firme", None) or args.firm)
    xlsx = args.workbook.expanduser() if args.workbook else default_excel_path()
    if not xlsx.exists():
        raise SystemExit(f"Classeur introuvable : {xlsx}")

    donnees = build_report_payload(xlsx, firm)
    wb_x = donnees["wb"]
    prods = donnees["produits"]
    mm_rows = donnees["mm_rows"]
    results = donnees["results"]
    product_bundles = donnees["product_bundles"]
    vue = donnees["vue_firme"]
    firm_results = vue.get("firm_results") or firm_scenario_results(results, firm)
    ref_scen = vue["ref_scen"]
    ref_res = vue["ref_res"]
    ref_sid = vue["ref_sid"]
    ref_prod = next((p for p in prods if p.model_name == ref_scen.model_name), donnees["focal_prod"])
    recs = donnees["recommandations_firme"]
    cross = donnees["cross_matrix_pct"]

    wb = Workbook()
    wb.remove(wb.active)

    # 1.1 Scénarios (+ note réactions concurrentes)
    ws = wb.create_sheet("1.1 Scénarios", 0)
    h1 = [
        "Scénario",
        "Firme",
        "Période",
        "Ventes",
        "CA",
        "Profit",
        "Marge",
        "PDM",
        "Service",
        "Note réactions concurrents",
    ]
    r1: list[list] = []
    for sid, owner, scen, res in firm_results:
        r1.append(
            [
                scenario_display_title(sid),
                owner,
                scen.period,
                res.sales,
                res.revenue,
                res.profit,
                res.margin,
                res.market_share,
                res.service_rate,
                competitor_reaction_blurb(sid, firm),
            ]
        )
    _write_sheet(ws, h1, r1, pct_cols={6, 7, 8})

    # 1.2 PDM Firmes
    ws = wb.create_sheet("1.2 PDM Firmes")
    h2 = ["Firme", "Unités réf.", "PDM", "Vente en ligne (%)", "Évolution vente en ligne"]
    r2 = []
    for code, units, share in load_firms_market_rows(wb_x):
        oc = FIRM_ONLINE_PCT.get(str(code), 0.0) / 100.0
        r2.append([code, units, share, oc, FIRM_ONLINE_TREND.get(str(code), "—")])
    _write_sheet(ws, h2, r2, pct_cols={2, 3})

    # 1.3 Segments (sans Prix préf.)
    ws = wb.create_sheet("1.3 Segments")
    h3 = ["Segment", "Description", "Part", "Unités réf.", "Online"]
    r3 = []
    ws_s = wb_x["SEGMENTS"]
    for row in range(2, ws_s.max_row + 1):
        if ws_s.cell(row, 1).value is None:
            continue
        r3.append(
            [
                ws_s.cell(row, 1).value,
                ws_s.cell(row, 2).value,
                float(ws_s.cell(row, 3).value or 0),
                float(ws_s.cell(row, 5).value or 0),
                float(ws_s.cell(row, 6).value or 0),
            ]
        )
    _write_sheet(ws, h3, r3, pct_cols={2, 4})

    # 1.4 Croissance
    ws = wb.create_sheet("1.4 Croissance")
    gr, seg_col = growth_rows(firm)
    h4 = ["Période", "Année", "Marché total", seg_col, "Croiss. vs P1"]
    r4 = [[a, b, c, d, e] for a, b, c, d, e in gr]
    _write_sheet(ws, h4, r4)

    # 1.5 Portefeuille
    ws = wb.create_sheet("1.5 Portefeuille")
    h5 = ["Code", "Nom modèle", "Segment", "Gamme", "Prix réf.", "Unités réf.", "Anomalie / note"]
    r5 = []
    for p in prods:
        note = portfolio_anomaly_note(p.product_key)
        if not note:
            note = production_band_note(float(p.units))
        r5.append(
            [
                p.product_key,
                p.model_name,
                f"{p.segment_idx} — {segment_label_from_idx(wb_x, p.segment_idx)}",
                canonical_gamme_label(p.range_raw),
                float(p.base_price),
                float(p.units),
                note or "",
            ]
        )
    _write_sheet(ws, h5, r5)

    # 1.6 Marketing
    ws = wb.create_sheet("1.6 Marketing")
    h6 = [
        "Segment",
        "Digital",
        "Social",
        "Influenceur",
        "Affichage",
        "Événements",
        "3 canaux min. (Top 3 recommandés)",
    ]
    r6 = []
    for row in mm_rows:
        vals: list = []
        for j in range(1, 6):
            try:
                vals.append(float(str(row[j]).replace(",", ".")))
            except (TypeError, ValueError):
                vals.append(row[j])
        r6.append([row[0]] + vals + [top3_channels_cell(row)])
    _write_sheet(ws, h6, r6)

    # 3.1 Budgets firme (scénario de référence)
    ws = wb.create_sheet("3.1 Budgets firme")
    rows_b = [["Indicateur", "Valeur"]] + _budget_rows_for_scenario(ref_scen, ref_res)
    _write_sheet(ws, rows_b[0], rows_b[1:])

    # Performance par modèle
    ws = wb.create_sheet("Performance modèles")
    hp = [
        "Modèle",
        "Statut",
        "PDM glob.",
        "PDM seg.",
        "Marge",
        "Service",
        "Stock dép.",
        "Production",
        "Ventes",
        "Stock fin.",
        "Prix cat.",
        "Promo",
        "Prix promo",
    ]
    rp = []
    for p in prods:
        b = product_bundles[p.product_key]
        sc = b["ref_scen"]
        rs = b["ref_res"]
        promo_price = sc.price * (1.0 + sc.promotion_rate)
        rp.append(
            [
                p.product_key,
                "Actif",
                rs.market_share,
                rs.market_share_segment,
                rs.margin,
                rs.service_rate,
                sc.opening_stock,
                sc.production,
                rs.sales,
                rs.forecast_ending_stock_units,
                sc.price,
                sc.promotion_rate,
                promo_price,
            ]
        )
    _write_sheet(ws, hp, rp, pct_cols={2, 3, 4, 5, 11})

    # Projection volumes
    ws = wb.create_sheet("Projection volumes")
    hproj = ["Production", "Ventes", "Stock final", "Service", "Alerte"]
    rproj = []
    for u in (1500, 2250, 3000):
        s = ref_scen.model_copy(update={"production": int(u), "scenario_name": f"{firm}_proj_{u}"})
        r = simulate(s)
        rproj.append(
            [
                u,
                r.sales,
                r.forecast_ending_stock_units,
                r.service_rate,
                "OK" if r.service_rate >= 0.85 else "Sous seuil 85 %",
            ]
        )
    _write_sheet(ws, hproj, rproj, pct_cols={3})

    # 1.7 Paramètres Réf
    ws = wb.create_sheet("1.7 Paramètres Réf")
    rev = max(ref_res.revenue, 1.0)
    base_rev = max(ref_res.base_revenue_before_premium or ref_res.revenue, 1.0)
    rng_key = RANGE_MAP.get(str(ref_prod.range_raw), "mid")
    uref = max(float(ref_prod.units), 1e-6)
    cog_ratio = MARKET_CONFIG["cogs_ratios"][rng_key]
    dist_rate = MARKET_CONFIG["ranges"][rng_key]["distribution_rate"]
    rows_p = [
        ["Paramètre", "Valeur"],
        ["Scénario (id)", ref_sid],
        ["Titre affiché", scenario_display_title(ref_sid)],
        ["Modèle", ref_prod.product_key],
        ["Prix catalogue ($)", ref_scen.price],
        ["Promotion (ratio)", ref_scen.promotion_rate],
        ["Production (u.)", ref_scen.production],
        ["Marketing ($)", ref_scen.marketing_budget],
        ["R&D ($)", ref_scen.rd_budget],
        ["CA simulé ($)", ref_res.revenue],
        ["Marge (ratio)", ref_res.margin],
        ["COGS / CA", ref_res.production_cost / rev],
        ["Distribution / CA hors prime", ref_res.distribution_cost / base_rev],
        ["PDM globale", ref_res.market_share],
        ["Théorie COGS gamme (ratio)", cog_ratio],
        ["Théorie distribution gamme (ratio)", dist_rate],
        ["Production / unités réf.", ref_scen.production / uref],
    ]
    _write_sheet(ws, rows_p[0], rows_p[1:])

    # Tableau Croisé
    ws = wb.create_sheet("Tableau Croisé")
    firms = cross["firms"]
    segs = cross["segments"]
    grid = cross["pct_grid"]
    rt = cross["row_totals_pct"]
    ct = cross["col_totals_pct"]
    gtot = cross["grand_total_pct"]
    hdr = ["Firme"] + [f"S{j+1}" for j in range(len(segs))] + ["TOTAL"]
    rcr = []
    for i, f in enumerate(firms):
        rcr.append([f] + [grid[i][j] / 100.0 for j in range(len(segs))] + [rt[i] / 100.0])
    rcr.append(["TOTAL"] + [ct[j] / 100.0 for j in range(len(segs))] + [gtot / 100.0])
    _write_sheet(ws, hdr, rcr, pct_cols=set(range(1, len(hdr) + 1)))

    # 1.10 Recommandations
    ws = wb.create_sheet("1.10 Exemples reco")
    h10 = ["Action", "Impact attendu", "Horizon", "Priorité"]
    r10 = [[r.action, r.impact, r.horizon, r.priority] for r in recs]
    _write_sheet(ws, h10, r10)

    enrich_workbook_product_scenarios(wb, firm)

    out = args.out
    if out is None:
        out = _ROOT / "reports" / f"Export_{firm}_{date.today().isoformat()}.xlsx"
    else:
        out = out.expanduser()
    if not out.is_absolute():
        out = Path.cwd() / out
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    wb_x.close()
    print(f"[OK] Excel : {out}")


if __name__ == "__main__":
    main()
