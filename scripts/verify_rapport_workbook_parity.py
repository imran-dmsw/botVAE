#!/usr/bin/env python3
"""Vérifie la concordance classeur Excel ↔ sources du rapport stratégique VAE."""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

from config.market_config import MARKET_CONFIG
from engine.simulation import total_market_size
from generate_rapport_vae_style_pdf import (
    load_firm_share,
    load_firms_market_rows,
    load_mm_rows,
    load_products,
)
from vae_rapport_firme_logic import FIRMES_VALIDES, PRICE_FLOORS, default_excel_path, resolve_firm_code


@dataclass
class CheckResult:
    name: str
    ok: bool
    detail: str


def _param_val(wb: openpyxl.Workbook, key: str) -> float | None:
    ws = wb["PARAM"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == key:
            raw = ws.cell(r, 2).value
            return float(raw) if raw is not None else None
    return None


def verify_workbook_parity(xlsx: Path, firm: str) -> list[CheckResult]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    checks: list[CheckResult] = []

    try:
        firms_rows = load_firms_market_rows(wb)
        checks.append(
            CheckResult(
                "FIRMS — 9 firmes",
                len(firms_rows) == 9,
                f"{len(firms_rows)} ligne(s) agrégée(s)",
            )
        )
        share_sum = sum(share for _, _, share in firms_rows)
        checks.append(
            CheckResult(
                "FIRMS — parts de marché",
                abs(share_sum - 1.0) < 0.02,
                f"somme Share_RefYear = {share_sum:.4f}",
            )
        )
        pdm = load_firm_share(wb, firm)
        checks.append(
            CheckResult(
                f"FIRMS — PDM {firm}",
                pdm > 0,
                f"Share_RefYear = {pdm:.4%}",
            )
        )

        ws_s = wb["SEGMENTS"]
        seg_count = sum(1 for r in range(2, ws_s.max_row + 1) if ws_s.cell(r, 1).value is not None)
        checks.append(
            CheckResult(
                "SEGMENTS — 6 segments",
                seg_count == 6,
                f"{seg_count} segment(s)",
            )
        )

        mm = load_mm_rows(wb)
        checks.append(
            CheckResult(
                "MARKETING_MATRIX — 6 lignes",
                len(mm) == 6,
                f"{len(mm)} ligne(s)",
            )
        )

        prods = load_products(wb, firm)
        checks.append(
            CheckResult(
                f"BASE_REFERENCE_MODEL — catalogue {firm}",
                len(prods) > 0,
                f"{len(prods)} modèle(s) : {', '.join(p.product_key for p in prods)}",
            )
        )

        base_units = _param_val(wb, "Base_Market_Units")
        growth = _param_val(wb, "Market_Growth")
        checks.append(
            CheckResult(
                "PARAM ↔ market_config — taille marché",
                base_units == MARKET_CONFIG["base_market_size"],
                f"Excel {base_units} vs config {MARKET_CONFIG['base_market_size']}",
            )
        )
        checks.append(
            CheckResult(
                "PARAM ↔ market_config — croissance",
                growth is not None and abs(growth - MARKET_CONFIG["growth_rate"]) < 1e-9,
                f"Excel {growth} vs config {MARKET_CONFIG['growth_rate']}",
            )
        )

        p1_market = total_market_size(1)
        expected_p1 = int(round((base_units or 0) * (1 + (growth or 0))))
        checks.append(
            CheckResult(
                "Croissance P1 (moteur)",
                abs(p1_market - expected_p1) <= 1,
                f"P1 = {p1_market:,} (attendu ≈ {expected_p1:,})",
            )
        )

        cogs_checks = [
            ("COGS_Bas", "entry", 0.60),
            ("COGS_Moyen", "mid", 0.57),
            ("COGS_Haut", "premium", 0.55),
        ]
        for param_key, range_key, fallback in cogs_checks:
            excel_val = _param_val(wb, param_key)
            cfg_val = MARKET_CONFIG["cogs_ratios"][range_key]
            checks.append(
                CheckResult(
                    f"PARAM ↔ COGS {range_key}",
                    excel_val is not None and abs(excel_val - cfg_val) < 1e-9,
                    f"Excel {excel_val} vs config {cfg_val}",
                )
            )

        floor_labels = sorted({p.range_raw for p in prods})
        missing_floors = [g for g in floor_labels if g not in PRICE_FLOORS and g.lower() not in {k.lower() for k in PRICE_FLOORS}]
        checks.append(
            CheckResult(
                "Planchers prix — gammes catalogue",
                not missing_floors,
                "OK" if not missing_floors else f"sans plancher : {', '.join(missing_floors)}",
            )
        )

        checks.append(
            CheckResult(
                "Périmètre rapport",
                True,
                "KPI simulés = moteur Python (pas MODEL_PERIOD / INPUT_FIRM). "
                "Contrôles saisie : run_tre_p1_e2e_test + generate_tre_p1_test_pdf.",
            )
        )
    finally:
        wb.close()

    return checks


def main() -> None:
    parser = argparse.ArgumentParser(description="Checklist Excel ↔ rapport stratégique VAE")
    parser.add_argument("--firme", "--firm", dest="firme", required=True)
    parser.add_argument("--workbook", type=Path, default=None)
    args = parser.parse_args()

    firm = resolve_firm_code(args.firme)
    if firm not in FIRMES_VALIDES:
        print(f"[ERREUR] Firme inconnue : {args.firme}")
        sys.exit(1)

    xlsx = args.workbook.expanduser() if args.workbook else default_excel_path()
    if not xlsx.exists():
        print(f"[ERREUR] Classeur introuvable : {xlsx}")
        sys.exit(1)

    print(f"[INFO] Classeur : {xlsx}")
    print(f"[INFO] Firme   : {firm}")
    results = verify_workbook_parity(xlsx, firm)
    failed = 0
    for item in results:
        tag = "OK" if item.ok else "ECHEC"
        if not item.ok:
            failed += 1
        print(f"[{tag}] {item.name} — {item.detail}")

    if failed:
        print(f"\n[ECHEC] {failed} contrôle(s) en échec.")
        sys.exit(1)
    print("\n[OK] Concordance référentiel Excel / sources rapport validée.")


if __name__ == "__main__":
    main()
