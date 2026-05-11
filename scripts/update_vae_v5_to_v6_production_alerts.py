"""
V5 -> V6 : seuils production/stockage (MODEL_PERIOD.TauxCouverture, CONTROLES,
CONTROLE_DECISIONS messages Production, TESTS_PRODUCTION, FORMULES_FR).
Ne modifie pas PARAM, SEGMENTS, FIRMS, MARKETING_MATRIX, BASE_REFERENCE_MODEL.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v5.xlsx")
DST = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v6.xlsx")

TITLE_FILL = PatternFill("solid", fgColor="1C2B4A")
TITLE_FONT = Font(color="FFFFFF", bold=True)

CONTROLES_BLOCK_START = 315  # 1-based row


def alerte_couverture_formula(e_col: str, row: int) -> str:
    c = f"{e_col}{row}"
    return (
        f'=IF({c}<0.9,"🔴 Alerte rupture",'
        f'IF({c}<1,"🟠 Production prudente",'
        f'IF({c}<=1.1,"🟢 Production équilibrée",'
        f'IF({c}<=1.2,"🟡 Production sécuritaire","🔴 Alerte surproduction"))))'
    )


def patch_model_period_taux(ws_mp):
    last_col = ws_mp.max_column + 1
    ws_mp.cell(1, last_col, "TauxCouverture")
    for r in range(2, ws_mp.max_row + 1):
        ws_mp.cell(r, last_col, f"=IF(U{r}>0,(S{r}+T{r})/U{r},1)")


def patch_controles_block_fixed(wb):
    """Version sans helper cassé — calcule lettre colonne TauxCouverture après ajout."""
    ws_c = wb["CONTROLES"]
    ws_mp = wb["MODEL_PERIOD"]
    tc_col = get_column_letter(ws_mp.max_column)

    row_title = CONTROLES_BLOCK_START
    ws_c.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=12)
    cell = ws_c.cell(row_title, 1, "Seuils production et stockage (une ligne par ligne MODEL_PERIOD)")
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")

    hdr = row_title + 1
    headers = [
        "Idx_MP",
        "Period_Index",
        "Firm",
        "ProductKey",
        "TauxCouverture",
        "Alerte_Couverture",
        "Ventes_perdues_est.",
        "Alerte_VentesPerdues",
        "Alerte_Surstock",
        "Alerte_CoutStockage",
    ]
    for i, h in enumerate(headers, 1):
        ws_c.cell(hdr, i, h)

    first_data = hdr + 1
    for i in range(616):
        mr = 2 + i
        rr = first_data + i
        ws_c.cell(rr, 1, i + 1)
        ws_c.cell(rr, 2, f"=MODEL_PERIOD!A{mr}")
        ws_c.cell(rr, 3, f"=MODEL_PERIOD!D{mr}")
        ws_c.cell(rr, 4, f"=MODEL_PERIOD!C{mr}")
        ws_c.cell(rr, 5, f"=MODEL_PERIOD!{tc_col}{mr}")
        ws_c.cell(rr, 6, alerte_couverture_formula("E", rr))
        ws_c.cell(rr, 7, f"=MAX(0,MODEL_PERIOD!U{mr}-(MODEL_PERIOD!S{mr}+MODEL_PERIOD!T{mr}))")
        ws_c.cell(
            rr,
            8,
            f'=IF(MODEL_PERIOD!U{mr}=0,"✓ OK",IF(G{rr}/MODEL_PERIOD!U{mr}>0.1,'
            f'"🔴 Ventes perdues > 10 % — risque de perte de PDM","✓ OK"))',
        )
        ws_c.cell(
            rr,
            9,
            f'=IF(MODEL_PERIOD!U{mr}=0,"✓ OK",IF(MODEL_PERIOD!AE{mr}/MODEL_PERIOD!U{mr}>0.2,'
            f'"🔴 Alerte surstock : production dépasse fortement la demande prévue",'
            f'IF(MODEL_PERIOD!AE{mr}/MODEL_PERIOD!U{mr}>0.1,'
            f'"🟡 Attention : risque de surstockage et coûts de stockage élevés","✓ OK")))',
        )
        ws_c.cell(
            rr,
            10,
            f'=IF(MODEL_PERIOD!AI{mr}>(MODEL_PERIOD!AD{mr}*MODEL_PERIOD!M{mr}*(1-MODEL_PERIOD!AF{mr}))*0.05,'
            f'"🟠 Attention : le stockage risque de réduire fortement votre profit (> 5 %)","✓ OK")',
        )


def production_message_g(sr: int) -> str:
    """Colonne G CONTROLE_DECISIONS pour ligne Production (agrégée firme/période)."""
    dem = "SUMIFS(MODEL_PERIOD!$U:$U,MODEL_PERIOD!$A:$A,$B{sr},MODEL_PERIOD!$D:$D,$A{sr})"
    dem = dem.format(sr=sr)
    return (
        f'=IF(F{sr}="✗ Bloqué","Production interdite pendant liquidation",'
        f'IF(D{sr}=0,"Production à zéro — modèle non actif ou liquidation",'
        f'IF(D{sr}<{dem}*0.85,'
        f'"🔴 Sous-production critique — ventes perdues estimées : "&TEXT(MAX(0,{dem}-D{sr}),"# ##0")&" u.",'
        f'IF(D{sr}<{dem}*0.9,'
        f'"🟠 Production prudente — fourchette recommandée : "&TEXT(ROUND({dem}*0.9,0),"# ##0")&" à "&TEXT(ROUND({dem}*1.1,0),"# ##0")&" u.",'
        f'IF(D{sr}<={dem}*1.1,'
        f'"🟢 Production dans la zone cible — stock estimé : "&TEXT(MAX(0,D{sr}-{dem}),"# ##0")&" u.",'
        f'IF(D{sr}<={dem}*1.2,'
        f'"🟡 Production sécuritaire — stock tampon estimé : "&TEXT(D{sr}-{dem},"# ##0")&" u.",'
        f'"🔴 Surproduction — stock final estimé : "&TEXT(D{sr}-{dem},"# ##0")&" u. — coût de stockage important")))))'
    )


def patch_controle_decisions(ws_cd):
    for sr in range(2, ws_cd.max_row + 1):
        if ws_cd.cell(sr, 3).value != "Production":
            continue
        ws_cd.cell(sr, 5, "Zone décision ±10 % / ±15 % vs demande agrégée (MODEL_PERIOD)")
        ws_cd.cell(sr, 7, production_message_g(sr))


def patch_tests_production_fix(ws):
    """Colonne J = part A/B (nombre); colonne K compare seuils sur valeur décimale (=J)."""
    h2 = [
        "Production",
        "DemandActive",
        "Ventes",
        "Stock final",
        "Taux service",
        "Cout stockage",
        "Profit estime",
        "Stock tampon",
        "Production recommandee",
        "Taux couverture",
        "Zone",
    ]
    for i, h in enumerate(h2, 1):
        ws.cell(2, i, h)

    for r in range(3, 7):
        ws.cell(r, 9, f"=CEILING(B{r}*1.05,100)")
        ws.cell(r, 10, f"=IF(B{r}>0,A{r}/B{r},0)")
        ws.cell(r, 10).number_format = "0.00%"
        ws.cell(
            r,
            11,
            f'=IF(J{r}<0.9,"🔴 Alerte rupture",IF(J{r}<1,"🟠 Production prudente",IF(J{r}<=1.1,"🟢 Production équilibrée",IF(J{r}<=1.2,"🟡 Production sécuritaire","🔴 Alerte surproduction"))))',
        )

    reminder_row = 31
    ws.cell(reminder_row, 1, "Rappel pédagogique :")
    ws.cell(reminder_row, 1).font = Font(bold=True)
    lines = [
        "- Produire trop peu = ventes perdues et perte de PDM.",
        "- Produire trop = coût de stockage (2,5 % × stock final × coût unitaire).",
        "- Zone idéale : taux de couverture entre 100 % et 110 %.",
        "- Si fort investissement marketing → viser le haut de la fourchette (+10 %).",
        "- Si investissement marketing faible → viser la fourchette basse (90–100 %).",
        "- Le stock de la période précédente est inclus automatiquement dans StockStart.",
    ]
    for k, line in enumerate(lines, 1):
        ws.cell(reminder_row + k, 1, line)


def patch_formules_fr(ws):
    start = ws.max_row + 1
    rows = [
        (
            25,
            "MODEL_PERIOD",
            "Taux de couverture",
            "TauxCouverture = (StockStart + Production) / DemandActive ; zone idéale 100 % à 110 %",
            "< 90 % alerte rupture ; > 120 % alerte surproduction (coût stockage)",
        ),
        (
            26,
            "CONTROLE_DECISIONS",
            "Zone de décision production",
            "Fourchette base DemandActive × [90 % ; 110 %] ; élargie [85 % ; 120 %]",
            "Marketing fort → haut de fourchette (+5 % à +10 %) ; marketing faible → bas (−5 % à −10 %)",
        ),
        (
            27,
            "MODEL_PERIOD / CONTROLES",
            "Alertes surstock et stockage",
            "Alerte modérée StockEnd/DemandActive > 10 % ; forte > 20 %",
            "Alerte coût : StockCost / (TotalSales×FinalPrice×(1-UnitCostRatio)) > 5 %",
        ),
    ]
    for i, (num, src, title, d, e) in enumerate(rows):
        r = start + i
        ws.cell(r, 1, num)
        ws.cell(r, 2, src)
        ws.cell(r, 3, title)
        ws.cell(r, 4, d)
        ws.cell(r, 5, e)


def main():
    if not SRC.exists():
        raise SystemExit(f"Manquant : {SRC}")
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    patch_model_period_taux(wb["MODEL_PERIOD"])
    patch_controles_block_fixed(wb)
    patch_controle_decisions(wb["CONTROLE_DECISIONS"])
    patch_tests_production_fix(wb["TESTS_PRODUCTION"])
    patch_formules_fr(wb["FORMULES_FR"])

    wb.save(DST)
    print(f"v6 sauvegardé : {DST}")
    print("  MODEL_PERIOD.TauxCouverture ; CONTROLES bloc seuils ; CONTROLE_DECISIONS Production ;")
    print("  TESTS_PRODUCTION ; FORMULES_FR étapes 25–27")

    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from update_vae_simulation_excel import add_report_function_and_generate

    wb_r = openpyxl.load_workbook(DST)
    add_report_function_and_generate(wb_r)
    wb_r.save(DST)
    print("  RAPPORT_*_P1 régénérés (Taux_couverture, Zone_production, Production_recommandee)")


if __name__ == "__main__":
    main()
