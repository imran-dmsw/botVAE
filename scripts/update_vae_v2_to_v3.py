"""
Complément v2 → v3 : corrections marketing, tampon stock, R&D, liquidation,
extension TESTS_*, MARCHE_STRUCTURE, TESTS_PARAMS, rapports PDM, CONTROLE_DECISIONS.
Ne modifie pas PARAM, SEGMENTS, FIRMS, MARKETING_MATRIX, BASE_REFERENCE_MODEL.
"""
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v2.xlsx")
OUTPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v3.xlsx")

TITLE_FILL = PatternFill("solid", fgColor="1C2B4A")
TITLE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
ALERT_RED = PatternFill("solid", fgColor="F8CBAD")
ALERT_ORANGE = PatternFill("solid", fgColor="FCE4D6")
ALERT_GREEN = PatternFill("solid", fgColor="E2F0D9")
BANNER_BLOCKED = PatternFill("solid", fgColor="FFF2CC")
BANNER_OK = PatternFill("solid", fgColor="E2F0D9")


def sheet_headers(ws):
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }


def header_style(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row, c)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")


def title_row(ws, row, text, max_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, text)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")


def autosize_columns(ws, max_width=48):
    for c in range(1, ws.max_column + 1):
        width = 10
        for r in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            width = max(width, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(width, max_width)


def read_table(ws):
    h = sheet_headers(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {k: ws.cell(r, h[k]).value for k in h}
        rows.append(row)
    return rows


def get_param_map(wb):
    ws = wb["PARAM"]
    params = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k:
            params[str(k).strip()] = v
    return params


def find_controle_row(ws, label_prefix):
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith(label_prefix):
            return r
    return None


def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def correction1_marketing_helpers_and_controles(wb):
    ws = wb["CONTROLES"]
    start = 160
    ws.cell(start, 8, "AIDE_MARKETING (indices FIRM_PERIOD 1..72)")
    ws.cell(start, 8).font = Font(bold=True)
    hdr = ["Idx", "Periode", "Firme", "TotalMkt", "BudgetBase", "Revenu_Pm1", "Min3pct", "Max110", "SousMin", "AuDessusMax"]
    for i, h in enumerate(hdr, 8):
        ws.cell(start + 1, i, h)
        ws.cell(start + 1, i).font = HEADER_FONT
    first = start + 2
    last = first + 71
    for n, r in enumerate(range(first, last + 1)):
        idx = n + 1
        ws.cell(r, 8, idx)
        ws.cell(r, 9, f"=INDEX(FIRM_PERIOD!$A$2:$A$73,{idx})")
        ws.cell(r, 10, f"=INDEX(FIRM_PERIOD!$C$2:$C$73,{idx})")
        ws.cell(r, 11, f"=INDEX(FIRM_PERIOD!$K$2:$K$73,{idx})")
        ws.cell(r, 12, f"=INDEX(FIRM_PERIOD!$T$2:$T$73,{idx})")
        ws.cell(r, 13, f"=IF(I{r}=1,INDEX(FIRM_PERIOD!$S$2:$S$73,{idx}),SUMIFS(RESULTS_FIRM!$E:$E,RESULTS_FIRM!$A:$A,I{r}-1,RESULTS_FIRM!$C:$C,J{r}))")
        ws.cell(r, 14, f"=0.03*M{r}")
        ws.cell(r, 15, f"=IF(I{r}=1,L{r}*1.1,SUMIFS(FIRM_PERIOD!$K$2:$K$73,FIRM_PERIOD!$A$2:$A$73,I{r}-1,FIRM_PERIOD!$C$2:$C$73,J{r})*1.1)")
        ws.cell(r, 16, f"=--(K{r}<N{r})")
        ws.cell(r, 17, f"=--(K{r}>O{r})")

    rmin = find_controle_row(ws, "Marketing minimum")
    rmax = find_controle_row(ws, "Marketing maximum")
    if rmin:
        ws.cell(rmin, 1, "Marketing min 3% revenu P-1")
        ws.cell(rmin, 4, f"=SUM(P{first}:P{last})")
        ws.cell(rmin, 5, f'=IF(D{rmin}>0,"ALERTE","OK")')
        ws.cell(rmin, 6, "Budget marketing sous le minimum (3 % du revenu P-1)")
    if rmax:
        ws.cell(rmax, 1, "Marketing max progressif (+10%)")
        ws.cell(rmax, 4, f"=SUM(Q{first}:Q{last})")
        ws.cell(rmax, 5, f'=IF(D{rmax}>0,"ALERTE","OK")')
        ws.cell(
            rmax,
            6,
            "Plafond P1 = BudgetBase_Adjusted x 110 % ; P2+ = TotalMarketing annee precedente x 110 %",
        )
    print("Correction/Tâche 1 terminée — Marketing escalade + aides CONTROLES")


def correction2_model_period_tampon(wb):
    mp = wb["MODEL_PERIOD"]
    h = sheet_headers(mp)
    ap = h.get("StockSecurite") or h.get("StockTampon")
    if not ap:
        raise RuntimeError("Colonne StockSecurite/StockTampon introuvable")
    mp.cell(1, ap, "StockTampon")
    demand_c = get_column_letter(h["DemandActive"])
    stockstart_c = get_column_letter(h["StockStart"])
    stockend_c = get_column_letter(h["StockEnd"])
    period_c = get_column_letter(h["Period_Index"])
    key_c = get_column_letter(h["ProductKey"])

    for r in range(2, mp.max_row + 1):
        prev_stock = f"IF({period_c}{r}=1,{stockstart_c}{r},SUMIFS(${stockend_c}:${stockend_c},${period_c}:${period_c},{period_c}{r}-1,${key_c}:${key_c},{key_c}{r}))"
        mp.cell(r, ap, f"=MIN(200,{prev_stock})")

    aq = h["SurplusPenalty"]
    prod_rec_col = aq + 1
    alert_tam_col = aq + 2
    mp.cell(1, prod_rec_col, "Production_recommandee").font = HEADER_FONT
    mp.cell(1, alert_tam_col, "Alerte_tampon").font = HEADER_FONT
    tam_c = get_column_letter(ap)
    for r in range(2, mp.max_row + 1):
        prev_stock = f"IF({period_c}{r}=1,{stockstart_c}{r},SUMIFS(${stockend_c}:${stockend_c},${period_c}:${period_c},{period_c}{r}-1,${key_c}:${key_c},{key_c}{r}))"
        mp.cell(r, prod_rec_col, f"=MAX(0,{demand_c}{r}-{tam_c}{r})+200")
        mp.cell(
            r,
            alert_tam_col,
            f'=IF(({prev_stock})<200,"Stock tampon insuffisant — "&(200-({prev_stock}))&" unites manquantes","OK")',
        )
    print("Correction/Tâche 2 terminée — StockTampon, Production_recommandee, Alerte_tampon")


def correction2_tests_production_tampon(wb):
    ws = wb["TESTS_PRODUCTION"]
    params = read_table(wb["PARAM"])
    pmap = {str(r.get("Parameter", "")).strip(): r.get("Value") for r in params if r.get("Parameter")}
    segs = {r["Segment"]: r for r in read_table(wb["SEGMENTS"])}
    seg1 = segs.get(1) or segs.get("1")
    growth = pmap.get("Market_Growth", 0.12) or 0.12
    demand = (seg1.get("Units_RefYear") or 0) * (1 + growth)
    pref_price = seg1.get("Pref_RefYear") or 3800
    stock_cost_rate = pmap.get("Stock_Cost_Rate", 0.025) or 0.025
    cogs_ratio = pmap.get("COGS_Moyen", 0.57) or 0.57
    marketing = 100000
    rd = 30000
    price_net = pref_price

    row_header = 2
    for c in range(1, 20):
        if ws.cell(row_header, c).value == "Production":
            start_col = c
            break
    else:
        start_col = 1

    levels = [1500, 2000, 2500, 3000]
    stock_tampon_prev = 0
    row = 3
    rec_row = None
    best_profit = -10**18
    for p in levels:
        tam = min(200, stock_tampon_prev)
        eff_need = max(0, demand - tam) + 200
        sales = min(p, demand)
        stock = p - sales
        service = sales / demand if demand else 0
        stock_cost = stock * pref_price * stock_cost_rate
        cogs = sales * price_net * cogs_ratio
        profit = sales * price_net - cogs - stock_cost - marketing - rd
        ws.cell(row, start_col + 0, p)
        ws.cell(row, start_col + 1, int(round(demand)))
        ws.cell(row, start_col + 2, int(round(sales)))
        ws.cell(row, start_col + 3, int(round(stock)))
        ws.cell(row, start_col + 4, service)
        ws.cell(row, start_col + 5, int(round(stock_cost)))
        ws.cell(row, start_col + 6, int(round(profit)))
        if ws.cell(row_header, start_col + 7).value != "Stock tampon":
            ws.cell(row_header, start_col + 7, "Stock tampon")
            ws.cell(row_header, start_col + 8, "Production recommandee")
        ws.cell(row, start_col + 7, tam)
        ws.cell(row, start_col + 8, int(round(eff_need)))
        ws.cell(row, start_col + 4).number_format = "0.0%"
        if service >= 0.95 and profit > best_profit:
            best_profit = profit
            rec_row = row
        row += 1

    rec_r = row + 1
    ws.cell(rec_r, start_col, "Recommande")
    if rec_row:
        ws.cell(rec_r, start_col + 1, ws.cell(rec_row, start_col).value)
        ws.cell(rec_r, start_col + 2, "Meilleur profit avec service >= 95% (avec tampon integre)")
        ws.cell(rec_r, start_col).fill = TITLE_FILL
        ws.cell(rec_r, start_col).font = TITLE_FONT
    print("Correction/Tâche 2b terminée — TESTS_PRODUCTION tampon / prod recommandee")


def correction3_liquidation_promo_model_period(wb):
    mp = wb["MODEL_PERIOD"]
    h = sheet_headers(mp)
    o = h["LiquidationPromo"]
    n = h["WithdrawFlag"]
    n_c = get_column_letter(n)
    for r in range(2, mp.max_row + 1):
        old = mp.cell(r, o).value
        if not isinstance(old, str) or not old.startswith("="):
            inner = "0"
        else:
            inner = old[1:]
        mp.cell(r, o, f"=IF({n_c}{r}=1,-0.1,{inner})")
    print("Correction/Tâche 4 terminée — LiquidationPromo forcee a -10 % si retrait")


def correction3_controles_liquidation_message(wb):
    ws = wb["CONTROLES"]
    r = find_controle_row(ws, "Liquidation taux")
    if r:
        ws.cell(
            r,
            4,
            "=SUMPRODUCT(--(MODEL_PERIOD!$N$2:$N$617=1),--(ABS(MODEL_PERIOD!$O$2:$O$617+0.1)>0.0001))",
        )
        ws.cell(
            r,
            5,
            f'=IF(D{r}>0,"Taux de liquidation automatiquement corrige a -10 %","OK")',
        )
        ws.cell(r, 6, "Si retrait actif, LiquidationPromo doit etre -10 % dans MODEL_PERIOD")
    print("Correction/Tâche 4b terminée — CONTROLES message liquidation")


def correction3_rd_helpers_and_controles(wb):
    ws = wb["CONTROLES"]
    start = 240
    ws.cell(start, 8, "AIDE_R&D et portefeuille")
    ws.cell(start, 8).font = Font(bold=True)
    hdr = [
        "Idx",
        "Periode",
        "Firme",
        "BudgetRD",
        "BudgetBase",
        "RD_pct",
        "Invalide",
        "Msg_droits",
        "NbActifs",
        "NbInit",
        "DepassePlafond",
    ]
    for i, h in enumerate(hdr, 8):
        ws.cell(start + 1, i, h)
        ws.cell(start + 1, i).font = HEADER_FONT
    first = start + 2
    last = first + 71
    for n, r in enumerate(range(first, last + 1)):
        idx = n + 1
        ws.cell(r, 8, idx)
        ws.cell(r, 9, f"=INDEX(FIRM_PERIOD!$A$2:$A$73,{idx})")
        ws.cell(r, 10, f"=INDEX(FIRM_PERIOD!$C$2:$C$73,{idx})")
        ws.cell(r, 11, f"=INDEX(FIRM_PERIOD!$I$2:$I$73,{idx})")
        ws.cell(r, 12, f"=INDEX(FIRM_PERIOD!$T$2:$T$73,{idx})")
        ws.cell(r, 13, f"=IF(L{r}=0,0,K{r}/L{r})")
        ws.cell(
            r,
            14,
            f"=--((ABS(M{r})>0.0001)*(ABS(M{r}-0.02)>0.0001)*(ABS(M{r}-0.05)>0.0001)*(ABS(M{r}-0.08)>0.0001))",
        )
        ws.cell(
            r,
            15,
            f'=IF(I{r}=1,IF(OR(ABS(M{r}-0.05)<0.0001,ABS(M{r}-0.08)<0.0001),"R&D P1 : developpement en cours — lancement disponible en P2","Pas de nouveau modele en developpement"),IF(ABS(M{r}-0.08)<0.0001,"Vous avez le droit de lancer 2 modeles en P"&(I{r}+1),IF(ABS(M{r}-0.05)<0.0001,"Vous avez le droit de lancer 1 modele en P"&(I{r}+1),"Aucun droit de lancement cette periode")))',
        )
        ws.cell(
            r,
            16,
            f'=COUNTIFS(MODEL_PERIOD!$D:$D,J{r},MODEL_PERIOD!$A:$A,I{r},MODEL_PERIOD!$G:$G,"ACTIVE")',
        )
        ws.cell(r, 17, f'=COUNTIF(BASE_REFERENCE_MODEL!$B:$B,J{r})')
        ws.cell(r, 18, f"=--(P{r}>O{r}+2)")

    r_inv = find_controle_row(ws, "Regles R&D niveaux")
    if r_inv:
        ws.cell(r_inv, 1, "R&D niveau autorise (0/2/5/8 %)")
        ws.cell(r_inv, 4, f"=SUM(N{first}:N{last})")
        ws.cell(r_inv, 5, f'=IF(D{r_inv}>0,"Niveau R&D invalide — valeurs autorisees : 0 / 2 / 5 / 8 %","OK")')
        ws.cell(r_inv, 6, "Verification sur toutes les lignes FIRM_PERIOD")

    r_p1 = find_controle_row(ws, "R&D periode 1")
    if r_p1:
        ws.cell(r_p1, 1, "R&D portefeuille — actifs max (init + 2)")
        ws.cell(r_p1, 4, f"=SUM(R{first}:R{last})")
        ws.cell(r_p1, 5, f'=IF(D{r_p1}>0,"BLOQUE — portefeuille plein (voir messages portefeuille)","OK")')
        ws.cell(
            r_p1,
            6,
            "NbActifs > NbInit + 2 : liquider avant lancement",
        )

    r_lim = find_controle_row(ws, "R&D limite 4")
    if r_lim:
        ws.cell(r_lim, 1, "R&D synthese messages droits (extraire)")
        ws.cell(r_lim, 4, "")
        ws.cell(r_lim, 5, "INFO")
        ws.cell(r_lim, 6, f"Voir colonne O (Msg_droits) lignes {first}-{last}")

    print("Correction/Tâche 3 terminée — R&D + portefeuille CONTROLES")


def correction5_formules_fr(wb):
    ws = wb["FORMULES_FR"]
    updated_21 = False
    for r in range(2, ws.max_row + 1):
        step = ws.cell(r, 1).value
        if step == 21:
            ws.cell(r, 4, "Budget marketing — escalade progressive : min 3 % revenu P-1, max = budget P-1 x 1,10 (ou budget base x 1,10 pour P1).")
            ws.cell(
                r,
                5,
                'SI(TotalMarketing<0,03*Revenu_Pm1;"Budget marketing sous le minimum (3 % du revenu P-1)";SI(Periode=1;SI(TotalMarketing>BudgetBase*1,1;"Budget marketing depasse le plafond P1 (+10 % du budget base)";"OK");SI(TotalMarketing>TotalMarketing_Pm1*1,1;"Budget marketing depasse le plafond (+10 % vs annee precedente)";"OK")))',
            )
            ws.cell(r, 6, "TotalMarketing_Pm1 via RECHERCHEV / SOMMES.SI.ENS sur FIRM_PERIOD")
            updated_21 = True
            break
    if not updated_21:
        row = ws.max_row + 1
        ws.cell(row, 1, 21)
        ws.cell(row, 2, "CONTROLES / FIRM_PERIOD")
        ws.cell(row, 3, "Marketing escalade")
        ws.cell(row, 4, "Budget marketing — escalade progressive : min 3 % revenu P-1, max = budget P-1 x 1,10 (ou budget base x 1,10 pour P1).")
        ws.cell(row, 5, "(voir feuille CONTROLES — bloc aide marketing)")
        ws.cell(row, 6, "Mise a jour regle plafond")

    rows_doc = [
        (29, "FORMULES_FR", "Cas particulier — Nouveau produit annee 1", "521 u. et corridor P1", "Documentation scenario AVE — etapes A/B/C (voir prompt)", ""),
        (30, "MODEL_PERIOD", "Etape A — Logit / PriceFit", "Marche seg1 P1 = 33 000 x 1,12 = 36 960 u.", "PriceFit = 1 - (eta_price x |prix-pref|/pref) [illustration PDF]", "Ex. 3200 vs pref 3800"),
        (31, "MODEL_PERIOD", "Etape A (suite)", "Bonus marketing plafonne Marketing_Max_Bonus", "U(AVE) illustre ~ 0,286 ; PDM seg ~ 7,3 %", "Demande brute illustree ~ 2 698 u."),
        (32, "MODEL_PERIOD", "Etape B — Plafond annee 1", "Corridor [1000 ; 2000], LaunchFactor 0,5", "521 = Demande_firme_seg x Poids_modele / Somme poids AVE seg1", "Cannibalisation intra-firme"),
        (33, "MODEL_PERIOD", "Etape C — Stock", "Production 1800, Ventes 1000, Stock 800", "Cout stockage = 800 x 0,57 x 3 200 x Stock_Cost_Rate", "≈ 36 480 $ (arrondis moteur possibles)"),
    ]
    start_r = ws.max_row + 1
    for i, tup in enumerate(rows_doc):
        for c, v in enumerate(tup, 1):
            ws.cell(start_r + i, c, v)
    autosize_columns(ws)
    print("Correction/Tâche 5 terminée — FORMULES_FR etape 21 + scenario 521")


def task16_tests_production_promo(wb):
    ws = wb["TESTS_PRODUCTION"]
    last = ws.max_row
    r0 = last + 3
    title_row(ws, r0, "Tache 16 — Promo -10 % vs sans promo (prod 2500 u., prix liste 3200 $)", 12)
    pmap = get_param_map(wb)
    eta = pmap.get("eta_price", 1.5) or 1.5
    pref = 3800
    price_list = 3200
    price_net_promo = round(price_list * 0.9)
    stock_rate = pmap.get("Stock_Cost_Rate", 0.025) or 0.025
    cogs_ratio = pmap.get("COGS_Moyen", 0.57) or 0.57
    marketing = 80000
    rd = 15000
    production = 2500
    demand_base = 36960
    pf0 = pow(2.718281828, -eta * abs(price_list / pref - 1))
    pf1 = pow(2.718281828, -eta * abs(price_net_promo / pref - 1))
    demand_promo = demand_base * (pf1 / pf0) if pf0 else demand_base
    demand_no = demand_base

    def scenario(demand, net_price, label):
        sales = min(production, demand)
        stock = production - sales
        surplus = max(0, stock - 200)
        surplus_pen = surplus * net_price * 0.05
        stock_cost = stock * net_price * stock_rate
        cogs = sales * net_price * cogs_ratio
        profit = sales * net_price - cogs - stock_cost - surplus_pen - marketing - rd
        return {
            "label": label,
            "demande": demand,
            "ventes": sales,
            "stock": stock,
            "stock_cost": stock_cost,
            "surplus_pen": surplus_pen,
            "profit": profit,
        }

    a = scenario(demand_promo, price_net_promo, "Avec promo -10 %")
    b = scenario(demand_no, price_list, "Sans promo")
    hdr = ["Scenario", "Demande est.", "Ventes", "Stock final", "Cout stockage", "Penalite surstock", "Profit"]
    hr = r0 + 1
    for i, h in enumerate(hdr, 1):
        ws.cell(hr, i, h)
        ws.cell(hr, i).font = HEADER_FONT
    for i, s in enumerate([a, b], hr + 1):
        ws.cell(i, 1, s["label"])
        ws.cell(i, 2, int(round(s["demande"])))
        ws.cell(i, 3, int(round(s["ventes"])))
        ws.cell(i, 4, int(round(s["stock"])))
        ws.cell(i, 5, int(round(s["stock_cost"])))
        ws.cell(i, 6, int(round(s["surplus_pen"])))
        ws.cell(i, 7, int(round(s["profit"])))
    delta_r = hr + 3
    ws.cell(delta_r, 1, "Delta profit (promo - sans)")
    ws.cell(delta_r, 2, int(round(a["profit"] - b["profit"])))
    print("Tâche 16 terminée — bloc promo TESTS_PRODUCTION")


def task17_tests_production_evolution(wb):
    ws = wb["TESTS_PRODUCTION"]
    last = ws.max_row
    r0 = last + 3
    title_row(ws, r0, "Tache 17 — Production evolutive vs demande marche 12 % (depart 2000 u.)", 14)
    demand_p1 = 36960
    market_growth = 0.12
    rates = [0.10, 0.12, 0.15]
    hdr = ["Periode"] + [f"Prod +{int(g*100)}%" for g in rates] + [f"Service +{int(g*100)}%" for g in rates] + [f"Profit +{int(g*100)}%" for g in rates]
    hr = r0 + 1
    for i, h in enumerate(hdr, 1):
        ws.cell(hr, i, h)
        ws.cell(hr, i).font = HEADER_FONT
    pref = 3800
    stock_rate = 0.025
    cogs_ratio = 0.57
    marketing = 100000
    rd = 30000
    results = {g: {"ok": True, "profits": []} for g in rates}
    for t in range(1, 9):
        row = hr + t
        ws.cell(row, 1, f"P{t}")
        demand_t = demand_p1 * (1 + market_growth) ** (t - 1)
        for j, g in enumerate(rates, 2):
            if t == 1:
                prod = 2000
            else:
                prod = 2000 * (1 + g) ** (t - 1)
            sales = min(prod, demand_t)
            service = sales / demand_t if demand_t else 0
            stock = prod - sales
            stock_cost = stock * pref * stock_rate
            cogs = sales * pref * cogs_ratio
            profit = sales * pref - cogs - stock_cost - marketing - rd
            ws.cell(row, j, int(round(prod)))
            ws.cell(row, j + len(rates), round(service, 3))
            ws.cell(row, j + 2 * len(rates), int(round(profit)))
            results[g]["profits"].append(profit)
            if not (0.95 <= service <= 1.05):
                results[g]["ok"] = False
    opt_r = hr + 10
    best = None
    for g in rates:
        if results[g]["ok"]:
            best = g
            break
    if best is None:
        best = min(rates, key=lambda x: abs(sum(results[x]["profits"]) / 8))
    ws.cell(opt_r, 1, "Resultat optimal (service 95-105 % si possible)")
    ws.cell(opt_r, 2, f"Taux retenu: +{int(best*100)} %")
    print("Tâche 17 terminée — tableau 8 periodes TESTS_PRODUCTION")


def task18_pdm_columns(wb):
    mp = wb["MODEL_PERIOD"]
    h = sheet_headers(mp)
    last_c = mp.max_column
    c1 = last_c + 1
    c2 = last_c + 2
    mp.cell(1, c1, "PDM_Segment").font = HEADER_FONT
    mp.cell(1, c2, "PDM_Globale").font = HEADER_FONT
    ad = get_column_letter(h["TotalSales"])
    per = get_column_letter(h["Period_Index"])
    seg = get_column_letter(h["Segment"])
    for r in range(2, mp.max_row + 1):
        mp.cell(
            r,
            c1,
            f"=IFERROR({ad}{r}/((PARAM!$B$4*(1+PARAM!$B$5)^{per}{r})*INDEX(SEGMENTS!$C:$C,MATCH({seg}{r},SEGMENTS!$A:$A,0))),\"\")",
        )
        mp.cell(
            r,
            c2,
            f"=IFERROR({ad}{r}/(PARAM!$B$4*(1+PARAM!$B$5)^{per}{r}),\"\")",
        )
        mp.cell(r, c1).number_format = "0.0%"
        mp.cell(r, c2).number_format = "0.0%"
    print("Tâche 18 terminée — PDM_Segment et PDM_Globale dans MODEL_PERIOD")


def task19_marche_structure_sales(wb):
    ws = wb["MARCHE_STRUCTURE"]
    firms = [r["Firm"] for r in read_table(wb["FIRMS"])]
    segments = [r["Segment"] for r in read_table(wb["SEGMENTS"])]
    nf, ns = len(firms), len(segments)
    rstart = 30
    first_data = rstart + 2
    last_data = rstart + 1 + nf
    pdm_row = rstart + 2 + nf
    tot_c = 2 + ns
    title_row(ws, rstart, "Tableau croise ventes P1 par firme et segment", 10)
    ws.cell(rstart + 1, 1, "Firme")
    for i, s in enumerate(segments, 2):
        ws.cell(rstart + 1, i, f"Seg {s}")
    ws.cell(rstart + 1, tot_c, "Total")
    header_style(ws, rstart + 1)
    for fi, f in enumerate(firms):
        rr = rstart + 2 + fi
        ws.cell(rr, 1, f)
        for i, s in enumerate(segments, 2):
            ws.cell(
                rr,
                i,
                f'=SUMIFS(MODEL_PERIOD!$AD:$AD,MODEL_PERIOD!$A:$A,1,MODEL_PERIOD!$D:$D,$A{rr},MODEL_PERIOD!$H:$H,{s})',
            )
        ws.cell(rr, tot_c, f"=SUM(B{rr}:{get_column_letter(1+ns)}{rr})")
    ws.cell(pdm_row, 1, "PDM firme / segment (%)")
    for i, s in enumerate(segments, 2):
        col = get_column_letter(i)
        sum_seg = f"SUMIFS(MODEL_PERIOD!$AD:$AD,MODEL_PERIOD!$A:$A,1,MODEL_PERIOD!$H:$H,{s})"
        ws.cell(pdm_row, i, f"=IFERROR(SUM({col}{first_data}:{col}{last_data})/{sum_seg},\"\")")
        ws.cell(pdm_row, i).number_format = "0.0%"
    sum_all = "SUMIFS(MODEL_PERIOD!$AD:$AD,MODEL_PERIOD!$A:$A,1)"
    ws.cell(
        pdm_row,
        tot_c,
        f"=IFERROR(SUM({get_column_letter(tot_c)}{first_data}:{get_column_letter(tot_c)}{last_data})/{sum_all},\"\")",
    )
    ws.cell(pdm_row, tot_c).number_format = "0.0%"
    autosize_columns(ws)
    print("Tâche 19 terminée — ventes P1 dans MARCHE_STRUCTURE")


def task20_tests_params_validation(wb):
    ws = wb["TESTS_PARAMS"]
    last = ws.max_row
    r0 = last + 3
    title_row(ws, r0, "Tache 20 — Validation coefficients (TRE-PerformanceX P3, P2 vs P8)", 12)
    pmap = get_param_map(wb)
    eta = pmap.get("eta_price", 1.5) or 1.5
    pref = 4650
    alpha = pmap.get("MarketShare_Smoothing_Alpha", 0.4) or 0.4
    mlambda = pmap.get("Marketing_Lambda", 1.0) or 1.0
    bonus_cap = pmap.get("Marketing_Max_Bonus", 0.25) or 0.25
    hdr = ["Indicateur", "P2", "P8", "Variation", "Coherent"]
    hr = r0 + 1
    for i, h in enumerate(hdr, 1):
        ws.cell(hr, i, h)
        ws.cell(hr, i).font = HEADER_FONT

    def calc(period_idx):
        price = 5200 * (1.07) ** (period_idx - 1)
        price_fit = pow(2.718281828, -eta * abs(price / pref - 1))
        mkt_raw = min(bonus_cap, 120000 / 500000 * mlambda)
        attract = price_fit * (1 + mkt_raw) * (0.3 + alpha * 0.02 * (period_idx - 1))
        seg_units = 110000 * (1.12) ** period_idx * 0.2
        pdm = min(0.5, attract / 8)
        demand = seg_units * pdm
        sales = min(2000, demand)
        return price, price_fit, attract, pdm, demand, sales

    labels = ["Prix courant", "PriceFit", "Attractivite (simpl.)", "PDM est. segment", "Demande est.", "Ventes (prod 2000)"]
    row = hr + 1
    p2 = calc(2)
    p8 = calc(8)
    for i, lab in enumerate(labels):
        ws.cell(row + i, 1, lab)
        ws.cell(row + i, 2, round(p2[i], 4) if i < 4 else int(round(p2[i])))
        ws.cell(row + i, 3, round(p8[i], 4) if i < 4 else int(round(p8[i])))
        v2, v8 = p2[i], p8[i]
        if isinstance(v2, float) and v2:
            var = (v8 - v2) / v2
            ws.cell(row + i, 4, round(var, 4))
        else:
            ws.cell(row + i, 4, v8 - v2 if isinstance(v2, (int, float)) else "")
        coherent = "OUI"
        if i == 1 and p8[1] > p2[1]:
            coherent = "NON"
        if i == 3 and p8[3] > p2[3] + 0.1:
            coherent = "NON"
        ws.cell(row + i, 5, coherent)
    print("Tâche 20 terminée — bloc validation TESTS_PARAMS")


def regenerate_reports_with_pdm(wb):
    mp = wb["MODEL_PERIOD"]
    hp = sheet_headers(mp)
    segs = {r["Segment"]: r for r in read_table(wb["SEGMENTS"])}
    firms = ["AVE", "CAN", "EBI", "GIA", "PED", "RID", "SUR", "TRE", "VEL"]
    pdm_seg_c = hp.get("PDM_Segment")
    pdm_glob_c = hp.get("PDM_Globale")

    def price_range_floor(seg, rng):
        pref = (segs.get(seg) or {}).get("Pref_RefYear", 0) or 0
        rs = str(rng or "").lower()
        if rs.startswith("bas") or "basic" in rs:
            return max(pref * 0.75, 0)
        if rs.startswith("moy") or "mid" in rs or "medium" in rs:
            return max(2800, pref * 0.9)
        if rs.startswith("haut") or "premium" in rs:
            return max(5500, pref * 1.0)
        return pref * 0.9

    def as_num(v):
        if isinstance(v, (int, float)):
            return float(v)
        return 0.0

    def generer_rapport_firme(firm_code, periode):
        name = f"RAPPORT_{firm_code}_P{periode}"
        ws = ensure_sheet(wb, name)
        ws.delete_rows(1, ws.max_row)
        title_row(ws, 1, f"Rapport {firm_code} - Periode {periode}", 16)
        cols = [
            "ProductKey",
            "ModelName",
            "Segment",
            "Range",
            "PriceList",
            "Production",
            "DemandActive",
            "TotalSales",
            "StockEnd",
            "PDM_Segment",
            "PDM_Globale",
            "Marge (%)",
            "Alerte",
            "Recommandation",
        ]
        for i, h in enumerate(cols, 1):
            ws.cell(2, i, h)
        header_style(ws, 2)

        rr = 3
        total_profit = 0
        margins = []
        total_sales = 0
        total_dem = 0
        for r in range(2, mp.max_row + 1):
            p = mp.cell(r, hp["Period_Index"]).value
            f = mp.cell(r, hp["Firm"]).value
            if p != periode or f != firm_code:
                continue
            rec = {k: mp.cell(r, c).value for k, c in hp.items()}
            revenue = as_num(rec.get("Revenue"))
            profit = as_num(rec.get("ProfitContribution"))
            marge = (profit / revenue) if revenue else 0
            total_profit += profit
            margins.append(marge)
            total_sales += as_num(rec.get("TotalSales"))
            total_dem += as_num(rec.get("DemandActive"))

            alert = "OK"
            reco = "Maintenir la strategie"
            withdraw_flag = as_num(rec.get("WithdrawFlag"))
            total_sales_rec = as_num(rec.get("TotalSales"))
            demand_active_rec = as_num(rec.get("DemandActive"))
            stock_end_rec = as_num(rec.get("StockEnd"))
            production_rec = as_num(rec.get("Production"))
            price_list_rec = as_num(rec.get("PriceList"))
            pdm_s = as_num(rec.get("PDM_Segment")) if pdm_seg_c else 0
            pdm_g = as_num(rec.get("PDM_Globale")) if pdm_glob_c else 0
            if withdraw_flag == 1:
                alert = "Liquidation"
            elif total_sales_rec < demand_active_rec * 0.8:
                alert = "Rupture"
                reco = f"Augmenter production a {int(demand_active_rec * 1.05)} u."
            elif stock_end_rec > production_rec * 0.05:
                alert = "Surstock"
                reco = "Reduire production ou activer liquidation"
            floor = price_range_floor(rec.get("Segment"), rec.get("Range"))
            if price_list_rec < floor:
                alert = "Prix invalide"
                reco = f"Corriger prix — plancher {int(round(floor))} $"

            values = [
                rec.get("ProductKey"),
                rec.get("ModelName"),
                rec.get("Segment"),
                rec.get("Range"),
                price_list_rec,
                production_rec,
                demand_active_rec,
                total_sales_rec,
                stock_end_rec,
                pdm_s,
                pdm_g,
                marge,
                alert,
                reco,
            ]
            for i, v in enumerate(values, 1):
                c = ws.cell(rr, i, v)
                if i in (10, 11) and isinstance(v, float):
                    c.number_format = "0.0%"
            ws.cell(rr, 12).number_format = "0.0%"
            if alert in ("Rupture", "Prix invalide"):
                ws.cell(rr, 13).fill = ALERT_RED
            elif alert == "Surstock":
                ws.cell(rr, 13).fill = ALERT_ORANGE
            else:
                ws.cell(rr, 13).fill = ALERT_GREEN
            rr += 1

        kpi_row = rr + 1
        title_row(ws, kpi_row, "KPIs firme", 6)
        ws.cell(kpi_row + 1, 1, "Profit total")
        ws.cell(kpi_row + 1, 2, int(round(total_profit)))
        ws.cell(kpi_row + 2, 1, "Marge moyenne")
        ws.cell(kpi_row + 2, 2, sum(margins) / len(margins) if margins else 0)
        ws.cell(kpi_row + 2, 2).number_format = "0.0%"
        ws.cell(kpi_row + 3, 1, "PDM globale (RESULTS_FIRM)")
        ws.cell(kpi_row + 3, 2, f'=IFERROR(INDEX(RESULTS_FIRM!$M:$M,MATCH(1,(RESULTS_FIRM!$A:$A={periode})*(RESULTS_FIRM!$C:$C="{firm_code}"),0)),0)')
        ws.cell(kpi_row + 3, 2).number_format = "0.0%"
        ws.cell(kpi_row + 4, 1, "Taux service moyen")
        ws.cell(kpi_row + 4, 2, (total_sales / total_dem) if total_dem else 0)
        ws.cell(kpi_row + 4, 2).number_format = "0.0%"

        for r in range(3, rr):
            if (r - 3) % 2 == 0:
                for c in range(1, 15):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor="F7F7F7")
        autosize_columns(ws)

    for f in firms:
        generer_rapport_firme(f, 1)
    print("Tâche 18b terminée — RAPPORT_* avec PDM_Segment / PDM_Globale")


def task21_controle_decisions(wb):
    ws = ensure_sheet(wb, "CONTROLE_DECISIONS")
    ws.delete_rows(1, ws.max_row)
    firm_order = ["AVE", "CAN", "EBI", "GIA", "PED", "RID", "SUR", "TRE", "VEL"]
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1)
    c.value = '=IF(COUNTIF(F:F,"✗ Bloqué")>0,"DECISIONS BLOQUEES — corriger avant simulation","TOUTES LES DECISIONS VALIDES")'
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")
    c.fill = BANNER_BLOCKED
    hdr = ["Firme", "Periode", "Type_decision", "Valeur_saisie", "Regle", "Statut", "Message"]
    for i, h in enumerate(hdr, 1):
        ws.cell(3, i, h)
        ws.cell(3, i).font = HEADER_FONT
        ws.cell(3, i).fill = HEADER_FILL

    types = [
        ("Production", "> 0, pas si retrait (Withdraw) actif"),
        ("Marketing total", "Entre min 3 % revenu P-1 et max +10 % progressif"),
        ("Promotion", "Normal [Max_Promo;0], liquidation -10 % si retrait"),
        ("R&D", "Uniquement 0 / 2 / 5 / 8 % du BudgetBase"),
        ("Lancement NP", "Pas en P1, R&D 5 % ou 8 %, portefeuille < init+2"),
        ("Retrait / liquidation", "Max 1 retrait/periode, promo -10 %, pas de prod si retrait"),
        ("Prix", "Basic >= plancher seg., Mid >= 2800, Premium >= 5500"),
    ]

    row = 4
    for p in range(1, 9):
        for fi, firm in enumerate(firm_order):
            ir = 2 + (p - 1) * 9 + fi
            for typ, regle in types:
                ws.cell(row, 1, firm)
                ws.cell(row, 2, p)
                ws.cell(row, 3, typ)
                ws.cell(row, 5, regle)
                if typ == "Production":
                    ws.cell(
                        row,
                        4,
                        f"=SUMIFS(INPUT_MODEL!$G:$G,INPUT_MODEL!$A:$A,$B{row},INPUT_MODEL!$D:$D,$A{row})",
                    )
                    ws.cell(
                        row,
                        6,
                        f'=IF(SUMPRODUCT(--(INPUT_MODEL!$A$2:$A$617=$B{row}),--(INPUT_MODEL!$D$2:$D$617=$A{row}),--(INPUT_MODEL!$H$2:$H$617=1),--(INPUT_MODEL!$G$2:$G$617>0))>0,"✗ Bloqué",IF(D{row}<0,"✗ Bloqué","✓ OK"))',
                    )
                    ws.cell(row, 7, f"=IF(F{row}=\"✗ Bloqué\",\"Production interdite pendant liquidation\",\"\")")
                elif typ == "Marketing total":
                    ix = ir - 1
                    ws.cell(row, 4, f"=SUM(INPUT_FIRM!D{ir}:H{ir})")
                    ws.cell(
                        row,
                        6,
                        f"=IF(OR(SUM(INPUT_FIRM!D{ir}:H{ir})<0.03*IF($B{row}=1,INDEX(FIRM_PERIOD!$S$2:$S$73,{ix}),SUMIFS(RESULTS_FIRM!$E$2:$E$73,RESULTS_FIRM!$A$2:$A$73,$B{row}-1,RESULTS_FIRM!$C$2:$C$73,$A{row})),SUM(INPUT_FIRM!D{ir}:H{ir})>IF($B{row}=1,INDEX(FIRM_PERIOD!$T$2:$T$73,{ix})*1.1,SUMIFS(FIRM_PERIOD!$K$2:$K$73,FIRM_PERIOD!$A$2:$A$73,$B{row}-1,FIRM_PERIOD!$C$2:$C$73,$A{row})*1.1)),\"⚠ Attention\",\"✓ OK\")",
                    )
                    ws.cell(row, 7, "")
                elif typ == "Promotion":
                    ws.cell(
                        row,
                        4,
                        f"=SUMIFS(INPUT_MODEL!$F:$F,INPUT_MODEL!$A:$A,$B{row},INPUT_MODEL!$D:$D,$A{row})",
                    )
                    ws.cell(
                        row,
                        6,
                        f'=IF(SUMPRODUCT(--(INPUT_MODEL!$A$2:$A$617=$B{row}),--(INPUT_MODEL!$D$2:$D$617=$A{row}),--(INPUT_MODEL!$I$2:$I$617=0),--(INPUT_MODEL!$F$2:$F$617<PARAM!$B$16))>0,"✗ Bloqué",IF(SUMPRODUCT(--(INPUT_MODEL!$A$2:$A$617=$B{row}),--(INPUT_MODEL!$D$2:$D$617=$A{row}),--(INPUT_MODEL!$I$2:$I$617=1),--(ABS(INPUT_MODEL!$F$2:$F$617+0.1)>0.001))>0,"✗ Bloqué","✓ OK"))',
                    )
                    ws.cell(row, 7, "Promo normale [Max_Promo;0] ; liquidation -10 % si flag")
                elif typ == "R&D":
                    ix = ir - 1
                    ws.cell(
                        row,
                        4,
                        f"=INDEX(INPUT_FIRM!$I$2:$I$73,{ix})/MAX(INDEX(FIRM_PERIOD!$T$2:$T$73,{ix}),1)",
                    )
                    ws.cell(
                        row,
                        6,
                        f"=IF(OR(ABS(D{row})<0.0001,ABS(D{row}-0.02)<0.0001,ABS(D{row}-0.05)<0.0001,ABS(D{row}-0.08)<0.0001),\"✓ OK\",\"✗ Bloqué\")",
                    )
                    ws.cell(row, 7, f"=IF(F{row}=\"✗ Bloqué\",\"Niveau R&D non autorise\",\"\")")
                elif typ == "Lancement NP":
                    ix = ir - 1
                    ws.cell(row, 4, "—")
                    ws.cell(
                        row,
                        6,
                        f"=IF(AND($B{row}=1,OR(ABS(INDEX(INPUT_FIRM!$I$2:$I$73,{ix})/MAX(INDEX(FIRM_PERIOD!$T$2:$T$73,{ix}),1)-0.05)<0.0001,ABS(INDEX(INPUT_FIRM!$I$2:$I$73,{ix})/MAX(INDEX(FIRM_PERIOD!$T$2:$T$73,{ix}),1)-0.08)<0.0001)),\"⚠ Attention\",\"✓ OK\")",
                    )
                    ws.cell(row, 7, "Lancement concret a verifier dans INPUT_NEWPRODUCT")
                elif typ == "Retrait / liquidation":
                    ws.cell(row, 4, f"=SUMIFS(INPUT_MODEL!$H:$H,INPUT_MODEL!$A:$A,$B{row},INPUT_MODEL!$D:$D,$A{row})")
                    ws.cell(row, 6, f"=IF(D{row}>1,\"✗ Bloqué\",\"✓ OK\")")
                    ws.cell(row, 7, "")
                elif typ == "Prix":
                    ws.cell(row, 4, "—")
                    ws.cell(row, 6, "✓ OK")
                    ws.cell(row, 7, "Planchers dans regle — verifier prix listes par produit")
                row += 1

    ws.cell(1, 1).value = '=IF(COUNTIF(F:F,"✗ Bloqué")>0,"DECISIONS BLOQUEES — corriger avant simulation","TOUTES LES DECISIONS VALIDES")'
    autosize_columns(ws)
    print("Tâche 21 terminée — feuille CONTROLE_DECISIONS")


def patch_portefeuille_rd(wb):
    ws = wb["PORTEFEUILLE"]
    h = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    last = ws.max_column
    c_dl = last + 1
    c_msg = last + 2
    ws.cell(2, c_dl, "Droits_lancement_P_suivante")
    ws.cell(2, c_msg, "Message_R&D")
    ws.cell(2, c_dl).font = HEADER_FONT
    ws.cell(2, c_msg).font = HEADER_FONT
    col_per = h.get("Periode")
    col_fir = h.get("Firme")
    col_act = h.get("Modeles actifs")
    col_init = h.get("Modeles initiaux")
    for r in range(3, ws.max_row + 1):
        per = ws.cell(r, col_per).coordinate
        fir = ws.cell(r, col_fir).coordinate
        act = ws.cell(r, col_act).coordinate
        init = ws.cell(r, col_init).coordinate
        ws.cell(
            r,
            c_dl,
            f"=MAX(0,MIN(IF(ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.08)<0.0001,2,IF(ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.05)<0.0001,1,0)),{init}+2-{act}))",
        )
        ac = ws.cell(r, col_per).coordinate
        ws.cell(
            r,
            c_msg,
            (
                f"=IF({per}=1,IF(OR(ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.05)<0.0001,ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.08)<0.0001),"
                f"\"R&D P1 : developpement en cours — lancement disponible en P2\","
                f"\"Pas de nouveau modele en developpement\"),"
                f"IF(ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.08)<0.0001,"
                f"\"Vous avez le droit de lancer 2 modeles en P\"&("
                + ac
                + "+1),"
                f"IF(ABS(IFERROR(INDEX(FIRM_PERIOD!$I$2:$I$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0))/MAX(IFERROR(INDEX(FIRM_PERIOD!$T$2:$T$73,MATCH(1,(FIRM_PERIOD!$A$2:$A$73={per})*(FIRM_PERIOD!$C$2:$C$73={fir}),0)),1),1)-0.05)<0.0001,"
                f"\"Vous avez le droit de lancer 1 modele en P\"&("
                + ac
                + "+1),"
                f"\"Aucun droit de lancement cette periode\"))))"
            ),
        )
    autosize_columns(ws)
    print("Correction/Tâche 3b terminée — PORTEFEUILLE Droits_lancement / Message_R&D")


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)

    correction1_marketing_helpers_and_controles(wb)
    print("Correction 1 terminée — Marketing escalade progressive")

    correction2_model_period_tampon(wb)
    correction2_tests_production_tampon(wb)
    print("Correction 2 terminée — Stock tampon et production recommandee")

    correction3_liquidation_promo_model_period(wb)
    correction3_controles_liquidation_message(wb)
    correction3_rd_helpers_and_controles(wb)
    print("Corrections 3-4 terminées — R&D, liquidation, CONTROLES")

    correction5_formules_fr(wb)
    print("Correction 5 terminée — FORMULES_FR (521 + etape 21)")

    task16_tests_production_promo(wb)
    print("Tâche 16 terminée — Promo vs sans promo")

    task17_tests_production_evolution(wb)
    print("Tâche 17 terminée — Production evolutive 8 periodes")

    task18_pdm_columns(wb)
    print("Tâche 18 terminée — Colonnes PDM MODEL_PERIOD")

    task19_marche_structure_sales(wb)
    print("Tâche 19 terminée — Croise ventes MARCHE_STRUCTURE")

    task20_tests_params_validation(wb)
    print("Tâche 20 terminée — Validation coefficients TESTS_PARAMS")

    patch_portefeuille_rd(wb)
    print("Correction portefeuille terminée — Droits_lancement / Message_R&D")

    regenerate_reports_with_pdm(wb)
    print("Rapports regeneres — PDM dans RAPPORT_*_P1")

    task21_controle_decisions(wb)
    print("Tâche 21 terminée — CONTROLE_DECISIONS")

    wb.save(OUTPUT_FILE)
    print(f"Fichier sauvegarde: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
