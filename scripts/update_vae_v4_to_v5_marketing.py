"""
Patch VAE classeur v4 -> v5 : seuils marketing (PARAM + CONTROLES + CONTROLE_DECISIONS + FORMULES_FR).
Ne modifie pas MODEL_PERIOD, FIRM_PERIOD, INPUT_FIRM, INPUT_MODEL.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

SRC = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v4.xlsx")
DST = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v5.xlsx")

PARAM_ROWS = [
    ("Marketing_Min_Abs", 50000),
    ("Marketing_Min_Pct", 0.03),
    ("Marketing_Ref_AVE", 80000),
    ("Marketing_Ref_CAN", 80000),
    ("Marketing_Ref_EBI", 80000),
    ("Marketing_Ref_GIA", 80000),
    ("Marketing_Ref_PED", 80000),
    ("Marketing_Ref_RID", 80000),
    ("Marketing_Ref_SUR", 80000),
    ("Marketing_Ref_TRE", 120000),
    ("Marketing_Ref_VEL", 80000),
]


def prev_mkt_sumifs(r: int) -> str:
    """Somme marketing saisi periode precedente (colonnes D a H INPUT_FIRM)."""
    return (
        f'SUMIFS(INPUT_FIRM!$D:$D,INPUT_FIRM!$A:$A,I{r}-1,INPUT_FIRM!$C:$C,J{r})'
        f'+SUMIFS(INPUT_FIRM!$E:$E,INPUT_FIRM!$A:$A,I{r}-1,INPUT_FIRM!$C:$C,J{r})'
        f'+SUMIFS(INPUT_FIRM!$F:$F,INPUT_FIRM!$A:$A,I{r}-1,INPUT_FIRM!$C:$C,J{r})'
        f'+SUMIFS(INPUT_FIRM!$G:$G,INPUT_FIRM!$A:$A,I{r}-1,INPUT_FIRM!$C:$C,J{r})'
        f'+SUMIFS(INPUT_FIRM!$H:$H,INPUT_FIRM!$A:$A,I{r}-1,INPUT_FIRM!$C:$C,J{r})'
    )


def base_l(r: int) -> str:
    return (
        f'IF(I{r}=1,VLOOKUP("Marketing_Ref_"&J{r},PARAM!$A:$B,2,FALSE),'
        f'{prev_mkt_sumifs(r)})'
    )


def marketing_min(r: int) -> str:
    return (
        f'MAX(VLOOKUP("Marketing_Min_Abs",PARAM!$A:$B,2,FALSE),'
        f'VLOOKUP("Marketing_Min_Pct",PARAM!$A:$B,2,FALSE)*({base_l(r)}))'
    )


def marketing_max(r: int) -> str:
    return f'({base_l(r)})*1.1'


def patch_param(ws):
    start = 39
    for i, (name, val) in enumerate(PARAM_ROWS):
        ws.cell(start + i, 1, name)
        ws.cell(start + i, 2, val)


def patch_controles_helpers(ws):
    ws.cell(160, 8, "AIDE_MARKETING (indices FIRM_PERIOD 1..72)")
    titles = {
        8: "Idx",
        9: "Periode",
        10: "Firme",
        11: "TotalMkt",
        12: "BaseMkt_ref_ou_Pm1",
        13: "Marketing_Min",
        14: "Marketing_Max",
        16: "SousMin",
        17: "AuDessusMax",
    }
    for col, title in titles.items():
        ws.cell(161, col, title)

    for r in range(162, 234):
        ws.cell(r, 12, f'={base_l(r)}')
        ws.cell(r, 13, f'={marketing_min(r)}')
        ws.cell(r, 14, f'={marketing_max(r)}')
        ws.cell(row=r, column=15, value=None)
        ws.cell(r, 16, f'=--(K{r}<M{r})')
        ws.cell(r, 17, f'=--(K{r}>N{r})')

    ws.cell(
        148,
        6,
        "Budget marketing sous le plancher (MAX(Min_Abs ; Min_Pct x base marketing ref ou P-1))",
    )
    ws.cell(
        149,
        6,
        "Plafond +10 % sur budget marketing ref (P1) ou total marketing reel periode precedente (P2+)",
    )


def cd_prev_sumifs(sr: int) -> str:
    return (
        f'SUMIFS(INPUT_FIRM!$D:$D,INPUT_FIRM!$A:$A,$B{sr}-1,INPUT_FIRM!$C:$C,$A{sr})'
        f'+SUMIFS(INPUT_FIRM!$E:$E,INPUT_FIRM!$A:$A,$B{sr}-1,INPUT_FIRM!$C:$C,$A{sr})'
        f'+SUMIFS(INPUT_FIRM!$F:$F,INPUT_FIRM!$A:$A,$B{sr}-1,INPUT_FIRM!$C:$C,$A{sr})'
        f'+SUMIFS(INPUT_FIRM!$G:$G,INPUT_FIRM!$A:$A,$B{sr}-1,INPUT_FIRM!$C:$C,$A{sr})'
        f'+SUMIFS(INPUT_FIRM!$H:$H,INPUT_FIRM!$A:$A,$B{sr}-1,INPUT_FIRM!$C:$C,$A{sr})'
    )


def cd_base(sr: int) -> str:
    return (
        f'IF($B{sr}=1,VLOOKUP("Marketing_Ref_"&$A{sr},PARAM!$A:$B,2,FALSE),'
        f'{cd_prev_sumifs(sr)})'
    )


def cd_marketing_min(sr: int) -> str:
    return (
        f'MAX(VLOOKUP("Marketing_Min_Abs",PARAM!$A:$B,2,FALSE),'
        f'VLOOKUP("Marketing_Min_Pct",PARAM!$A:$B,2,FALSE)*({cd_base(sr)}))'
    )


def cd_marketing_max(sr: int) -> str:
    return f'({cd_base(sr)})*1.1'


def patch_controle_decisions(ws):
    for sr in range(2, ws.max_row + 1):
        if ws.cell(sr, 3).value != "Marketing total":
            continue
        dcell = ws.cell(sr, 4).value
        if not dcell or "SUM(INPUT_FIRM!" not in str(dcell):
            continue
        m = re.search(r"SUM\(INPUT_FIRM!D(\d+):H(\d+)\)", str(dcell))
        if not m:
            continue
        ir = m.group(1)
        sum_expr = f"SUM(INPUT_FIRM!D{ir}:H{ir})"
        m_min = cd_marketing_min(sr)
        m_max = cd_marketing_max(sr)
        ws.cell(sr, 5, "Plancher MAX(Min_Abs ; Min_Pct x base) ; plafond base x 110 % (base = ref P1 ou MKT reel P-1)")
        ws.cell(
            sr,
            6,
            f'=IF({sum_expr}<{m_min},"⚠ Attention",IF({sum_expr}>{m_max},"✗ Bloqué","✓ OK"))',
        )
        ws.cell(
            sr,
            7,
            f'=IF({sum_expr}<{m_min},"Budget marketing sous le plancher ("&TEXT({m_min},"# ##0 $")&" min)",IF({sum_expr}>{m_max},"Budget marketing dépasse le plafond ("&TEXT({m_max},"# ##0 $")&" max)",""))',
        )


def patch_formules_fr(ws):
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 3).value == "Limites marketing":
            ws.cell(
                r,
                4,
                "Budget marketing — règle corrigée (ref PARAM, pas BudgetBase_Adjusted)",
            )
            ws.cell(
                r,
                5,
                "Plancher P1 = MAX(Min_Abs ; Min_Pct x Marketing_Ref_firme). Plafond P1 = Marketing_Ref x 1,10. "
                "P2+ : base = total marketing reel P-1 (INPUT_FIRM). BudgetBase_Adjusted (revenu firme) n entre pas dans ce calcul.",
            )
            break


def main():
    if not SRC.exists():
        raise SystemExit(f"Fichier source introuvable : {SRC}")
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    patch_param(wb["PARAM"])
    patch_controles_helpers(wb["CONTROLES"])
    patch_controle_decisions(wb["CONTROLE_DECISIONS"])
    patch_formules_fr(wb["FORMULES_FR"])

    wb.save(DST)
    print("Correction marketing terminée — v5 sauvegardé")
    print(f"  -> {DST}")
    print("TRE P1 : plancher ≈ 50 000 $, plafond = 132 000 $")
    print("AVE P1 : plancher ≈ 50 000 $, plafond = 88 000 $")


if __name__ == "__main__":
    main()
