#!/usr/bin/env python3
"""
Post-traitement openpyxl du classeur export plan (ex. reports/Export_TRE_plan.xlsx).

1) Supprime les lignes dont la colonne A est un libellé ROI (marketing / R&D).
2) Normalise les gammes (Standard → Medium) ; note P037 sur 1.5 Portefeuille si besoin.
3) Ajoute le bloc « Simulation de référence — 2 250 unités » sur chaque feuille Pxxx Scénario.

Termine par un JSON sur stdout avec status \"success\" ou \"error\".
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parent.parent
_SCRIPTS = Path(__file__).resolve().parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_SCRIPTS))

from openpyxl import load_workbook  # noqa: E402
from openpyxl.workbook.workbook import Workbook  # noqa: E402

from excel_2250_reference import (  # noqa: E402
    enrich_workbook_product_scenarios,
    header_map,
    SCENARIO_SHEET_RE,
)
from vae_rapport_firme_logic import resolve_firm_code  # noqa: E402

log = logging.getLogger("recalc")

GAMME_CORR_NOTE = "⚠ Gamme corrigée : Standard → Medium."
P037_CODE = "P037"
ROI_LABELS = frozenset({"roi marketing", "roi r&d"})


def _norm_col_a(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().lower().replace("&amp;", "&")


def _is_roi_label(norm: str) -> bool:
    return norm in ROI_LABELS or norm == "roi r&amp;d"


def correction1_remove_roi_rows(wb: Workbook) -> int:
    removed = 0
    for ws in wb.worksheets:
        for r in range(ws.max_row, 0, -1):
            norm = _norm_col_a(ws.cell(r, 1).value)
            if _is_roi_label(norm):
                ws.delete_rows(r, 1)
                removed += 1
    return removed


def correction2_standard_medium(wb: Workbook) -> dict[str, int]:
    stats = {"portfolio_gamme": 0, "portfolio_note": 0, "scenario_gamme": 0, "global_replace": 0}

    def normalize_gamme_column(ws: Any, *, is_portfolio: bool) -> None:
        nonlocal stats
        hmap = header_map(ws)
        if "Gamme" not in hmap:
            return
        gc = hmap["Gamme"]
        cc = hmap.get("Code")
        nc = hmap.get("Anomalie / note") or hmap.get("Anomalie / Note")
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, gc).value or "").strip() != "Standard":
                continue
            ws.cell(r, gc, "Medium")
            if is_portfolio:
                stats["portfolio_gamme"] += 1
            else:
                stats["scenario_gamme"] += 1
            if is_portfolio and cc and nc and str(ws.cell(r, cc).value or "").strip() == P037_CODE:
                note_s = str(ws.cell(r, nc).value or "")
                if GAMME_CORR_NOTE not in note_s:
                    sep = "\n" if note_s else ""
                    ws.cell(r, nc, value=f"{note_s}{sep}{GAMME_CORR_NOTE}")
                    stats["portfolio_note"] += 1

    if "1.5 Portefeuille" in wb.sheetnames:
        normalize_gamme_column(wb["1.5 Portefeuille"], is_portfolio=True)

    for name in wb.sheetnames:
        if not SCENARIO_SHEET_RE.match(name):
            continue
        normalize_gamme_column(wb[name], is_portfolio=False)

    pat = re.compile(r"\bStandard\b")

    def replace_in_sheet(title: str) -> None:
        nonlocal stats
        if title not in wb.sheetnames:
            return
        ws = wb[title]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and pat.search(v):
                    c.value = pat.sub("Medium", v)
                    stats["global_replace"] += 1

    replace_in_sheet("Réaction Concurrents")
    replace_in_sheet("1.10 Recommandations")
    return stats


def run(path: Path, firm: str) -> dict[str, Any]:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    firm = resolve_firm_code(firm)
    if not path.exists():
        raise FileNotFoundError(str(path))

    wb = load_workbook(path, data_only=False)
    n_roi = correction1_remove_roi_rows(wb)
    c2 = correction2_standard_medium(wb)
    c3 = enrich_workbook_product_scenarios(wb, firm)
    wb.save(path)

    return {
        "status": "success",
        "workbook": str(path.resolve()),
        "firm": firm,
        "roi_rows_removed": n_roi,
        "correction2": c2,
        "correction3": c3,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Post-traitement Export_*_plan.xlsx (ROI, gammes, bloc 2 250 u.)")
    ap.add_argument(
        "--workbook",
        type=Path,
        default=_ROOT / "reports" / "Export_TRE_plan.xlsx",
        help="Chemin du classeur export à modifier",
    )
    ap.add_argument("--firme", default="TRE", help="Code firme (pour scénarios P2)")
    args = ap.parse_args()
    try:
        report = run(args.workbook.expanduser(), args.firme)
        print(json.dumps(report, ensure_ascii=False, indent=2))
    except Exception as e:
        log.exception("recalc")
        print(json.dumps({"status": "error", "error": str(e)}, ensure_ascii=False, indent=2))
        raise SystemExit(1) from e


if __name__ == "__main__":
    main()
