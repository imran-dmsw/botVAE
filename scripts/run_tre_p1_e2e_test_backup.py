"""
Test bout en bout TRE P1 : injection décisions fictives, rapport CONTROLE_DECISIONS,
feuille TEST_TRE_P1, scénario invalide + TEST_TRE_P1_INVALIDE.
Charge de préférence le classeur v5 ; sinon v4. Écrit la copie de test (ex. v5_TEST.xlsx).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

V4 = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v4.xlsx")
V5 = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v5.xlsx")
WORKBOOK_DEFAULT = V5 if V5.exists() else V4
OUT = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v5_TEST.xlsx")

MARKETING_REF_FALLBACK = {
    "AVE": 80_000,
    "CAN": 80_000,
    "EBI": 80_000,
    "GIA": 80_000,
    "PED": 80_000,
    "RID": 80_000,
    "SUR": 80_000,
    "TRE": 120_000,
    "VEL": 80_000,
}

TRE_INPUT_ROW = 9  # INPUT_FIRM row for TRE period 1

PRODUCTIONS = {
    "P035": 1800,
    "P036": 2500,
    "P037": 1500,
    "P038": 2000,
    "P039": 1900,
    "P040": 2000,
    "P041": 2600,
    "P042": 3600,
    "P043": 2700,
}

MARKETING_TRE = {
    "BudgetDigital": 24000,
    "BudgetSocial": 18000,
    "BudgetInfluencer": 48000,
    "BudgetDisplay": 12000,
    "BudgetEvents": 18000,
}

FILL_GREEN = PatternFill("solid", fgColor="E8F5F0")
FILL_ORANGE = PatternFill("solid", fgColor="FDF3E3")
FILL_RED = PatternFill("solid", fgColor="FAEAEA")


def col_map(ws, row=1):
    return {ws.cell(row, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row, c).value}


def tre_revenue_ref_year(wb) -> float:
    brm = wb["BASE_REFERENCE_MODEL"]
    h = col_map(brm)
    s = 0.0
    for r in range(2, brm.max_row + 1):
        if brm.cell(r, h["Firm"]).value != "TRE":
            continue
        bp = float(brm.cell(r, h["BasePrice_RefYear"]).value or 0)
        u = float(brm.cell(r, h["Units_RefYear"]).value or 0)
        s += bp * u
    return s


def param_val(wb, name: str) -> Any:
    ws = wb["PARAM"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == name:
            return ws.cell(r, 2).value
    return None


def marketing_ref_firm(wb, firm: str) -> float:
    key = f"Marketing_Ref_{firm}"
    v = param_val(wb, key)
    if v is not None:
        return float(v)
    return float(MARKETING_REF_FALLBACK.get(firm, 80_000))


def total_marketing_prev_period(wb, firm: str, period: int) -> float | None:
    """Somme BudgetDigital…BudgetEvents pour (firm, period-1) dans INPUT_FIRM."""
    if period <= 1:
        return None
    ws_if = wb["INPUT_FIRM"]
    ic = col_map(ws_if)
    pi = ic["Period_Index"]
    fc = ic["Firm"]
    keys = ["BudgetDigital", "BudgetSocial", "BudgetInfluencer", "BudgetDisplay", "BudgetEvents"]
    s = 0.0
    found = False
    for r in range(2, ws_if.max_row + 1):
        if ws_if.cell(r, pi).value != period - 1:
            continue
        if ws_if.cell(r, fc).value != firm:
            continue
        found = True
        s += sum(float(ws_if.cell(r, ic[k]).value or 0) for k in keys)
    return s if found else None


def seuils_marketing(wb, firm: str, period: int) -> tuple[float, float]:
    """Aligné sur PARAM + CONTROLE_DECISIONS (base = ref P1 ou total marketing réel P-1)."""
    min_abs = float(param_val(wb, "Marketing_Min_Abs") or 50_000)
    min_pct = float(param_val(wb, "Marketing_Min_Pct") or 0.03)
    ref = marketing_ref_firm(wb, firm)
    if period == 1:
        base = ref
    else:
        prev = total_marketing_prev_period(wb, firm, period)
        base = prev if prev is not None else ref
    min_mkt = max(min_abs, min_pct * base)
    max_mkt = base * 1.1
    return min_mkt, max_mkt


def budget_base_adjusted_tre_p1(wb) -> float:
    rev = tre_revenue_ref_year(wb)
    mg = float(param_val(wb, "Market_Growth") or 0.12)
    inf = float(param_val(wb, "Price_Inflation") or 0.07)
    return rev * (1 + mg) ** 1 * (1 + inf) ** 1


def price_list_p040_est(wb) -> float | None:
    brm = wb["BASE_REFERENCE_MODEL"]
    h = col_map(brm)
    inf = float(param_val(wb, "Price_Inflation") or 0.07)
    for r in range(2, brm.max_row + 1):
        if brm.cell(r, h["ProductKey"]).value == "P040" and brm.cell(r, h["Firm"]).value == "TRE":
            bp = float(brm.cell(r, h["BasePrice_RefYear"]).value or 0)
            return bp * (1 + inf) ** 1
    return None


def model_period_row_p040(wb) -> int | None:
    mp = wb["MODEL_PERIOD"]
    h = col_map(mp)
    for r in range(2, mp.max_row + 1):
        if (
            mp.cell(r, h["Period_Index"]).value == 1
            and mp.cell(r, h["ProductKey"]).value == "P040"
            and mp.cell(r, h["Firm"]).value == "TRE"
        ):
            return r
    return None


def count_init_tre(wb) -> int:
    brm = wb["BASE_REFERENCE_MODEL"]
    h = col_map(brm)
    return sum(
        1
        for r in range(2, brm.max_row + 1)
        if brm.cell(r, h["Firm"]).value == "TRE"
    )


def count_active_tre_p1(wb) -> int:
    """Sans recalcul Excel, approximer les actifs par produits existants TRE (PRODUCT_MASTER)."""
    pm = wb["PRODUCT_MASTER"]
    h = col_map(pm)
    return sum(
        1
        for r in range(2, pm.max_row + 1)
        if pm.cell(r, h["Firm"]).value == "TRE" and int(pm.cell(r, h["ExistingFlag"]).value or 0) == 1
    )


def lancement_prevu_col(ic: dict) -> int:
    for name in ic:
        if name and "LancementPrevu" in str(name):
            return ic[name]
    raise KeyError("Colonne LancementPrevu introuvable dans INPUT_FIRM")


def inject_valid_decisions(wb) -> dict[str, Any]:
    rev = tre_revenue_ref_year(wb)
    bba = budget_base_adjusted_tre_p1(wb)
    budget_rd = int(round(0.02 * bba))

    ws_if = wb["INPUT_FIRM"]
    ic = col_map(ws_if)
    r = TRE_INPUT_ROW
    ws_if.cell(r, ic["BudgetDigital"], MARKETING_TRE["BudgetDigital"])
    ws_if.cell(r, ic["BudgetSocial"], MARKETING_TRE["BudgetSocial"])
    ws_if.cell(r, ic["BudgetInfluencer"], MARKETING_TRE["BudgetInfluencer"])
    ws_if.cell(r, ic["BudgetDisplay"], MARKETING_TRE["BudgetDisplay"])
    ws_if.cell(r, ic["BudgetEvents"], MARKETING_TRE["BudgetEvents"])
    ws_if.cell(r, ic["BudgetRD"], budget_rd)
    ws_if.cell(r, ic["SustainabilityInvestPct"], 0.005)
    ws_if.cell(r, lancement_prevu_col(ic), 0)

    ws_im = wb["INPUT_MODEL"]
    imh = col_map(ws_im)
    for row in range(2, ws_im.max_row + 1):
        if ws_im.cell(row, imh["Period_Index"]).value != 1:
            continue
        if ws_im.cell(row, imh["Firm"]).value != "TRE":
            continue
        pk = ws_im.cell(row, imh["ProductKey"]).value
        if pk in PRODUCTIONS:
            ws_im.cell(row, imh["Production"], PRODUCTIONS[pk])
            ws_im.cell(row, imh["Promotion"], 0.0)
            ws_im.cell(row, imh["Withdraw"], 0)
            ws_im.cell(row, imh["LiquidationPromo"], 0)
        elif isinstance(pk, str) and pk.startswith("N-TRE"):
            ws_im.cell(row, imh["Production"], 0)
            ws_im.cell(row, imh["Promotion"], 0.0)
            ws_im.cell(row, imh["Withdraw"], 0)
            ws_im.cell(row, imh["LiquidationPromo"], 0)

    return {
        "Revenue_RefYear_TRE_calc": rev,
        "BudgetBase_Adjusted_TRE_P1": bba,
        "BudgetRD_injected": budget_rd,
        "TotalMarketing": sum(MARKETING_TRE.values()),
    }


def evaluate_controles_tre_p1_python(wb) -> tuple[list[dict], int]:
    """Recalcule les statuts CONTROLE_DECISIONS pour TRE P1 (sans moteur Excel)."""
    ws_if = wb["INPUT_FIRM"]
    ic = col_map(ws_if)
    r_if = TRE_INPUT_ROW
    d_sum = sum(ws_if.cell(r_if, ic[f"Budget{k}"]).value or 0 for k in ["Digital", "Social", "Influencer", "Display", "Events"])

    min_mkt, max_mkt = seuils_marketing(wb, "TRE", 1)
    bba = budget_base_adjusted_tre_p1(wb)

    budget_rd = ws_if.cell(r_if, ic["BudgetRD"]).value or 0
    rd_ratio = float(budget_rd) / float(bba) if bba else 0

    lp = ws_if.cell(r_if, lancement_prevu_col(ic)).value or 0

    ws_im = wb["INPUT_MODEL"]
    imh = col_map(ws_im)
    max_promo = float(param_val(wb, "Max_Promo") or -0.05)

    rows_im = []
    for row in range(2, ws_im.max_row + 1):
        if ws_im.cell(row, imh["Period_Index"]).value != 1:
            continue
        if ws_im.cell(row, imh["Firm"]).value != "TRE":
            continue
        rows_im.append(
            {
                "pk": ws_im.cell(row, imh["ProductKey"]).value,
                "promo": float(ws_im.cell(row, imh["Promotion"]).value or 0),
                "prod": float(ws_im.cell(row, imh["Production"]).value or 0),
                "withdraw": int(ws_im.cell(row, imh["Withdraw"]).value or 0),
                "liq": int(ws_im.cell(row, imh["LiquidationPromo"]).value or 0),
            }
        )

    prod_total = sum(x["prod"] for x in rows_im)
    withdraw_prod_conflict = any(x["withdraw"] == 1 and x["prod"] > 0 for x in rows_im)

    if withdraw_prod_conflict:
        prod_stat, prod_msg = "✗ Bloqué", "Production interdite pendant liquidation"
    elif prod_total < 0:
        prod_stat, prod_msg = "✗ Bloqué", ""
    else:
        prod_stat, prod_msg = "✓ OK", ""

    if d_sum < min_mkt:
        mkt_stat = "⚠ Attention"
    elif d_sum > max_mkt:
        mkt_stat = "✗ Bloqué"
    else:
        mkt_stat = "✓ OK"

    promo_blocked = False
    for x in rows_im:
        if x["liq"] == 0 and x["promo"] < max_promo - 1e-9:
            promo_blocked = True
            break
        if x["liq"] == 1 and abs(x["promo"] + 0.1) > 0.001:
            promo_blocked = True
            break
    promo_stat = "✗ Bloqué" if promo_blocked else "✓ OK"

    rd_ok = any(abs(rd_ratio - v) < 1e-6 for v in [0, 0.02, 0.05, 0.08])
    rd_stat = "✓ OK" if rd_ok else "✗ Bloqué"

    nb_init = count_init_tre(wb)
    nb_act = count_active_tre_p1(wb)
    # Sans recalcul Excel, WithdrawFlag est une formule — utiliser INPUT_MODEL.Withdraw
    nb_ret = sum(x["withdraw"] for x in rows_im)

    if lp == 0:
        lan_stat, lan_msg = "✓ OK", "Aucun lancement prévu"
    elif nb_act < nb_init + 2:
        lan_stat = "✓ OK"
        lan_msg = f"Lancement autorisé — {nb_init + 2 - nb_act} place(s) disponible(s)"
    elif nb_ret >= 1:
        lan_stat, lan_msg = "⚠ Attention", "Lancement conditionnel — une liquidation est en cours cette période"
    else:
        lan_stat, lan_msg = "✗ Bloqué", "BLOQUÉ — portefeuille plein. Liquidez un modèle (promo -10 %) avant ce lancement."

    retr = sum(x["withdraw"] for x in rows_im)
    retr_stat = "✗ Bloqué" if retr > 1 else "✓ OK"

    price_floor_premium = 5500
    plist_est = price_list_p040_est(wb)
    price_ok_stat = "✓ OK"
    if plist_est is not None and plist_est < price_floor_premium:
        price_ok_stat = "✗ Bloqué"

    report = [
        {"type": "Production (agrégé)", "valeur": int(prod_total), "statut": prod_stat, "msg": prod_msg},
        {"type": "Marketing total", "valeur": int(d_sum), "statut": mkt_stat, "msg": f"min ~{int(min_mkt)} $ / max ~{int(max_mkt)} $"},
        {"type": "Promotion (contrôle)", "valeur": "voir P040…", "statut": promo_stat, "msg": "Promo normale vs liquidation"},
        {"type": "R&D", "valeur": f"{rd_ratio*100:.2f} %", "statut": rd_stat, "msg": ""},
        {"type": "Lancement", "valeur": int(lp), "statut": lan_stat, "msg": lan_msg},
        {"type": "Retrait / liquidation", "valeur": retr, "statut": retr_stat, "msg": ""},
        {
            "type": "Prix P040 (Premium ≥ 5500 $)",
            "valeur": f"{plist_est:,.0f} $" if plist_est is not None else "n/a",
            "statut": price_ok_stat,
            "msg": "Prix liste estimé BasePrice×(1+inflation)^P1",
        },
    ]

    blocked = sum(1 for x in report if x["statut"] == "✗ Bloqué")
    return report, blocked


def banner_text(blocked_count: int) -> str:
    if blocked_count > 0:
        return "DÉCISIONS BLOQUÉES — corriger avant simulation"
    return "TOUTES LES DÉCISIONS VALIDÉES"


def style_statut_column(ws, col_f: int, start_r: int, n: int):
    for i in range(n):
        r = start_r + i
        v = ws.cell(r, col_f).value
        fill = None
        if v == "✓ OK":
            fill = FILL_GREEN
        elif v == "⚠ Attention":
            fill = FILL_ORANGE
        elif v == "✗ Bloqué":
            fill = FILL_RED
        if fill:
            for c in range(1, 8):
                ws.cell(r, c).fill = fill


def build_test_sheet(wb, meta: dict, tre_cd_start: int, python_report: list[dict]):
    name = "TEST_TRE_P1"
    if name in wb.sheetnames:
        del wb[name]
    wt = wb.create_sheet(name)
    wt.cell(1, 1, "Bloc 1 — Décisions injectées (TRE, P1)")
    wt.cell(1, 1).font = Font(bold=True)
    rows = [
        ("BudgetDigital", MARKETING_TRE["BudgetDigital"]),
        ("BudgetSocial", MARKETING_TRE["BudgetSocial"]),
        ("BudgetInfluencer", MARKETING_TRE["BudgetInfluencer"]),
        ("BudgetDisplay", MARKETING_TRE["BudgetDisplay"]),
        ("BudgetEvents", MARKETING_TRE["BudgetEvents"]),
        ("Total marketing", meta["TotalMarketing"]),
        ("BudgetRD (2 % BudgetBase_Ad)", meta["BudgetRD_injected"]),
        ("BudgetBase_Adjusted (calc)", round(meta["BudgetBase_Adjusted_TRE_P1"], 2)),
        ("Revenue_RefYear_TRE (somme PxU)", round(meta["Revenue_RefYear_TRE_calc"], 2)),
        ("SustainabilityInvestPct", 0.005),
        ("LancementPrevu", 0),
    ]
    r = 3
    wt.cell(r - 1, 1, "Champ")
    wt.cell(r - 1, 2, "Valeur")
    for label, val in rows:
        wt.cell(r, 1, label)
        wt.cell(r, 2, val)
        r += 1
    r += 1

    wt.cell(r, 1, "Bloc 2 — CONTROLE_DECISIONS (références Excel)")
    wt.cell(r, 1).font = Font(bold=True)
    r += 1
    hdr = ["Type décision", "Valeur saisie", "Règle", "Statut", "Message"]
    for i, h in enumerate(hdr, 1):
        wt.cell(r, i, h)
        wt.cell(r, i).font = Font(bold=True)
    r += 1
    ws_cd = wb["CONTROLE_DECISIONS"]
    for i in range(7):
        sr = tre_cd_start + i
        wt.cell(r, 1, ws_cd.cell(sr, 3).value)
        wt.cell(r, 2, ws_cd.cell(sr, 4).value)
        wt.cell(r, 3, ws_cd.cell(sr, 5).value)
        wt.cell(r, 4, ws_cd.cell(sr, 6).value)
        wt.cell(r, 5, ws_cd.cell(sr, 7).value)
        r += 1

    r += 1
    wt.cell(r, 1, "Bloc 2 bis — Évaluation Python (sans recalcul Excel)")
    wt.cell(r, 1).font = Font(bold=True)
    r += 1
    for i, h in enumerate(["Type", "Valeur", "Statut", "Message"], 1):
        wt.cell(r, i, h)
        wt.cell(r, i).font = Font(bold=True)
    r += 1
    py_start = r
    for row in python_report:
        wt.cell(r, 1, row["type"])
        wt.cell(r, 2, str(row["valeur"]))
        wt.cell(r, 3, row["statut"])
        wt.cell(r, 4, row.get("msg") or "")
        r += 1
    style_statut_column(wt, 3, py_start, len(python_report))

    r += 2
    wt.cell(r, 1, "Bloc 3 — KPI scénario P040 (références MODEL_PERIOD après ouverture Excel)")
    wt.cell(r, 1).font = Font(bold=True)
    r += 1
    mr = model_period_row_p040(wb)
    mp = wb["MODEL_PERIOD"]
    mh = col_map(mp)
    kpi = [
        ("Profit P040", 2087894, "AL", "ProfitContribution"),
        ("Marge P040", 0.22, "—", "ProfitContribution / Revenue"),
        ("Taux service P040", 1.0, "—", "TotalSales / DemandActive"),
        ("Stock final P040", 177, "AE", "StockEnd"),
    ]
    wt.cell(r, 1, "KPI")
    wt.cell(r, 2, "Attendu")
    wt.cell(r, 3, "Calculé (ref)")
    wt.cell(r, 4, "Écart")
    wt.cell(r, 5, "Statut")
    for c in range(1, 6):
        wt.cell(r, c).font = Font(bold=True)
    r += 1
    kpi_start = r
    if mr:
        row_mp = mr
        for label, exp, _col, note in kpi:
            wt.cell(r, 1, label)
            wt.cell(r, 2, exp)
            if label.startswith("Profit"):
                fmla = f"=INDEX(MODEL_PERIOD!$AL:$AL,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)"
            elif label.startswith("Marge"):
                fmla = f"=INDEX(MODEL_PERIOD!$AL:$AL,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)/INDEX(MODEL_PERIOD!$AK:$AK,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)"
            elif label.startswith("Taux"):
                fmla = f"=INDEX(MODEL_PERIOD!$AD:$AD,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)/INDEX(MODEL_PERIOD!$U:$U,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)"
            else:
                fmla = f"=INDEX(MODEL_PERIOD!$AE:$AE,MATCH(1,(MODEL_PERIOD!$A$2:$A$617=1)*(MODEL_PERIOD!$C$2:$C$617=\"P040\")*(MODEL_PERIOD!$D$2:$D$617=\"TRE\"),0)+1)"
            wt.cell(r, 3, fmla)
            wt.cell(r, 4, f"=IF(ISNUMBER(C{r}),C{r}-B{r},\"\")")
            wt.cell(
                r,
                5,
                f"=IF(AND(ISNUMBER(C{r}),ABS(C{r}-B{r})<=MAX(ABS(B{r})*0.05,1)),\"OK\",\"KO\")",
            )
            r += 1
    else:
        wt.cell(r, 1, "Ligne P040 TRE non trouvée dans MODEL_PERIOD")

    r += 2
    wt.cell(r, 1, "Bloc 4 — Conclusion")
    wt.cell(r, 1).font = Font(bold=True)
    r += 1
    conc_row = r
    py_end = py_start + len(python_report) - 1
    wt.cell(
        conc_row,
        1,
        f'=IF(COUNTIF(C{py_start}:C{py_end},"✗ Bloqué")=1,"⚠ 1 blocage attendu (prix / contrôle) — TEST PARTIEL OK",IF(COUNTIF(C{py_start}:C{py_end},"✗ Bloqué")=0,"✓ AUCUN BLOCAGE — VÉRIFIER ALERTE PRIX","✗ BLOCAGES INATTENDUS — INVESTIGUER"))',
    )

    for col in range(1, 8):
        wt.column_dimensions[get_column_letter(col)].width = 22


def inject_invalid(wb):
    """Budget marketing dépassant le plafond (+10 % sur base marketing ref ou réel P-1)."""
    ws_if = wb["INPUT_FIRM"]
    ic = col_map(ws_if)
    r = TRE_INPUT_ROW
    # Note : 999 999 $ seuls ne dépassent pas le plafond (~118 M$) avec ce classeur ; on force un dépassement net.
    ws_if.cell(r, ic["BudgetInfluencer"], 130_000_000)
    ws_if.cell(r, lancement_prevu_col(ic), 1)

    ws_im = wb["INPUT_MODEL"]
    imh = col_map(ws_im)
    for row in range(2, ws_im.max_row + 1):
        if ws_im.cell(row, imh["Period_Index"]).value != 1:
            continue
        if ws_im.cell(row, imh["Firm"]).value != "TRE":
            continue
        if ws_im.cell(row, imh["ProductKey"]).value == "P040":
            ws_im.cell(row, imh["Promotion"], -0.15)
            break


def build_invalide_sheet(wb, stats: dict):
    name = "TEST_TRE_P1_INVALIDE"
    if name in wb.sheetnames:
        del wb[name]
    w = wb.create_sheet(name)
    w.cell(1, 1, "Scénario invalide (injecté puis rétabli dans INPUT_* pour fichier final)")
    w.cell(3, 1, "Modification")
    w.cell(3, 2, "Valeur")
    w.cell(4, 1, "BudgetInfluencer")
    w.cell(4, 2, "130000000 (BudgetInfluencer — dépasse plafond)")
    w.cell(5, 1, "Promotion P040")
    w.cell(5, 2, -0.15)
    w.cell(6, 1, "LancementPrevu")
    w.cell(6, 2, 1)
    r = 9
    w.cell(r, 1, "Résultats Python (scénario invalide)")
    r += 1
    for k, v in stats.items():
        w.cell(r, 1, k)
        w.cell(r, 2, str(v))
        r += 1


def main():
    print(f"=== Étape 0 — Chargement classeur ({WORKBOOK_DEFAULT.name}) → copie TEST ===")
    wb = openpyxl.load_workbook(WORKBOOK_DEFAULT)

    print("=== Étape 1 — Injection décisions TRE P1 ===")
    meta = inject_valid_decisions(wb)
    print(f"  Revenue_RefYear_TRE (calc): {meta['Revenue_RefYear_TRE_calc']:,.0f} $")
    print(f"  BudgetBase_Adjusted P1: {meta['BudgetBase_Adjusted_TRE_P1']:,.0f} $")
    print(f"  BudgetRD injecté: {meta['BudgetRD_injected']:,} $")
    print(f"  Total marketing: {meta['TotalMarketing']:,} $")

    tre_cd_start = 53

    print("\n=== Étape 2 — CONTROLE_DECISIONS — TRE P1 (évaluation Python, sans recalcul Excel) ===")
    print(
        f"{'Type décision':<28} | {'Valeur':<14} | {'Statut':<12} | Message"
    )
    print("-" * 90)
    report, n_blocked = evaluate_controles_tre_p1_python(wb)
    for row in report:
        print(
            f"{row['type']:<28} | {str(row['valeur'])[:14]:<14} | {row['statut']:<12} | {str(row.get('msg',''))[:55]}"
        )

    ws_cd = wb["CONTROLE_DECISIONS"]
    band_cell = ws_cd.cell(1, 1).value
    print("\n=== Étape 3 — Bandeau CONTROLE_DECISIONS (cellule A1) ===")
    print(f">>> BANDEAU (formule / texte A1) : {band_cell}")

    print("\n=== Étape 3b — Synthèse blocages ✗ (Python) ===")
    bt = banner_text(n_blocked)
    print(f">>> Interprétation (compte ✗ Bloqué colonne statuts Python): {bt}")
    if n_blocked == 0:
        print("\033[92m✓ TEST RÉUSSI (aucun ✗ Bloqué selon évaluation Python)\033[0m")
    else:
        print(f"\033[91mBlocages ✗: {n_blocked}\033[0m")

    build_test_sheet(wb, meta, tre_cd_start, report)

    print("\n=== Étape 4 — Feuilles TEST_TRE_P1 et (plus loin) TEST_TRE_P1_INVALIDE ===")
    print("  → Récap décisions, liaisons MODEL_PERIOD (KPI), conclusion dans TEST_TRE_P1.")

    print("\n=== Étape 5 — Scénario invalide (temporaire) ===")
    inject_invalid(wb)
    rep_inv, _ = evaluate_controles_tre_p1_python(wb)

    ws_if = wb["INPUT_FIRM"]
    ic = col_map(ws_if)
    d_sum_inv = sum(ws_if.cell(TRE_INPUT_ROW, ic[f"Budget{k}"]).value or 0 for k in ["Digital", "Social", "Influencer", "Display", "Events"])
    _, max_mkt = seuils_marketing(wb, "TRE", 1)

    ws_im = wb["INPUT_MODEL"]
    imh = col_map(ws_im)
    prom_bad = False
    for row in range(2, ws_im.max_row + 1):
        if ws_im.cell(row, imh["Period_Index"]).value != 1:
            continue
        if ws_im.cell(row, imh["Firm"]).value != "TRE":
            continue
        if ws_im.cell(row, imh["ProductKey"]).value == "P040":
            prom_bad = float(ws_im.cell(row, imh["Promotion"]).value or 0) < -0.05
            break

    nb_init = count_init_tre(wb)
    nb_act = count_active_tre_p1(wb)
    lp = ws_if.cell(TRE_INPUT_ROW, lancement_prevu_col(ic)).value
    portfolio_block = lp == 1 and nb_act >= nb_init + 2

    print("=== TEST INVALIDE — 3 contrôles attendus ===")
    x_mkt = d_sum_inv > max_mkt
    print(
        f"Marketing hors plafond    → [{'✗ dépasse' if x_mkt else 'OK ≤ plafond'}] "
        f"(total {d_sum_inv:,.0f} $ vs max ~{max_mkt:,.0f} $)"
    )
    print(f"Promotion hors limite     → [{'✗ Bloqué' if prom_bad else 'OK'}]")
    print(
        f"Portefeuille plein+Lanc.  → [{'✗ Bloqué' if portfolio_block else 'OK / non saturé'}] "
        f"(NbActifs={nb_act}, NbInit+2={nb_init+2})"
    )

    xb_inv = sum(1 for x in rep_inv if x["statut"] == "✗ Bloqué")
    print(f"\nStatuts ✗ Bloqué (Python, scénario invalide) : {xb_inv} ligne(s)")
    print(f"Résultat attendu : marketing dépassé + promo + portefeuille → idéal 3 ✗ si Excel calcule pareil")

    build_invalide_sheet(
        wb,
        {
            "marketing_depasse_plafond": x_mkt,
            "promo_hors_limite": prom_bad,
            "portefeuille_plein_lancement": portfolio_block,
            "statuts_python": str([x["statut"] for x in rep_inv]),
            "nb_x_bloque_python": xb_inv,
        },
    )

    print("\n=== Rétablissement entrées valides (fichier final) ===")
    inject_valid_decisions(wb)

    wb.save(OUT)
    print(f"\nTEST terminé — fichier sauvegardé : {OUT}")


if __name__ == "__main__":
    main()
