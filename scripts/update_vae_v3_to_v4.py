"""
V3 → V4 : colonne LancementPrevu (INPUT_FIRM), vérification portefeuille (CONTROLE_DECISIONS), FORMULES_FR.
Ne modifie que INPUT_FIRM, CONTROLE_DECISIONS, FORMULES_FR.
"""
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

INPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v3.xlsx")
OUTPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v4.xlsx")

HEADER_TEXT = "LancementPrevu (0=non / 1=oui)"
HEADER_COMMENT = (
    "Mettre 1 si la firme prévoit de lancer un nouveau modèle cette période. "
    "Déclenche la vérification de capacité du portefeuille."
)


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)

    ws_in = wb["INPUT_FIRM"]
    # Insérer après SustainabilityInvestPct (colonne 10) → nouvelle colonne 11
    ws_in.insert_cols(11, 1)
    h = ws_in.cell(1, 11, HEADER_TEXT)
    h.comment = Comment(HEADER_COMMENT, "Moteur VAE")
    for r in range(2, ws_in.max_row + 1):
        ws_in.cell(r, 11, 0)

    ws_cd = wb["CONTROLE_DECISIONS"]
    for r in range(4, ws_cd.max_row + 1):
        typ = ws_cd.cell(r, 3).value
        if typ not in ("Lancement NP", "Lancement"):
            continue
        ws_cd.cell(r, 3, "Lancement")
        ws_cd.cell(
            r,
            4,
            f"=INDEX(INPUT_FIRM!$K$2:$K$73,MATCH(1,(INPUT_FIRM!$A$2:$A$73=$B{r})*(INPUT_FIRM!$C$2:$C$73=$A{r}),0))",
        )
        ws_cd.cell(
            r,
            5,
            "NbActifs vs NbInitiaux+2 ; retraits période si LancementPrevu=1",
        )
        # Statut
        ws_cd.cell(
            r,
            6,
            (
                f'=IF(D{r}=0,"✓ OK",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'<COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2),'
                f'"✓ OK",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'>=COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$N:$N,1)>=1),'
                f'"⚠ Attention",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'>=COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$N:$N,1)=0),'
                f'"✗ Bloqué","✓ OK"))))'
            ),
        )
        # Message
        ws_cd.cell(
            r,
            7,
            (
                f'=IF(D{r}=0,"Aucun lancement prévu",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'<COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2),'
                f'"Lancement autorisé — "&(COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2-'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE"))'
                f'&" place(s) disponible(s)",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'>=COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$N:$N,1)>=1),'
                f'"Lancement conditionnel — une liquidation est en cours cette période",'
                f'IF(AND(D{r}=1,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$G:$G,"ACTIVE")'
                f'>=COUNTIF(BASE_REFERENCE_MODEL!$B:$B,$A{r})+2,'
                f'COUNTIFS(MODEL_PERIOD!$D:$D,$A{r},MODEL_PERIOD!$A:$A,$B{r},MODEL_PERIOD!$N:$N,1)=0),'
                f'"BLOQUÉ — portefeuille plein. Liquidez un modèle (promo -10 %) avant ce lancement.",'
                f'""))))'
            ),
        )

    ws_fr = wb["FORMULES_FR"]
    next_row = ws_fr.max_row + 1
    ws_fr.cell(next_row, 1, 24)
    ws_fr.cell(next_row, 2, "CONTROLE_DECISIONS / INPUT_FIRM")
    ws_fr.cell(next_row, 3, "LancementPrevu")
    ws_fr.cell(
        next_row,
        4,
        "Vérification lancement — si LancementPrevu=1 dans INPUT_FIRM, le moteur contrôle NbActifs vs NbInitiaux+2 "
        "avant d'autoriser le lancement. Un retrait en cours la même période est accepté (lancement conditionnel).",
    )
    ws_fr.cell(next_row, 5, "Voir CONTROLE_DECISIONS, type Lancement")
    ws_fr.cell(next_row, 6, "")

    wb.save(OUTPUT_FILE)
    print("V4 terminée — colonne LancementPrevu + vérification 4 CONTROLE_DECISIONS")


if __name__ == "__main__":
    main()
