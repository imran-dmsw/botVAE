"""Feuilles « Pxxx Scénario » : paramètres P2, KPIs et tableau de référence 2 250 u."""
from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from engine.excel_reference_2250 import (
    REFERENCE_2250_HEADERS,
    compute_reference_2250_row,
)
from engine.simulation import simulate
from generate_rapport_vae_style_pdf import ProdRow, firm_ref_revenue, scenario_from_product

log = logging.getLogger("excel_2250_reference")

BLOCK_MARKER = "Simulation de référence — Production fixée à 2 250 unités"

HEADER_FILL = PatternFill("solid", fgColor="1A82D4")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10, color="000000")
TITLE_FILL = PatternFill("solid", fgColor="0D1B2A")
TITLE_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F4F6F8")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
NEG_FONT = Font(name="Arial", size=10, color="C0392B")

SCENARIO_SHEET_RE = re.compile(r"^P\d{3}\s+Scénario$")


def header_map(ws: Any, row: int = 1) -> dict[str, int]:
    return {
        str(ws.cell(row, c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row, c).value
    }


def parse_segment_idx(seg_val: Any) -> int:
    s = str(seg_val or "").strip()
    if "—" in s:
        return int(str(s.split("—")[0]).strip())
    if " - " in s:
        return int(str(s.split(" - ")[0]).strip())
    raise ValueError(f"Segment illisible : {seg_val!r}")


def read_products_from_portfolio(ws: Any) -> list[ProdRow]:
    h = header_map(ws)
    name_col = "Nom produit" if "Nom produit" in h else "Nom modèle"
    need = ["Code", name_col, "Segment", "Gamme", "Prix réf.", "Unités réf."]
    for k in need:
        if k not in h:
            raise KeyError(f"Colonne manquante en 1.5 Portefeuille : {k}")
    cols = {"Code": h["Code"], "Nom": h[name_col], "Segment": h["Segment"], "Gamme": h["Gamme"], "Prix réf.": h["Prix réf."], "Unités réf.": h["Unités réf."]}
    out: list[ProdRow] = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, cols["Code"]).value
        if code is None or str(code).strip() == "":
            continue
        out.append(
            ProdRow(
                product_key=str(code).strip(),
                model_name=str(ws.cell(r, cols["Nom"]).value or ""),
                segment_idx=parse_segment_idx(ws.cell(r, cols["Segment"]).value),
                range_raw=str(ws.cell(r, cols["Gamme"]).value or ""),
                base_price=float(ws.cell(r, cols["Prix réf."]).value or 0),
                units=float(ws.cell(r, cols["Unités réf."]).value or 0),
            )
        )
    return out


def remove_outdated_reference_block(ws: Any) -> None:
    """Supprime un bloc 2 250 u. à l'ancien ordre de colonnes (CA en 3e position)."""
    for r in range(1, min(ws.max_row + 1, 800)):
        v = ws.cell(r, 1).value
        if not isinstance(v, str) or BLOCK_MARKER not in v:
            continue
        next_r = r + 1
        if next_r > ws.max_row:
            return
        if ws.cell(next_r, 3).value == "CA ($)":
            ws.delete_rows(r, 4)
        return


def sheet_has_reference_block(ws: Any) -> bool:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and BLOCK_MARKER in v:
                return True
    return False


def last_used_row(ws: Any) -> int:
    last = 1
    for row in ws.iter_rows():
        for c in row:
            if c.value not in (None, ""):
                last = max(last, c.row)
    return last


def _write_styled_table_header(ws: Any, r: int, labels: tuple[str, ...] | list[str]) -> None:
    for i, lab in enumerate(labels, start=1):
        cell = ws.cell(r, i, lab)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_reference_data_row(
    ws: Any,
    r: int,
    scenario_txt: str,
    d: dict[str, float | int],
) -> None:
    """Colonnes : Scénario, Ventes, Profit, Taux de service, CA, Marge, Stock."""
    ws.cell(r, 1, scenario_txt)
    ws.cell(r, 2, d["ventes"])
    ws.cell(r, 3, d["profit"])
    ws.cell(r, 4, d["service"])
    ws.cell(r, 5, d["ca"])
    ws.cell(r, 6, d["marge"])
    ws.cell(r, 7, d["stock"])

    for c in range(1, 8):
        ws.cell(r, c).font = BODY_FONT
        ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.cell(r, 2).number_format = "#,##0"
    ws.cell(r, 3).number_format = "#,##0.00"
    ws.cell(r, 4).number_format = "0.00%"
    ws.cell(r, 5).number_format = "#,##0.00"
    ws.cell(r, 6).number_format = "0.00%"
    ws.cell(r, 7).number_format = "#,##0"

    for c in range(1, 8):
        ws.cell(r, c).fill = ALT_FILL

    if float(d["profit"]) < 0:
        ws.cell(r, 3).font = NEG_FONT
    if float(d["marge"]) < 0:
        ws.cell(r, 6).font = NEG_FONT


def write_param_kpi_blocks(ws: Any, start_row: int, prod: ProdRow, scen: Any, res: Any) -> int:
    r = start_row
    ws.cell(r, 1, "Paramètres (P2 — moteur)")
    ws.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
    r += 1
    params = [
        ("Produit", prod.product_key),
        ("Nom", prod.model_name),
        ("Prix catalogue P2 ($)", scen.price),
        ("Production simulée (u.)", scen.production),
        ("Marketing alloué ($)", scen.marketing_budget),
        ("R&D allouée ($)", scen.rd_budget),
    ]
    for lab, val in params:
        ws.cell(r, 1, lab)
        ws.cell(r, 2, val)
        ws.cell(r, 1).font = BODY_FONT
        ws.cell(r, 2).font = BODY_FONT
        if isinstance(val, float):
            ws.cell(r, 2).number_format = "#,##0.00"
        r += 1

    r += 1
    ws.cell(r, 1, "KPIs — simulation P2")
    ws.cell(r, 1).font = Font(name="Arial", size=10, bold=True)
    r += 1
    kpis = [
        ("Ventes (u.)", res.sales),
        ("CA ($)", res.revenue),
        ("Profit ($)", res.profit),
        ("Marge", res.margin),
        ("Taux de service", res.service_rate),
    ]
    for lab, val in kpis:
        ws.cell(r, 1, lab)
        ws.cell(r, 2, val)
        ws.cell(r, 1).font = BODY_FONT
        ws.cell(r, 2).font = BODY_FONT
        if lab in ("Marge", "Taux de service"):
            ws.cell(r, 2).number_format = "0.00%"
        elif lab == "Ventes (u.)":
            ws.cell(r, 2).number_format = "#,##0"
        else:
            ws.cell(r, 2).number_format = "#,##0.00"
        if lab == "Profit ($)" and isinstance(val, (int, float)) and val < 0:
            ws.cell(r, 2).font = NEG_FONT
        if lab == "Marge" and isinstance(val, (int, float)) and val < 0:
            ws.cell(r, 2).font = NEG_FONT
        r += 1
    return r + 1


def write_reference_2250_section(
    ws: Any,
    start_row: int,
    prod: ProdRow,
    scen: Any,
    margin_warnings: list[str],
) -> None:
    d = compute_reference_2250_row(
        units_ref=float(prod.units),
        price_p2=float(scen.price),
        marketing_alloc=float(scen.marketing_budget),
        rd_alloc=float(scen.rd_budget),
        range_raw=str(prod.range_raw),
    )
    marge = float(d["marge"])
    if marge < -0.5 or marge > 0.5:
        msg = f"{prod.product_key}: marge pédagogique hors plage ±50 % ({marge:.2%})."
        log.warning(msg)
        margin_warnings.append(msg)

    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    tcell = ws.cell(r, 1, BLOCK_MARKER)
    tcell.fill = TITLE_FILL
    tcell.font = TITLE_FONT
    tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rd = ws.row_dimensions.get(r)
    if rd is None or rd.height is None:
        ws.row_dimensions[r].height = 22
    r += 1

    _write_styled_table_header(ws, r, REFERENCE_2250_HEADERS)
    r += 1
    scenario_txt = f"Prod. 2 250 u. — {prod.model_name}"
    _write_reference_data_row(ws, r, scenario_txt, d)
    r += 1

    note = (
        "Estimation pédagogique — moteur simplifié (COGS + distribution + budgets fixes). "
        "Comparer avec le scénario de référence P2 ci-dessus pour mesurer l'impact "
        "d'une production normalisée à 2 250 unités."
    )
    ws.cell(r, 1, note)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(r, 1).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")

    for c in range(1, 8):
        col = get_column_letter(c)
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 14)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 38)


def enrich_workbook_product_scenarios(wb: Workbook, firm: str) -> dict[str, Any]:
    """Crée ou complète les feuilles « Pxxx Scénario » à partir de « 1.5 Portefeuille »."""
    if "1.5 Portefeuille" not in wb.sheetnames:
        raise RuntimeError("Feuille « 1.5 Portefeuille » introuvable.")
    prods = read_products_from_portfolio(wb["1.5 Portefeuille"])
    adj = firm_ref_revenue(prods)
    product_keys = {f"P{n:03d}" for n in range(35, 44)}
    targets = [p for p in prods if p.product_key in product_keys]
    targets.sort(key=lambda x: x.product_key)

    out: dict[str, Any] = {"sheets_touched": [], "margin_warnings": []}

    for prod in targets:
        sheet_name = f"{prod.product_key} Scénario"
        scen = scenario_from_product(
            prod,
            firm,
            2,
            f"{prod.product_key}_P2ref",
            adj,
            products=prods,
            production_mult=1.08,
            sustainability_tranches=2,
        )
        res = simulate(scen)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            remove_outdated_reference_block(ws)
            if sheet_has_reference_block(ws):
                continue
            start = last_used_row(ws) + 2
        else:
            ws = wb.create_sheet(sheet_name)
            start = write_param_kpi_blocks(ws, 1, prod, scen, res)

        write_reference_2250_section(ws, start, prod, scen, out["margin_warnings"])
        out["sheets_touched"].append(sheet_name)

    return out
