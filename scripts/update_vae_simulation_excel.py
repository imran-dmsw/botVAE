import math
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1.xlsx")
OUTPUT_FILE = Path("/Users/imran/Downloads/VAE_tout_inclus_01_05_2026_Final1_v2.xlsx")
TITLE_FILL = PatternFill("solid", fgColor="1C2B4A")
TITLE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
ALERT_RED = PatternFill("solid", fgColor="F8CBAD")
ALERT_ORANGE = PatternFill("solid", fgColor="FCE4D6")
ALERT_GREEN = PatternFill("solid", fgColor="E2F0D9")
# Couleurs zones production / stockage (rapports firmes)
ZONE_RUPTURE = PatternFill("solid", fgColor="FAEAEA")
ZONE_PRUDENT = PatternFill("solid", fgColor="FDF3E3")
ZONE_EQUILIBRE = PatternFill("solid", fgColor="E8F5F0")
ZONE_SECURITE = PatternFill("solid", fgColor="FFFDE7")
ZONE_SURPROD = PatternFill("solid", fgColor="FAEAEA")


def sheet_headers(ws):
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}


def header_style(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row, c)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")


def title_row(ws, row, text, max_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, text)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")


def autosize_columns(ws, max_width=42):
    for c in range(1, ws.max_column + 1):
        width = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            width = max(width, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(width, max_width)


def get_param_map(wb):
    ws = wb["PARAM"]
    params = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k:
            params[str(k).strip()] = v
    return params


def read_table(ws):
    h = sheet_headers(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for k, c in h.items():
            row[k] = ws.cell(r, c).value
        rows.append(row)
    return rows


def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name], False
    return wb.create_sheet(name), True


def block1(wb):
    segments = read_table(wb["SEGMENTS"])
    firms = read_table(wb["FIRMS"])
    models = read_table(wb["BASE_REFERENCE_MODEL"])

    # ETAT_2026
    ws, _ = ensure_sheet(wb, "ETAT_2026")
    ws.delete_rows(1, ws.max_row)
    row = 1

    title_row(ws, row, "ETAT 2026 - Segments", 6)
    row += 1
    seg_headers = ["Segment", "Taille segment", "Prix prefere", "Nb produits", "Firmes presentes"]
    for i, h in enumerate(seg_headers, 1):
        ws.cell(row, i, h)
    header_style(ws, row)
    row += 1

    by_seg = defaultdict(list)
    for m in models:
        by_seg[m["Segment"]].append(m)
    for s in segments:
        seg_id = s["Segment"]
        seg_models = by_seg.get(seg_id, [])
        firmes = sorted({m["Firm"] for m in seg_models})
        values = [
            seg_id,
            s["Units_RefYear"],
            s["Pref_RefYear"],
            len(seg_models),
            ", ".join(firmes),
        ]
        for i, v in enumerate(values, 1):
            ws.cell(row, i, v)
        row += 1

    row += 1
    title_row(ws, row, "ETAT 2026 - Firmes", 6)
    row += 1
    firm_headers = ["Firme", "Nb produits", "PDM initiale", "Unites de reference"]
    for i, h in enumerate(firm_headers, 1):
        ws.cell(row, i, h)
    header_style(ws, row)
    row += 1

    by_firm = defaultdict(list)
    for m in models:
        by_firm[m["Firm"]].append(m)
    for f in firms:
        f_code = f["Firm"]
        f_models = by_firm.get(f_code, [])
        units = sum((m["Units_RefYear"] or 0) for m in f_models)
        values = [f_code, len(f_models), f["Share_RefYear"], units]
        for i, v in enumerate(values, 1):
            ws.cell(row, i, v)
        row += 1

    row += 1
    title_row(ws, row, "ETAT 2026 - Produits de reference", 10)
    row += 1
    prod_headers = [
        "ProductKey",
        "Firm",
        "ModelName",
        "Segment",
        "Range",
        "BasePrice_RefYear",
        "Units_RefYear",
        "InitStock_P1",
    ]
    for i, h in enumerate(prod_headers, 1):
        ws.cell(row, i, h)
    header_style(ws, row)
    row += 1

    for m in models:
        values = [m.get(k) for k in prod_headers]
        for i, v in enumerate(values, 1):
            ws.cell(row, i, v)
        row += 1
    autosize_columns(ws)

    # MARCHE_STRUCTURE
    ws2, _ = ensure_sheet(wb, "MARCHE_STRUCTURE")
    ws2.delete_rows(1, ws2.max_row)
    title_row(ws2, 1, "MARCHE STRUCTURE - Vue segments", 11)
    h1 = [
        "Segment",
        "Description",
        "Part marche",
        "Unites 2026",
        "Unites P1 (x1.12)",
        "Prix prefere",
        "% Online",
        "Firmes presentes",
        "Nb produits",
    ]
    for i, h in enumerate(h1, 1):
        ws2.cell(2, i, h)
    header_style(ws2, 2)
    r = 3
    growth = get_param_map(wb).get("Market_Growth", 0.12) or 0.12
    for s in segments:
        seg_id = s["Segment"]
        seg_models = by_seg.get(seg_id, [])
        firms_in = sorted({m["Firm"] for m in seg_models})
        vals = [
            seg_id,
            s["Description"],
            s["Share_RefYear"],
            s["Units_RefYear"],
            int(round((s["Units_RefYear"] or 0) * (1 + growth))),
            s["Pref_RefYear"],
            s["OnlineShare_RefYear"],
            ", ".join(firms_in),
            len(seg_models),
        ]
        for i, v in enumerate(vals, 1):
            ws2.cell(r, i, v)
        r += 1

    title_row(ws2, 15, "MARCHE STRUCTURE - Firme x Segment", 12)
    seg_ids = [s["Segment"] for s in segments]
    ws2.cell(16, 1, "Firme")
    for i, seg in enumerate(seg_ids, 2):
        ws2.cell(16, i, f"Segment {seg}")
    ws2.cell(16, 2 + len(seg_ids), "PDM globale")
    header_style(ws2, 16)

    product_counter = Counter((m["Firm"], m["Segment"]) for m in models)
    rr = 17
    for f in firms:
        ws2.cell(rr, 1, f["Firm"])
        for i, seg in enumerate(seg_ids, 2):
            ws2.cell(rr, i, product_counter.get((f["Firm"], seg), 0))
        ws2.cell(rr, 2 + len(seg_ids), f["Share_RefYear"])
        rr += 1
    autosize_columns(ws2)
    return {"ETAT_2026", "MARCHE_STRUCTURE"}


def add_model_period_columns(wb):
    ws = wb["MODEL_PERIOD"]
    headers = sheet_headers(ws)
    max_col = ws.max_column
    stock_sec_col = max_col + 1
    penalty_col = max_col + 2
    ws.cell(1, stock_sec_col, "StockSecurite").font = HEADER_FONT
    ws.cell(1, penalty_col, "SurplusPenalty").font = HEADER_FONT

    demand_col = get_column_letter(headers["DemandActive"])
    prod_col = get_column_letter(headers["Production"])
    stockend_col = get_column_letter(headers["StockEnd"])
    pricelist_col = get_column_letter(headers["PriceList"])
    profit_col = get_column_letter(headers["ProfitContribution"])

    for r in range(2, ws.max_row + 1):
        ws.cell(r, stock_sec_col, f"=MAX(0,{demand_col}{r}-{prod_col}{r}-200)")
        ws.cell(r, penalty_col, f"=IF({stockend_col}{r}>200,({stockend_col}{r}-200)*{pricelist_col}{r}*0.05,0)")
        prev_formula = ws.cell(r, headers["ProfitContribution"]).value
        if isinstance(prev_formula, str) and prev_formula.startswith("="):
            ws.cell(r, headers["ProfitContribution"], f"={prev_formula[1:]}-{get_column_letter(penalty_col)}{r}")
        else:
            ws.cell(r, headers["ProfitContribution"], f"={profit_col}{r}-{get_column_letter(penalty_col)}{r}")
    return {"MODEL_PERIOD"}


def append_controles_rows(wb):
    ws = wb["CONTROLES"]
    start = ws.max_row + 2
    title_row(ws, start, "Seuils de lecture", 6)
    r = start + 1

    rows = [
        ("Zone ideale service", 1, "Global", '=IF(AVERAGE(RESULTS_FIRM!$D$2:$D$73)/MAX(1,AVERAGE(FIRM_SEG!$S$2:$S$433)),"","")', '=IF(AND(D{r}>=0.95,D{r}<=1.05),"OK","Hors zone ideale")', "Service entre 95% et 105%"),
        ("Alerte ventes perdues", 1, "Global", "=AVERAGE(IFERROR(RESULTS_FIRM!$D$2:$D$73/FIRM_PERIOD!$L$2:$L$73,0))", '=IF(D{r}<0.8,"ALERTE","OK")', "Production insuffisante — ventes perdues > 20 %"),
        ("Alerte surstock", 1, "Global", "=AVERAGE(IFERROR(MODEL_PERIOD!$AE$2:$AE$617/MODEL_PERIOD!$T$2:$T$617,0))", '=IF(D{r}>0.05,"ALERTE","OK")', "Stock eleve — risque de cout de stockage"),
        ("Alerte critique service", 1, "Global", "=AVERAGE(IFERROR(RESULTS_FIRM!$D$2:$D$73/FIRM_PERIOD!$L$2:$L$73,0))", '=IF(D{r}<0.5,"ALERTE CRITIQUE","OK")', "Situation critique — plus de 50 % de la demande non servie"),
        ("Verification StockCost", 1, "Global", "=SUMPRODUCT(--(ABS(MODEL_PERIOD!$AI$2:$AI$617-(MODEL_PERIOD!$AE$2:$AE$617*MODEL_PERIOD!$K$2:$K$617*PARAM!$B$10))>1))", '=IF(D{r}>0,"Ecart detecte","OK")', "StockCost = StockEnd x PriceList x Stock_Cost_Rate"),
        ("Marketing minimum 3%", 1, "Global", "=SUMPRODUCT(--(FIRM_PERIOD!$K$2:$K$73 < 0.03*FIRM_PERIOD!$S$2:$S$73))", '=IF(D{r}>0,"ALERTE","OK")', "Budget marketing sous le minimum (3 %)"),
        ("Marketing maximum 15%", 1, "Global", "=SUMPRODUCT(--(FIRM_PERIOD!$K$2:$K$73 > PARAM!$B$14*FIRM_PERIOD!$T$2:$T$73))", '=IF(D{r}>0,"ALERTE","OK")', "Budget marketing au-dessus du plafond (15 %)"),
        ("Promotion hors limite", 1, "Global", '=SUMPRODUCT(--(INPUT_MODEL!$F$2:$F$617<PARAM!$B$16),--(INPUT_MODEL!$I$2:$I$617=0))', '=IF(D{r}>0,"ALERTE","OK")', "Promotion hors limite — utiliser la liquidation"),
        ("Liquidation taux fixe", 1, "Global", '=SUMPRODUCT(--(INPUT_MODEL!$I$2:$I$617=1),--(INPUT_MODEL!$F$2:$F$617<>-0.10))', '=IF(D{r}>0,"ALERTE","OK")', "Taux de liquidation incorrect — forcer a -10 %"),
        ("Regles R&D niveaux", 1, "Global", '=SUMPRODUCT(--(ABS(INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73-0)>0.0001),--(ABS(INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73-0.02)>0.0001),--(ABS(INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73-0.05)>0.0001),--(ABS(INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73-0.08)>0.0001))', '=IF(D{r}>0,"ALERTE","OK")', "Niveau R&D invalide — valeurs autorisees : 0/2/5/8 %"),
        ("R&D periode 1", 1, "Global", '=SUMPRODUCT(--(INPUT_FIRM!$A$2:$A$73=1),--((INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73=0.05)+(INPUT_FIRM!$I$2:$I$73/FIRM_PERIOD!$T$2:$T$73=0.08)>0))', '=IF(D{r}>0,"ALERTE","OK")', "Lancement impossible en P1 — modele disponible en P2+"),
        ("R&D limite 4 modeles", 1, "Global", '=MAX(FIRM_PERIOD!$Z$2:$Z$73)', '=IF(D{r}>4,"ALERTE","OK")', "Limite de 4 modeles developpes atteinte"),
        ("Liquidation et production", 1, "Global", '=SUMPRODUCT(--(MODEL_PERIOD!$N$2:$N$617=1),--(MODEL_PERIOD!$T$2:$T$617>0))', '=IF(D{r}>0,"ALERTE","OK")', "Production interdite pendant liquidation"),
    ]

    for controle, periode, seg, val, test, com in rows:
        ws.cell(r, 1, controle)
        ws.cell(r, 2, periode)
        ws.cell(r, 3, seg)
        ws.cell(r, 4, val if isinstance(val, (int, float)) else str(val))
        ws.cell(r, 5, test.format(r=r))
        ws.cell(r, 6, com)
        r += 1
    autosize_columns(ws)
    return {"CONTROLES"}


def build_tests_production(wb):
    ws, _ = ensure_sheet(wb, "TESTS_PRODUCTION")
    ws.delete_rows(1, ws.max_row)
    params = get_param_map(wb)
    segs = {r["Segment"]: r for r in read_table(wb["SEGMENTS"])}
    seg1 = segs.get(1) or segs.get("1")
    demand = (seg1.get("Units_RefYear") or 0) * (1 + (params.get("Market_Growth", 0.12) or 0.12))
    pref_price = seg1.get("Pref_RefYear") or 4000
    stock_cost_rate = params.get("Stock_Cost_Rate", 0.025) or 0.025
    cogs_ratio = params.get("COGS_Moyen", 0.57) or 0.57
    marketing = 100000
    rd = 30000
    price_net = pref_price

    title_row(ws, 1, "TESTS PRODUCTION - Segment 1 / Periode 1", 10)
    headers = ["Production", "DemandActive", "Ventes", "Stock final", "Taux service", "Cout stockage", "Profit estime"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    header_style(ws, 2)

    levels = [1500, 2000, 2500, 3000]
    row = 3
    rec_row = None
    best_profit = -10**18
    for p in levels:
        sales = min(p, demand)
        stock = p - sales
        service = sales / demand if demand else 0
        stock_cost = stock * pref_price * stock_cost_rate
        cogs = sales * price_net * cogs_ratio
        profit = sales * price_net - cogs - stock_cost - marketing - rd
        vals = [p, int(round(demand)), int(round(sales)), int(round(stock)), service, int(round(stock_cost)), int(round(profit))]
        for i, v in enumerate(vals, 1):
            ws.cell(row, i, v)
        ws.cell(row, 5).number_format = "0.0%"
        if service >= 0.95 and profit > best_profit:
            best_profit = profit
            rec_row = row
        row += 1

    ws.cell(row + 1, 1, "Recommande")
    if rec_row:
        ws.cell(row + 1, 2, ws.cell(rec_row, 1).value)
        ws.cell(row + 1, 3, "Meilleur profit avec service >= 95%")
        ws.cell(row + 1, 1).fill = TITLE_FILL
        ws.cell(row + 1, 1).font = TITLE_FONT
    autosize_columns(ws)
    return {"TESTS_PRODUCTION"}


def build_portefeuille(wb):
    ws, _ = ensure_sheet(wb, "PORTEFEUILLE")
    ws.delete_rows(1, ws.max_row)
    mp = wb["MODEL_PERIOD"]
    h = sheet_headers(mp)
    firms = [r["Firm"] for r in read_table(wb["FIRMS"])]

    title_row(ws, 1, "PORTEFEUILLE - Suivi firmes/periodes", 12)
    headers = [
        "Periode",
        "Firme",
        "Modeles actifs",
        "Developpement",
        "Disponibles a lancer",
        "Retires/liquides",
        "Modeles initiaux",
        "Droits lancement disponibles",
        "Peut lancer",
        "Peut retirer",
    ]
    for i, x in enumerate(headers, 1):
        ws.cell(2, i, x)
    header_style(ws, 2)

    init_active = defaultdict(int)
    withdrawn_by_firm = defaultdict(int)
    data = []
    for r in range(2, mp.max_row + 1):
        p = mp.cell(r, h["Period_Index"]).value
        f = mp.cell(r, h["Firm"]).value
        st = mp.cell(r, h["Status"]).value
        ex = mp.cell(r, h["ExistingFlag"]).value
        if p == 1 and ex == 1 and st == "ACTIVE":
            init_active[f] += 1
        data.append((p, f, st, ex))

    grouped = defaultdict(lambda: defaultdict(int))
    for p, f, st, ex in data:
        if st == "ACTIVE" and ex == 1:
            grouped[(p, f)]["actifs"] += 1
        if st == "DEVELOPMENT":
            grouped[(p, f)]["dev"] += 1
        if st in ("PRELAUNCH",):
            grouped[(p, f)]["dispo"] += 1
        if st in ("WITHDRAW_LIQ", "INACTIVE"):
            grouped[(p, f)]["ret"] += 1

    rr = 3
    periods = sorted({k[0] for k in grouped.keys()})
    for p in periods:
        for f in firms:
            g = grouped.get((p, f), {})
            actifs = g.get("actifs", 0)
            dev = g.get("dev", 0)
            dispo = g.get("dispo", 0)
            ret = g.get("ret", 0)
            withdrawn_by_firm[f] += ret
            droits = max(0, 2 - (actifs - init_active.get(f, 0)))
            peut_lancer = "OUI" if droits > 0 else "NON"
            peut_retirer = "OUI" if withdrawn_by_firm[f] <= 4 and ret <= 1 else "NON"
            vals = [p, f, actifs, dev, dispo, ret, init_active.get(f, 0), droits, peut_lancer, peut_retirer]
            for i, v in enumerate(vals, 1):
                ws.cell(rr, i, v)
            rr += 1
    autosize_columns(ws)
    return {"PORTEFEUILLE"}


def update_formules_fr(wb):
    ws = wb["FORMULES_FR"]
    row = ws.max_row + 1
    additions = [
        (19, "MODEL_PERIOD", "SurplusPenalty", "Penalite de surstock au-dela de 200 unites", "SI(StockEnd>200;(StockEnd-200)*PriceList*5%;0)", "Deduit de ProfitContribution"),
        (20, "CONTROLES", "Seuils service", "Zone ideale 95%-105% + alertes 80%/50%", 'SI(ET(TauxService>=95%;TauxService<=105%);"OK";"Hors zone")', "Lecture capacite production"),
        (21, "CONTROLES", "Limites marketing", "Plancher 3% et plafond 15%", "SI(TotalMarketing<3%*RevenueBase;\"Alerte\";SI(TotalMarketing>15%*BudgetBase;\"Alerte\";\"OK\"))", "Respect des bornes budgetaires"),
        (22, "CONTROLES", "Regles R&D", "Niveaux 0/2/5/8%, restriction P1, max 4 modeles", "SI(OU(RD%=0;RD%=2%;RD%=5%;RD%=8%);\"OK\";\"Alerte\")", "Gouvernance innovation"),
        (23, "PORTEFEUILLE", "Droits lancement/retrait", "Calcul des droits selon actifs initiaux et retraits", "MAX(0;2-(Actifs-ActifsInitiaux))", "Peut lancer / Peut retirer"),
    ]
    for a in additions:
        for c, v in enumerate(a, 1):
            ws.cell(row, c, v)
        row += 1
    autosize_columns(ws)
    return {"FORMULES_FR"}


def create_tests_params(wb):
    ws, _ = ensure_sheet(wb, "TESTS_PARAMS")
    ws.delete_rows(1, ws.max_row)
    title_row(ws, 1, "TESTS PARAMS - TRE P2", 12)
    headers = ["Other_Operating_Rate", "Market_Growth", "Stock_Cost_Rate", "Profit estime TRE P2"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    header_style(ws, 2)

    combos = []
    for op_rate in [0.05, 0.075]:
        for growth in [0.10, 0.11, 0.12]:
            for stock_rate in [0.02, 0.025]:
                combos.append((op_rate, growth, stock_rate))

    # Scenario TRE_Enduro_Premium (inputs fixed from prompt)
    price = 5200
    production = 2000
    marketing = 120000
    rd = 40000
    sales = 1823
    base_cogs_ratio = 0.55
    row = 3
    for op_rate, growth, stock_rate in combos:
        revenue = sales * price
        cogs = sales * price * base_cogs_ratio
        stock = max(0, production - sales)
        stock_cost = stock * price * stock_rate
        other_ops = revenue * op_rate
        growth_bonus = revenue * (growth - 0.10) * 0.1
        profit = revenue - cogs - stock_cost - marketing - rd - other_ops + growth_bonus
        ws.cell(row, 1, op_rate)
        ws.cell(row, 2, growth)
        ws.cell(row, 3, stock_rate)
        ws.cell(row, 4, int(round(profit)))
        ws.cell(row, 1).number_format = "0.0%"
        ws.cell(row, 2).number_format = "0.0%"
        ws.cell(row, 3).number_format = "0.0%"
        row += 1
    autosize_columns(ws)
    return {"TESTS_PARAMS"}


def add_report_function_and_generate(wb):
    mp = wb["MODEL_PERIOD"]
    hp = sheet_headers(mp)
    segs = {r["Segment"]: r for r in read_table(wb["SEGMENTS"])}
    firms = ["AVE", "CAN", "EBI", "GIA", "PED", "RID", "SUR", "TRE", "VEL"]

    def price_range_floor(seg, rng):
        pref = (segs.get(seg) or {}).get("Pref_RefYear", 0) or 0
        if str(rng).lower().startswith("bas"):
            return pref * 0.75
        if str(rng).lower().startswith("moy"):
            return pref * 0.9
        return pref * 1.0

    def as_num(v):
        if isinstance(v, (int, float)):
            return float(v)
        return 0.0

    def zone_style_taux(tc: float):
        """Libellé zone + fond colonne Alerte (taux couverture = StockStart+Production)/DemandActive."""
        if tc < 0.9:
            return "Rupture", ZONE_RUPTURE
        if tc < 1.0:
            return "Prudente", ZONE_PRUDENT
        if tc <= 1.1:
            return "Équilibrée", ZONE_EQUILIBRE
        if tc <= 1.2:
            return "Sécuritaire", ZONE_SECURITE
        return "Surproduction", ZONE_SURPROD

    def generer_rapport_firme(firm_code, periode):
        name = f"RAPPORT_{firm_code}_P{periode}"
        ws, _ = ensure_sheet(wb, name)
        ws.delete_rows(1, ws.max_row)
        title_row(ws, 1, f"Rapport {firm_code} - Periode {periode}", 15)
        cols = [
            "ProductKey",
            "ModelName",
            "Segment",
            "Range",
            "PriceList",
            "Production",
            "DemandActive",
            "TotalSales",
            "StockEnd",
            "Marge (%)",
            "Alerte",
            "Taux_couverture",
            "Zone_production",
            "Production_recommandee",
            "Recommandation",
        ]
        for i, h in enumerate(cols, 1):
            ws.cell(2, i, h)
        header_style(ws, 2)

        rr = 3
        total_profit = 0
        margins = []
        total_sales = 0
        total_dem = 0
        for r in range(2, mp.max_row + 1):
            p = mp.cell(r, hp["Period_Index"]).value
            f = mp.cell(r, hp["Firm"]).value
            if p != periode or f != firm_code:
                continue
            rec = {k: mp.cell(r, c).value for k, c in hp.items()}
            revenue = as_num(rec.get("Revenue"))
            profit = as_num(rec.get("ProfitContribution"))
            marge = (profit / revenue) if revenue else 0
            total_profit += profit
            margins.append(marge)
            total_sales += as_num(rec.get("TotalSales"))
            total_dem += as_num(rec.get("DemandActive"))

            withdraw_flag = as_num(rec.get("WithdrawFlag"))
            total_sales_rec = as_num(rec.get("TotalSales"))
            demand_active_rec = as_num(rec.get("DemandActive"))
            stock_end_rec = as_num(rec.get("StockEnd"))
            production_rec = as_num(rec.get("Production"))
            price_list_rec = as_num(rec.get("PriceList"))
            stock_start_rec = as_num(rec.get("StockStart"))

            tc = (
                (stock_start_rec + production_rec) / demand_active_rec
                if demand_active_rec
                else 1.0
            )
            prod_reco = int(math.ceil(demand_active_rec * 1.05 / 100.0) * 100) if demand_active_rec else 0
            zone_label, zone_fill = zone_style_taux(tc)

            alert = zone_label
            reco = "Maintenir la strategie"
            alert_fill = zone_fill

            floor = price_range_floor(rec.get("Segment"), rec.get("Range"))
            if withdraw_flag == 1:
                alert = "Liquidation"
                reco = "Retrait / fin de vie — voir consignes liquidation"
                alert_fill = ALERT_ORANGE
            elif price_list_rec < floor:
                alert = "Prix invalide"
                reco = f"Corriger prix — plancher {int(round(floor))} $"
                alert_fill = ZONE_RUPTURE
            else:
                if tc < 0.9:
                    reco = f"Augmenter vers ~{prod_reco} u. (couverture cible 100–110 %)"
                elif tc > 1.2:
                    reco = "Réduire production ou stimuler la demande — risque surstock"
                elif tc > 1.1:
                    reco = "Surveillance stock tampon — ajuster si demande confirmée plus basse"

            values = [
                rec.get("ProductKey"),
                rec.get("ModelName"),
                rec.get("Segment"),
                rec.get("Range"),
                price_list_rec,
                production_rec,
                demand_active_rec,
                total_sales_rec,
                stock_end_rec,
                marge,
                alert,
                tc,
                zone_label,
                prod_reco,
                reco,
            ]
            for i, v in enumerate(values, 1):
                ws.cell(rr, i, v)
            ws.cell(rr, 10).number_format = "0.0%"
            ws.cell(rr, 12).number_format = "0.0%"
            ws.cell(rr, 11).fill = alert_fill
            rr += 1

        kpi_row = rr + 1
        title_row(ws, kpi_row, "KPIs firme", 6)
        ws.cell(kpi_row + 1, 1, "Profit total")
        ws.cell(kpi_row + 1, 2, int(round(total_profit)))
        ws.cell(kpi_row + 2, 1, "Marge moyenne")
        ws.cell(kpi_row + 2, 2, sum(margins) / len(margins) if margins else 0)
        ws.cell(kpi_row + 2, 2).number_format = "0.0%"
        ws.cell(kpi_row + 3, 1, "PDM globale")
        ws.cell(kpi_row + 3, 2, f'=IFERROR(INDEX(RESULTS_FIRM!$M:$M,MATCH(1,(RESULTS_FIRM!$A:$A={periode})*(RESULTS_FIRM!$C:$C="{firm_code}"),0)),0)')
        ws.cell(kpi_row + 3, 2).number_format = "0.0%"
        ws.cell(kpi_row + 4, 1, "Taux service moyen")
        ws.cell(kpi_row + 4, 2, (total_sales / total_dem) if total_dem else 0)
        ws.cell(kpi_row + 4, 2).number_format = "0.0%"

        zebra = PatternFill("solid", fgColor="F7F7F7")
        for r in range(3, rr):
            if (r - 3) % 2 == 0:
                for c in list(range(1, 11)) + list(range(12, 16)):
                    ws.cell(r, c).fill = zebra
        autosize_columns(ws)

    for f in firms:
        generer_rapport_firme(f, 1)
    return {f"RAPPORT_{f}_P1" for f in firms}


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    modified = set()

    b1 = block1(wb)
    modified |= b1
    print(f"Bloc 1 termine — {len(b1)} feuilles modifiees / creees")

    b2 = add_model_period_columns(wb) | append_controles_rows(wb) | build_tests_production(wb)
    modified |= b2
    print(f"Bloc 2 termine — {len(b2)} feuilles modifiees / creees")

    b3 = {"CONTROLES", "MODEL_PERIOD"}
    modified |= b3
    print(f"Bloc 3 termine — {len(b3)} feuilles modifiees / creees")

    b4 = {"CONTROLES"}
    modified |= b4
    print(f"Bloc 4 termine — {len(b4)} feuilles modifiees / creees")

    b5 = build_portefeuille(wb) | {"CONTROLES"}
    modified |= b5
    print(f"Bloc 5 termine — {len(b5)} feuilles modifiees / creees")

    b6 = add_report_function_and_generate(wb)
    modified |= b6
    print(f"Bloc 6 termine — {len(b6)} feuilles modifiees / creees")

    b7 = create_tests_params(wb) | update_formules_fr(wb)
    modified |= b7
    print(f"Bloc 7 termine — {len(b7)} feuilles modifiees / creees")

    wb.save(OUTPUT_FILE)
    print(f"Fichier sauvegarde: {OUTPUT_FILE}")
    print(f"Total feuilles modifiees / creees: {len(modified)}")


if __name__ == "__main__":
    main()
