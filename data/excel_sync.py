from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, Iterable, Optional


@dataclass(frozen=True)
class ExcelSyncReport:
    workbook_path: str
    sheet_used: str
    extracted_params: Dict[str, float]
    warnings: list[str]


def _iter_param_rows(ws) -> Iterable[tuple[Optional[str], Optional[float]]]:
    """
    Heuristique: feuille PARAM où la colonne A contient un identifiant texte
    et la colonne B la valeur numérique (ou équivalent).
    """
    for r in range(1, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k is None and v is None:
            continue
        key = str(k).strip() if k is not None else None
        val = _to_float(v)
        yield key, val


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(" ", "").replace(",", ".")
        return float(s)
    except Exception:
        return None


def extract_params_from_workbook(xlsx_path: str, *, sheet_name: str = "PARAM") -> ExcelSyncReport:
    """
    Extraction minimale des paramètres d'un chiffrier Excel.

    Le bot n'importe pas automatiquement le classeur: cette fonction sert de garde-fou
    lorsque vous fournissez un .xlsx et souhaitez comparer rapidement les paramètres clés.
    """
    warnings: list[str] = []
    extracted: Dict[str, float] = {}

    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl est requis pour lire un .xlsx") from e

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        # fallback: case-insensitive match
        match = next((n for n in wb.sheetnames if n.strip().lower() == sheet_name.lower()), None)
        if match is None:
            raise ValueError(f"Feuille '{sheet_name}' introuvable. Feuilles: {wb.sheetnames}")
        sheet_name = match

    ws = wb[sheet_name]
    for key, val in _iter_param_rows(ws):
        if not key:
            continue
        if val is None:
            continue
        extracted[key] = val

    if not extracted:
        warnings.append("Aucun paramètre numérique détecté (format de feuille inattendu).")

    return ExcelSyncReport(
        workbook_path=xlsx_path,
        sheet_used=sheet_name,
        extracted_params=extracted,
        warnings=warnings,
    )


def compare_excel_params_to_market_config(
    extracted_params: Dict[str, float],
    market_config: Dict[str, Any],
) -> dict:
    """
    Compare quelques paramètres attendus si présents.

    La correspondance des noms dépend du chiffrier ; on supporte des variantes fréquentes.
    """
    aliases = {
        "base_market_size": ["Base_Market_Units", "BASE_MARKET_UNITS", "MarketUnits_RefYear"],
        "growth_rate": ["Market_Growth", "MARKET_GROWTH", "GrowthRate"],
        "price_inflation_rate": ["Price_Inflation", "PRICE_INFLATION", "InflationRate"],
        "promo_standard_max": ["Max_Promo", "MAX_PROMO"],
        "promo_liquidation_max": ["Max_Liquidation_Promo", "MAX_LIQUIDATION_PROMO"],
    }

    def get_excel_value(keys: list[str]) -> Optional[float]:
        for k in keys:
            if k in extracted_params:
                return extracted_params[k]
        # fallback: case-insensitive
        lower_map = {k.lower(): v for k, v in extracted_params.items()}
        for k in keys:
            if k.lower() in lower_map:
                return lower_map[k.lower()]
        return None

    rows = []
    for cfg_key, excel_keys in aliases.items():
        excel_val = get_excel_value(excel_keys)
        cfg_val = None
        if cfg_key in market_config:
            cfg_val = float(market_config[cfg_key])
        else:
            # constraints subtree
            if cfg_key in market_config.get("constraints", {}):
                cfg_val = float(market_config["constraints"][cfg_key])
        if excel_val is None or cfg_val is None:
            continue
        delta = excel_val - cfg_val
        rows.append(
            {
                "key": cfg_key,
                "excel": excel_val,
                "config": cfg_val,
                "delta": delta,
                "delta_pct": (delta / cfg_val) if cfg_val else None,
            }
        )

    return {
        "compared": rows,
        "notes": [
            "Comparaison indicative: la nomenclature Excel varie selon les versions.",
            "Pour une synchronisation complète (prix par modèle/période, FIRMS, etc.), une logique dédiée est requise.",
        ],
    }

